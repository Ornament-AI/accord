"""Contract tests for the PII-free canonical workbook extractor."""

from __future__ import annotations

import json
import importlib.util
import re
from pathlib import Path

import pytest
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[3]
CONTRACT_PATH = ROOT / "fixtures" / "sanitized" / "june-2026" / "canonical_export_contract.json"
SCRIPT_PATH = ROOT / "scripts" / "extract_canonical_export_contract.py"
SPEC = importlib.util.spec_from_file_location("extract_canonical_export_contract", SCRIPT_PATH)
assert SPEC is not None and SPEC.loader is not None
EXTRACTOR = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(EXTRACTOR)
CANONICAL_SHEET_NAMES = EXTRACTOR.CANONICAL_SHEET_NAMES
PAY_BILL_COLUMNS = EXTRACTOR.PAY_BILL_COLUMNS
extract_contract = EXTRACTOR.extract_contract


def test_checked_in_contract_is_pii_free_and_complete() -> None:
    contract = json.loads(CONTRACT_PATH.read_text(encoding="utf-8"))

    assert contract["contract_version"] == 1
    assert tuple(sheet["name"] for sheet in contract["sheets"]) == CANONICAL_SHEET_NAMES
    assert tuple(contract["pay_bill"]["columns"]) == PAY_BILL_COLUMNS
    assert contract["pay_bill"]["normalized_june_2026_totals"] == {
        "employee_count": 28,
        "salary_earnings": "5073200.00",
        "employer_share": "29785.00",
        "gross_bill": "5102985.00",
        "total_deductions": "1264890.00",
        "net_payable": "3838095.00",
    }
    pay_bill = next(sheet for sheet in contract["sheets"] if sheet["name"] == "Pay Bill")
    assert pay_bill["print_area"] == "'Pay Bill'!$A$1:$AB$208"
    assert pay_bill["print_titles"] == "'Pay Bill'!$2:$7"
    assert pay_bill["page_setup"]["orientation"] == "landscape"
    assert pay_bill["page_setup"]["paperSize"] == "9"
    assert pay_bill["manual_row_breaks"] == [67, 134]

    serialized = json.dumps(contract, ensure_ascii=False)
    assert re.search(r"\b(?:Shri|Smt)\.?\b", serialized, re.IGNORECASE) is None
    assert re.search(r"\b[A-Z]{5}[0-9]{4}[A-Z]\b", serialized) is None
    EXTRACTOR._assert_no_pii(contract)


def test_extractor_records_structure_without_cell_values(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in CANONICAL_SHEET_NAMES:
        sheet = workbook.create_sheet(sheet_name)
        sheet["A1"] = "safe structural label"
        sheet["B2"] = "=1+1"
    for hidden_name in ("Motor car Ad", "Motor cycale Ad (2)", "Festival"):
        workbook[hidden_name].sheet_state = "hidden"
    pay_bill = workbook["Pay Bill"]
    pay_bill.merge_cells("O2:Z2")
    pay_bill.print_area = "A1:AB208"
    pay_bill.print_title_rows = "2:7"
    pay_bill.page_setup.orientation = "landscape"
    pay_bill.page_setup.paperSize = "9"
    pay_bill.column_dimensions["A"].width = 5.5
    pay_bill.row_dimensions[3].height = 40

    source = tmp_path / "source.xlsx"
    workbook.save(source)
    contract = extract_contract(source)

    assert tuple(sheet["name"] for sheet in contract["sheets"]) == CANONICAL_SHEET_NAMES
    extracted_pay_bill = next(sheet for sheet in contract["sheets"] if sheet["name"] == "Pay Bill")
    assert extracted_pay_bill["formula_count"] == 1
    assert extracted_pay_bill["merged_cells"] == ["O2:Z2"]
    assert "safe structural label" not in json.dumps(contract)


def test_pii_guard_rejects_employee_values() -> None:
    with pytest.raises(ValueError, match="PII-like value"):
        EXTRACTOR._assert_no_pii({"employee": "Shri Example Person"})
