"""Safety tests for the clean-room canonical structural template builder."""

from __future__ import annotations

from io import BytesIO
import importlib.util
import os
from pathlib import Path

from openpyxl import Workbook
import pytest


ROOT = Path(__file__).resolve().parents[3]
SCRIPT_PATH = ROOT / "scripts" / "build_canonical_schedule_template.py"
TEMPLATE_PATH = (
    ROOT / "backend" / "app" / "reports" / "templates" / "canonical_schedule_structure.xlsx"
)
SPEC = importlib.util.spec_from_file_location("build_canonical_schedule_template", SCRIPT_PATH)
assert SPEC is not None and SPEC.loader is not None
BUILDER = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(BUILDER)


def test_checked_in_template_passes_fail_closed_ooxml_validator() -> None:
    BUILDER.validate_template(TEMPLATE_PATH.read_bytes())


def test_ooxml_validator_rejects_any_cell_payload() -> None:
    workbook = Workbook()
    workbook.active["A1"] = "source value"
    buffer = BytesIO()
    workbook.save(buffer)
    with pytest.raises(ValueError, match="Worksheet payload"):
        BUILDER.validate_template(buffer.getvalue())


def test_build_is_byte_deterministic_for_same_source(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in BUILDER.CANONICAL_SHEETS:
        sheet = workbook.create_sheet(name)
        sheet.merge_cells("A1:B1")
        sheet.column_dimensions["A"].width = 12.5
        sheet.row_dimensions[1].height = 22
    source = tmp_path / "synthetic-source.xlsx"
    workbook.save(source)

    first = BUILDER.build_template(source)
    second = BUILDER.build_template(source)
    assert first == second


def test_local_canonical_source_rebuild_matches_checked_in_template() -> None:
    source = os.environ.get("ACCORD_CANONICAL_SOURCE_XLSX")
    if not source:
        pytest.skip("set ACCORD_CANONICAL_SOURCE_XLSX for byte-for-byte source proof")
    assert BUILDER.build_template(Path(source)) == TEMPLATE_PATH.read_bytes()
