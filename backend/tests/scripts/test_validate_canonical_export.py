"""Tests for the independent canonical export acceptance validator."""

from __future__ import annotations

import importlib.util
from pathlib import Path
import shutil

from openpyxl import Workbook, load_workbook
import pytest


ROOT = Path(__file__).resolve().parents[3]
SCRIPT_PATH = ROOT / "scripts" / "validate_canonical_export.py"
CONTRACT_PATH = ROOT / "fixtures" / "sanitized" / "june-2026" / "canonical_export_contract.json"
TEMPLATE_PATH = (
    ROOT / "backend" / "app" / "reports" / "templates" / "canonical_schedule_structure.xlsx"
)
SPEC = importlib.util.spec_from_file_location("validate_canonical_export", SCRIPT_PATH)
assert SPEC is not None and SPEC.loader is not None
VALIDATOR = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(VALIDATOR)


def test_raw_template_matches_contract_except_normalized_payslip() -> None:
    issues = VALIDATOR.validate_structure(TEMPLATE_PATH, VALIDATOR.load_contract(CONTRACT_PATH))
    assert issues
    assert all(issue.location.startswith("sheet['PaySlip']") for issue in issues)


def test_normalized_payslip_structure_accepts_all_28_repeated_blocks() -> None:
    semantics = VALIDATOR.load_contract(CONTRACT_PATH)["canonical_semantics"]["front_sheets"][
        "PaySlip"
    ]
    structure = semantics["normalized_structure"]
    starts = semantics["block_starts"]
    rows = []
    for start in starts:
        rows.append({"row": start, "height": structure["start_row_height"]})
        rows.extend(
            {"row": row, "height": structure["default_row_height"]}
            for row in range(start + 1, start + 13)
        )
        rows.append({"row": start + 13, "height": structure["total_row_height"]})
        rows.append({"row": start + 14, "height": structure["net_row_height"]})
    actual = {
        "used_range": structure["used_range"],
        "print_area": structure["print_area"],
        "column_dimensions": [
            {"min": index, "max": index, "width": width}
            for index, width in enumerate(structure["column_widths"], start=1)
        ],
        "row_dimensions": rows,
        "merged_cells": sorted(VALIDATOR._normalized_payslip_merges(starts)),
        "manual_row_breaks": [starts[index - 1] + 17 for index in range(2, len(starts), 2)],
        "page_setup": structure["page_setup"],
    }
    issues = []
    VALIDATOR._validate_normalized_payslip_structure(issues, actual, semantics)
    assert issues == []


def test_disbursement_and_nps_contract_are_distinct_from_pay_bill_net() -> None:
    semantics = VALIDATOR.load_contract(CONTRACT_PATH)["canonical_semantics"]
    front = semantics["front_sheets"]
    assert front["office tip"]["totals"][0]["expected"] == "3838095.00"
    assert front["Bank Tip"]["totals"][0]["expected"] == "3991038.00"
    assert front["PaySlip"]["expected_total"] == "3991038.00"

    reconciliations = {item["name"]: item for item in semantics["reconciliations"]}
    assert "Bank Tip!G42" not in reconciliations["net payable"]["members"]
    assert reconciliations["bank disbursement"]["members"] == [
        "Pay Bill!AA208",
        "Pension Sub (2)!I41",
    ]
    nps_targets = reconciliations["NPS contribution"]["target_members"]
    assert len(nps_targets) == 20
    assert "Pay Bill!Q208" not in nps_targets
    assert "Pay Bill!R208" not in nps_targets
    assert reconciliations["NPS contribution"]["expected"] == "262188.00"


def test_structure_reports_visibility_width_merge_and_break_differences(tmp_path: Path) -> None:
    workbook = load_workbook(TEMPLATE_PATH)
    sheet = workbook["Pay Bill"]
    sheet.sheet_state = "hidden"
    sheet.column_dimensions["A"].width = 99
    sheet.unmerge_cells("O2:Z2")
    sheet.row_breaks.brk = []
    broken = tmp_path / "broken.xlsx"
    workbook.save(broken)

    issues = VALIDATOR.validate_structure(broken, VALIDATOR.load_contract(CONTRACT_PATH))
    codes = {issue.code for issue in issues}
    assert "sheet.state" in codes
    assert "sheet.column_dimensions" in codes
    assert "sheet.merged_cells" in codes
    assert "sheet.manual_row_breaks" in codes


def _write_formula_role_row(sheet, row: int) -> None:
    for column in (*range(3, 11), 12, 13, *range(16, 26)):
        letter = sheet.cell(row, column).column_letter
        sheet.cell(row, column, f"=SUM({letter}{row - 4}:{letter}{row - 1})")
    sheet.cell(row, 11, f"=SUM(C{row}:J{row})")
    sheet.cell(row, 14, f"=K{row}+L{row}-M{row}")
    sheet.cell(row, 26, f"=SUM(P{row}:Y{row})")
    sheet.cell(row, 27, f"=N{row}-Z{row}")


def test_pay_bill_semantics_catches_ordinal_and_formula_role_errors() -> None:
    workbook = load_workbook(TEMPLATE_PATH)
    sheet = workbook["Pay Bill"]
    for column in range(1, 29):
        sheet.cell(8, column, column)
    sheet["A10"] = 1
    sheet["B15"] = "Total Rs."
    _write_formula_role_row(sheet, 15)
    sheet["K15"] = "=SUM(C15:I15)"
    sheet["AB8"] = 29

    issues = VALIDATOR.validate_pay_bill_semantics(sheet, VALIDATOR.load_contract(CONTRACT_PATH))
    codes = {issue.code for issue in issues}
    assert "pay_bill.ordinal_row" in codes
    assert "pay_bill.formula_role" in codes


def _semantic_contract(*, front=None, schedules=None, reconciliations=None):
    return {
        "canonical_semantics": {
            "front_sheets": front or {},
            "schedules": schedules or {},
            "reconciliations": reconciliations or [],
        }
    }


def test_contract_covers_every_non_pay_bill_sheet_without_pii_values() -> None:
    contract = VALIDATOR.load_contract(CONTRACT_PATH)
    semantics = contract["canonical_semantics"]
    covered = set(semantics["front_sheets"]) | set(semantics["schedules"])
    expected = {sheet["name"] for sheet in contract["sheets"]} - {"Pay Bill"}

    assert covered == expected
    assert len(covered) == 17


def test_front_sheet_body_and_total_formula_are_semantically_required() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "office tip"
    sheet["E36"] = "=SUM(E8:E34)"
    contract = _semantic_contract(
        front={
            "office tip": {
                "body": {
                    "rows": [8, 35],
                    "serial_column": "B",
                    "required_columns": ["C", "D"],
                    "expected_count": 28,
                },
                "totals": [{"cell": "E36", "sum_range": "E8:E35", "expected": "3838095.00"}],
            }
        }
    )

    issues = VALIDATOR.validate_non_pay_bill_semantics(
        workbook, contract, require_june_totals=False
    )
    codes = {issue.code for issue in issues}
    assert "sheet.body_empty" in codes
    assert "sheet.total_formula" in codes


def test_gpf_destination_row_buckets_and_total_are_checked() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "GPF-Nagpur"
    sheet["A36"] = 1
    sheet["B36"] = "account"
    sheet["C36"] = "employee"
    sheet["D36"] = "post"
    sheet["I36"] = "=F36"
    sheet["I41"] = "=SUM(I36:I40)"
    sheet["B20"] = "Accountant General, Mumbai"
    sheet["A23"] = "office"
    sheet["A27"] = "officer"
    contract = _semantic_contract(
        schedules={
            "GPF-Nagpur": {
                "body": {
                    "rows": [36, 40],
                    "serial_column": "A",
                    "required_columns": ["B", "C", "D"],
                    "expected_count": 1,
                },
                "destination": {
                    "cell": "B20",
                    "jurisdiction": "nagpur",
                    "required_cells": ["B20", "A23", "A27"],
                },
                "row_formula": {"target_column": "I", "source_columns": ["F", "G"]},
                "totals": [{"cell": "I41", "sum_range": "I36:I40", "expected": "1"}],
            }
        }
    )

    issues = VALIDATOR.validate_non_pay_bill_semantics(
        workbook, contract, require_june_totals=False
    )
    codes = {issue.code for issue in issues}
    assert "gpf.jurisdiction" in codes
    assert "sheet.row_formula" in codes


def test_nps_account_head_allocation_roles_are_checked() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pension Sub (2)"
    for offset, coordinate in enumerate(("C11", "C12", "C13", "D11", "F11", "G11", "H11", "I11")):
        sheet[coordinate] = offset + 1
    sheet["H41"] = "=SUM(H11:H40)"
    sheet["I41"] = "=SUM(I11:I40)"
    sheet["I44"] = "=H41"
    sheet["I45"] = "=I41"
    sheet["I46"] = "=SUM(I44:I45)"
    sheet["C47"] = "=I46"
    sheet["B44"] = "employee account head"
    sheet["B45"] = "employer account head"
    contract = _semantic_contract(
        schedules={
            "Pension Sub (2)": {
                "block_starts": [11],
                "formula_roles": {
                    "H41": "=SUM(H11:H40)",
                    "I41": "=SUM(I11:I40)",
                    "I44": "=H41",
                    "I45": "=I41",
                    "I46": "=SUM(I42:I45)",
                    "C47": "=I46",
                },
                "required_cells": ["B42", "B43", "B44", "B45"],
                "totals": [{"cell": "I46", "expected": "1"}],
            }
        }
    )

    issues = VALIDATOR.validate_non_pay_bill_semantics(
        workbook, contract, require_june_totals=False
    )
    codes = {issue.code for issue in issues}
    assert "sheet.formula_role" in codes
    assert "sheet.required_value" in codes


def test_cross_sheet_schedule_reconciliation_uses_cached_values() -> None:
    formulas = Workbook()
    values = Workbook()
    values.active.title = "WORLI"
    values["WORLI"]["H8"] = 1250
    mumbai = values.create_sheet("Mumbai")
    mumbai["J11"] = 10419
    pay_bill = values.create_sheet("Pay Bill")
    pay_bill["W208"] = 10000
    contract = _semantic_contract(
        reconciliations=[
            {
                "name": "accommodation recovery",
                "members": ["WORLI!H8", "Mumbai!J11"],
                "target": "Pay Bill!W208",
                "expected": "11669.00",
                "operation": "sum",
            }
        ]
    )

    issues = VALIDATOR.validate_non_pay_bill_semantics(formulas, contract, value_workbook=values)
    assert {issue.code for issue in issues} == {"workbook.reconciliation"}


@pytest.mark.skipif(shutil.which("soffice") is None, reason="LibreOffice is unavailable")
def test_libreoffice_recalculation_exposes_formula_errors(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active["A1"] = "=1/0"
    source = tmp_path / "error.xlsx"
    workbook.save(source)

    recalculated = VALIDATOR.recalculate_with_libreoffice(source)
    try:
        assert VALIDATOR._formula_error_cells(recalculated) == ["Sheet!A1"]
    finally:
        shutil.rmtree(recalculated.parents[1], ignore_errors=True)
