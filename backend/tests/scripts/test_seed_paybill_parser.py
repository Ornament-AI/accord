"""Source-fidelity regressions for the canonical June pay-bill importer."""

from __future__ import annotations

import importlib.util
import os
from types import SimpleNamespace
import sys
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import Workbook

SCRIPT = Path(__file__).resolve().parents[3] / "scripts" / "seed_paybill_xlsx.py"
SPEC = importlib.util.spec_from_file_location("seed_paybill_xlsx", SCRIPT)
assert SPEC is not None and SPEC.loader is not None
seed_paybill_xlsx = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = seed_paybill_xlsx
SPEC.loader.exec_module(seed_paybill_xlsx)


def test_seed_component_specs_cover_canonical_register_columns() -> None:
    from app.domain.payroll.export_metadata import register_column_matches_classification

    expected = {
        "BASIC": "basic_pay",
        "DA": "dearness_allowance",
        "CLA": "city_compensatory_allowance",
        "HRA": "house_rent_allowance",
        "WASH_ALLOWANCE": "wash_child_other_charges",
        "OTHER_ALLOWANCE": "other_reimbursement_salary_increment_difference",
        "ADDITIONAL_ALLOWANCE": "additional_conveyance_transport_allowance",
        "TRANSPORT": "transport_pta_honorarium",
        "GPF_SUBSCRIPTION": "gpf_subscription_refund_arrears",
        "NPS_EMPLOYER_TRANSFER": "pension_employer_share",
        "EPF_EMPLOYER_TRANSFER": "pension_employer_share",
        "NPS_EMPLOYEE": "pension_employee_share",
        "EPF_EMPLOYEE": "pension_employee_share",
        "EPF_EMPLOYER": "employer_share",
        "INCOME_TAX": "income_tax",
        "PROFESSIONAL_TAX": "professional_tax",
        "GIS": "insurance",
        "HBA_INSTALLMENT": "advances",
        "ACCOMMODATION_LICENSE_FEE": "house_rent_service_charge_arrears",
    }

    specs = {
        code: (classification, register_column)
        for code, _name, classification, register_column in seed_paybill_xlsx.COMPONENT_SPECS
    }
    assert {
        code: register_column for code, (_classification, register_column) in specs.items()
    } == expected
    assert all(
        register_column_matches_classification(classification, register_column)
        for classification, register_column in specs.values()
    )


def test_seed_component_transfer_metadata_preserves_disbursement_contract() -> None:
    specs = {
        code: (name, classification, register_column)
        for code, name, classification, register_column in seed_paybill_xlsx.COMPONENT_SPECS
    }

    def payload(code: str) -> dict[str, object]:
        name, classification, register_column = specs[code]
        return seed_paybill_xlsx.component_create_payload(
            code=code,
            name=name,
            classification=classification,
            register_column=register_column,
            display_order=1,
        )

    assert payload("NPS_EMPLOYER_TRANSFER")["employer_transfer"] is True
    assert "transfer_of" not in payload("NPS_EMPLOYER_TRANSFER")
    assert payload("EPF_EMPLOYER_TRANSFER") == {
        "code": "EPF_EMPLOYER_TRANSFER",
        "name": "EPF Employer Transfer",
        "classification": "ag_deduction",
        "register_column": "pension_employer_share",
        "display_order": 1,
        "employer_transfer": True,
        "transfer_of": "EPF_EMPLOYER",
    }


def test_wipe_sends_sql_on_psql_stdin_for_variable_substitution(monkeypatch) -> None:
    captured: dict[str, object] = {}

    def fake_run(command, **kwargs):
        captured["command"] = command
        captured.update(kwargs)
        return SimpleNamespace(returncode=0, stderr="", stdout="")

    monkeypatch.setattr("subprocess.run", fake_run)
    organization_id = str(uuid4())

    seed_paybill_xlsx.wipe_org_master_data(
        organization_id,
        dsn_env={"PGDATABASE": "accord_acceptance"},
    )

    command = captured["command"]
    assert isinstance(command, list)
    assert "-c" not in command
    assert command[-2:] == ["-d", "accord_acceptance"]
    assert f"org_id={organization_id}" in command
    assert "WHERE organization_id = :'org_id'::uuid" in str(captured["input"])


def _synthetic_workbook(path: Path, *, invalid_other_total: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pay Bill"
    ws.cell(9, 1, 25)
    ws.cell(9, 2, "Post of Assistant Accounts Officer")
    ws.cell(10, 2, "Shri Source Faithful")
    ws.cell(10, 3, "t ")  # Canonical employee-25 defect.
    ws.cell(11, 2, "Accounts Officer")
    ws.cell(12, 2, "Basic @ Rs.66000/-")
    ws.cell(13, 2, "ABCDE1234F")
    ws.cell(14, 2, "Total Rs.")
    for column, value in {
        3: 66000,
        4: 38280,
        5: 240,
        6: 19800,
        7: 0,
        8: "not-money" if invalid_other_total else 2300,
        9: 14600,
        10: 2700,
        11: 143920,
        16: 0,
        17: 14600,
        18: 10428,
        19: 0,
        21: 9000,
        24: 200,
    }.items():
        ws.cell(14, column, value)

    for title in (
        "GPF-Mumbai",
        "GPF-Nagpur",
        "GIS",
        "Income Tax",
        "HBA Ad",
        "Pension Sub (2)",
        "Bank Tip",
        "WORLI",
        "Mumbai",
    ):
        wb.create_sheet(title)

    ws = wb["GIS"]
    ws.cell(7, 2, 1)
    ws.cell(7, 3, "Shri Source Faithful")
    ws.cell(7, 5, 480)

    ws = wb["Income Tax"]
    ws.cell(5, 1, 1)
    ws.cell(5, 2, "Shri Source Faithful")
    ws.cell(5, 3, "Accounts Officer")
    ws.cell(5, 4, "ABCDE1234F")
    ws.cell(5, 6, 9000)

    ws = wb["Pension Sub (2)"]
    ws.cell(11, 2, 1)
    ws.cell(11, 3, "PENSION-ACCOUNT-25")
    ws.cell(11, 4, "Shri Source Faithful")
    ws.cell(11, 8, 10428)
    ws.cell(11, 9, 14600)
    ws.cell(12, 3, "SEVARTH-25")
    ws.cell(13, 3, "000000000025")

    ws = wb["Bank Tip"]
    ws.cell(14, 2, 1)
    ws.cell(14, 3, "Shri Source Faithful")
    ws.cell(14, 4, "State Bank of India, Source Branch")
    ws.cell(14, 5, "0000000012345")
    ws.cell(14, 6, "SBIN0000123")
    ws.cell(14, 7, 109212)

    wb.save(path)


def test_parser_recovers_bad_basic_preserves_column_i_and_real_bank(tmp_path: Path) -> None:
    path = tmp_path / "canonical-shape.xlsx"
    _synthetic_workbook(path)

    employees = seed_paybill_xlsx.parse_paybill(path)

    assert len(employees) == 1
    employee = employees[0]
    assert employee.sr == 25
    assert employee.basic == 66000
    assert employee.additional_allowance == 14600
    assert employee.gross == 143920
    assert employee.bank_account == "0000000012345"
    assert employee.ifsc == "SBIN0000123"
    assert employee.bank_name == "State Bank of India, Source Branch"
    assert employee.bank_branch is None
    assert employee.regime == "nps"
    assert employee.pension_account == "PENSION-ACCOUNT-25"
    assert employee.pay_bill_group is not None
    assert employee.pay_bill_group.order == 1
    assert employee.pay_bill_group.heading == "Assistant Accounts Officer"
    assert employee.pay_bill_group.owner_designation == "Accounts Officer"

    summary = seed_paybill_xlsx.reconciliation_summary(employees)
    assert summary["headcount"] == 1
    assert summary["salary_earnings"] == 143920
    assert summary["additional_allowance"] == 14600


def test_employee_payload_does_not_fabricate_unknown_profile_or_bank_values(
    tmp_path: Path,
) -> None:
    path = tmp_path / "canonical-shape.xlsx"
    _synthetic_workbook(path)
    employee = seed_paybill_xlsx.parse_paybill(path)[0]

    pay_bill_post_id = uuid4()
    payload = seed_paybill_xlsx.employee_create_payload(
        employee,
        office_id=uuid4(),
        post_id=uuid4(),
        pay_bill_post_id=pay_bill_post_id,
    )

    assert payload["profile"]["date_of_birth"] is None
    assert payload["profile"]["date_of_joining"] is None
    assert payload["profile"]["epf_number"] is None
    assert payload["profile"]["gpf_account_number"] is None
    assert payload["profile"]["sevarth_id"] == "SEVARTH-25"
    assert payload["profile"]["pran"] == "000000000025"
    assert payload["profile"]["pension_account"] == "PENSION-ACCOUNT-25"
    assert payload["pay"] == {"pay_matrix_level": None, "basic_pay": "66000.00"}
    assert payload["posting"]["pay_bill_post_id"] == str(pay_bill_post_id)
    assert payload["bank"] == {
        "account_number": "0000000012345",
        "ifsc": "SBIN0000123",
        "bank_name": "State Bank of India, Source Branch",
        "branch": None,
        "is_primary_salary": True,
    }


def test_nonblank_invalid_money_is_not_silently_coerced_to_zero(tmp_path: Path) -> None:
    path = tmp_path / "invalid-money.xlsx"
    _synthetic_workbook(path, invalid_other_total=True)

    with pytest.raises(seed_paybill_xlsx.SeedError, match="invalid nonblank money"):
        seed_paybill_xlsx.parse_paybill(path)


@pytest.mark.skipif(
    not os.environ.get("ACCORD_CANONICAL_PAYBILL_XLSX"),
    reason="set ACCORD_CANONICAL_PAYBILL_XLSX to run the PII-bearing local acceptance test",
)
def test_real_canonical_workbook_normalized_reconciliation() -> None:
    path = Path(os.environ["ACCORD_CANONICAL_PAYBILL_XLSX"])
    employees = seed_paybill_xlsx.parse_paybill(path)
    summary = seed_paybill_xlsx.reconciliation_summary(employees)

    assert summary == {
        "headcount": 28,
        "salary_earnings": 5073200,
        "employer_share": 29785,
        "gross_bill": 5102985,
        "total_deductions": 1264890,
        "net_payable": 3838095,
        "offbill_employer_remittance": 152943,
        "employee_disbursement": 3991038,
        "additional_allowance": 157943,
        "gpf_total": 280000,
        "pension_employer": 182728,
        "pension_employee": 139030,
        "income_tax": 550700,
        "gis": 22440,
        "professional_tax": 5600,
    }
    assert [employee.sr for employee in employees] == list(range(1, 29))
    assert all(employee.bank_account for employee in employees)
    assert [
        employees[index - 1].pay_bill_group.order
        for index in (1, 2, 3, 16, 25, 26)
        if employees[index - 1].pay_bill_group is not None
    ] == [1, 2, 3, 4, 5, 6]
