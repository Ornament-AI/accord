"""Golden tests for Income Tax, Professional Tax, and GIS schedule builders."""

from __future__ import annotations

import re
from collections import Counter
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from uuid import UUID, uuid4

import pytest
import sqlalchemy as sa
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy.ext.asyncio import AsyncSession

from app.exceptions import ConflictError
from app.models.accommodation import AccommodationAssignment, accommodation_charge_versions
from app.models.advances import AdvanceAccount, advance_installment_versions
from app.models.employees import (
    Employee,
    employee_pay_versions,
    employee_posting_versions,
    employee_profile_versions,
)
from app.models.org_structure import Office, Post
from app.models.pay_components import PayComponent, component_rate_versions
from app.models.payroll_runs import (
    PayrollPeriod,
    PayrollRun,
    payroll_employee_results,
    payroll_result_lines,
)
from app.models.platform import PayrollApproval
from app.models.recurring_instructions import (
    RecurringInstruction,
    recurring_instruction_versions,
)
from app.reports.base import ReportContext, ReportRegistry
from app.reports.excel import MONEY_FORMAT
from app.reports.families.statutory import (
    REPORT_TYPE_GIS,
    REPORT_TYPE_INCOME_TAX,
    REPORT_TYPE_PROFESSIONAL_TAX,
    gis_builder,
    income_tax_builder,
    professional_tax_builder,
    register,
    statutory_to_excel,
    statutory_to_pdf,
)
from app.reports.formatting import format_inr
from app.services import versioning
from app.services.run_calculation import calculate_run_command
from app.services.run_posting import post_run
from app.tenancy import bind_tenant_context
from tests.roster_helpers import initialize_run_roster
from tests.e2e.fixture_loader import (
    ACCOMMODATION_COMPONENT_CODES,
    ADVANCE_COMPONENT_CODES,
    BASIC_CODE,
    RECURRING_COMPONENT_CODES,
    EmployeeSeed,
    line_amount,
    load_june_fixture,
    map_quarters_location,
    map_regime,
    money,
)
from tests.identity_helpers import seed_organization, seed_user

EFFECTIVE_FROM = date(2026, 1, 1)
PERIOD_YEAR = 2026
PERIOD_MONTH = 6
TEMPLATE_VERSION = "v1"

_PAN_RE = re.compile(r"^ZZZPZ\d{4}Z$")

_CLASSIFICATION_DB = {
    "earning": "earning",
    "employer_contribution": "employer_contribution",
    "AG_deduction": "ag_deduction",
    "treasury_deduction": "treasury_deduction",
    "gross_adjustment": "gross_adjustment",
    "external_recovery": "external_recovery",
}

_TWO = Decimal("0.01")
_ZERO = Decimal("0.00")

# Module-level cache so the expensive June seed runs once per process.
_CACHED_WORLD: dict | None = None


def _dec(value: object) -> Decimal:
    return Decimal(str(value)).quantize(_TWO)


async def _bind(session: AsyncSession, org_id: UUID, user_id: UUID) -> None:
    if session.in_transaction():
        await session.rollback()
    await session.begin()
    await bind_tenant_context(session, organization_id=org_id, user_id=user_id)


async def _seed_one_employee(
    session: AsyncSession,
    *,
    org_id: UUID,
    user_id: UUID,
    employee: EmployeeSeed,
    office_ids: dict[str, UUID],
    post_id: UUID,
    component_ids: dict[str, UUID],
) -> UUID:
    basic = line_amount(employee, BASIC_CODE)
    assert basic is not None, f"{employee.fixture_id} missing BASIC"
    regime, gpf_jurisdiction = map_regime(employee.regime)

    header = Employee(organization_id=org_id, employee_number=employee.fixture_id)
    session.add(header)
    await session.flush()

    profile_values: dict = {
        "name": employee.name,
        "sevarth_id": employee.sevarth_id,
        "pan": employee.pan,
        "date_of_birth": date(1985, 1, 15),
        "date_of_joining": date(2010, 6, 1),
        "retirement_regime": regime,
        "gpf_jurisdiction": gpf_jurisdiction,
        "pran": None,
        "gpf_account_number": None,
        "epf_number": None,
        "pension_account": None,
    }
    if regime == "gpf":
        profile_values["gpf_account_number"] = employee.gpf_account
        if employee.pran:
            profile_values["pran"] = employee.pran
    elif regime == "nps":
        profile_values["pran"] = employee.pran or f"9000{employee.fixture_id[-4:].zfill(8)}"
    elif regime == "epf":
        profile_values["epf_number"] = employee.epf_number or f"SYNTEPF/{employee.fixture_id}/UAN"

    await versioning.insert_version(
        session,
        employee_profile_versions,
        organization_id=org_id,
        header_id=header.id,
        effective_from=EFFECTIVE_FROM,
        values=profile_values,
        change_reason=None,
        created_by=user_id,
    )
    await versioning.insert_version(
        session,
        employee_posting_versions,
        organization_id=org_id,
        header_id=header.id,
        effective_from=EFFECTIVE_FROM,
        values={
            "office_id": office_ids[employee.office_id],
            "post_id": post_id,
        },
        change_reason=None,
        created_by=user_id,
    )
    await versioning.insert_version(
        session,
        employee_pay_versions,
        organization_id=org_id,
        header_id=header.id,
        effective_from=EFFECTIVE_FROM,
        values={"pay_matrix_level": "L10", "basic_pay": money(basic).quantize(_TWO)},
        change_reason=None,
        created_by=user_id,
    )

    for line in employee.lines:
        code = line.component_code
        if code == BASIC_CODE:
            continue
        if code in RECURRING_COMPONENT_CODES:
            instruction = RecurringInstruction(
                organization_id=org_id,
                employee_id=header.id,
                component_id=component_ids[code],
            )
            session.add(instruction)
            await session.flush()
            await versioning.insert_version(
                session,
                recurring_instruction_versions,
                organization_id=org_id,
                header_id=instruction.id,
                effective_from=EFFECTIVE_FROM,
                values={
                    "amount": money(line.amount).quantize(_TWO),
                    "rate": None,
                    "reason": f"June fixture {code}",
                },
                change_reason=None,
                created_by=user_id,
            )
            continue
        if code in ADVANCE_COMPONENT_CODES:
            principal = max(line.amount * Decimal("24"), line.amount)
            advance = AdvanceAccount(
                organization_id=org_id,
                employee_id=header.id,
                advance_type="hba",
                principal=money(principal).quantize(_TWO),
                sanctioned_on=EFFECTIVE_FROM,
                reference=f"HBA-{employee.fixture_id}",
            )
            session.add(advance)
            await session.flush()
            await versioning.insert_version(
                session,
                advance_installment_versions,
                organization_id=org_id,
                header_id=advance.id,
                effective_from=EFFECTIVE_FROM,
                values={
                    "installment_amount": money(line.amount).quantize(_TWO),
                    "installments_total": 24,
                    "installments_recovered_opening": 0,
                },
                change_reason=None,
                created_by=user_id,
            )
            continue
        if code == "ACCOMMODATION_LICENSE_FEE":
            assert employee.accommodation is not None
            foregone = line_amount(employee, "FOREGONE_HRA")
            assignment = AccommodationAssignment(
                organization_id=org_id,
                employee_id=header.id,
                quarters_location=map_quarters_location(employee.accommodation.location),
                quarters_identifier=f"Q-{employee.fixture_id}",
            )
            session.add(assignment)
            await session.flush()
            await versioning.insert_version(
                session,
                accommodation_charge_versions,
                organization_id=org_id,
                header_id=assignment.id,
                effective_from=EFFECTIVE_FROM,
                values={
                    "license_fee": money(line.amount).quantize(_TWO),
                    "informational_hra_foregone": (
                        money(foregone).quantize(_TWO) if foregone is not None else None
                    ),
                },
                change_reason=None,
                created_by=user_id,
            )
            continue
        if code in ACCOMMODATION_COMPONENT_CODES:
            continue
        raise AssertionError(f"No seeding strategy for {employee.fixture_id} line {code}")

    return header.id


async def _seed_posted_june(session: AsyncSession) -> dict:
    """Seed fixture_loader June world, calculate, approve directly, and post."""
    fixture = load_june_fixture()
    assert len(fixture.employees) == 32

    if session.in_transaction():
        await session.rollback()

    org = await seed_organization(
        session,
        name=fixture.organization.name,
        slug=f"statutory-{uuid4().hex[:10]}",
    )
    user = await seed_user(session, workos_user_id=f"statutory_{uuid4().hex[:10]}")
    await session.commit()

    await _bind(session, org.id, user.id)

    office_ids: dict[str, UUID] = {}
    for office in fixture.organization.offices:
        row = Office(
            organization_id=org.id,
            name=office.name,
            jurisdiction=office.jurisdiction,
        )
        session.add(row)
        await session.flush()
        office_ids[office.fixture_id] = row.id

    post = Post(
        organization_id=org.id,
        designation="Synthetic Clerk",
        class_="III",
    )
    session.add(post)
    await session.flush()

    component_ids: dict[str, UUID] = {}
    display_order = 0
    for comp in fixture.components:
        display_order += 1
        api_cls = comp.api_classification
        if api_cls is None:
            assert comp.code == "FOREGONE_HRA"
            continue
        db_cls = _CLASSIFICATION_DB[comp.fixture_classification]
        component = PayComponent(
            organization_id=org.id,
            code=comp.code,
            name=comp.name,
            classification=db_cls,
            display_order=display_order,
            employer_transfer=comp.employer_transfer,
            transfer_of=comp.transfer_of,
        )
        session.add(component)
        await session.flush()
        component_ids[comp.code] = component.id

        if comp.code in RECURRING_COMPONENT_CODES or comp.code == BASIC_CODE:
            await versioning.insert_version(
                session,
                component_rate_versions,
                organization_id=org.id,
                header_id=component.id,
                effective_from=EFFECTIVE_FROM,
                values={
                    "calc_kind": "fixed_recurring_amount",
                    "amount": Decimal("0.00"),
                    "rate": None,
                    "basis": None,
                    "rounding_rule": "ROUND_HALF_UP_RUPEE",
                },
                change_reason=None,
                created_by=user.id,
            )

    employee_ids: dict[str, UUID] = {}
    for emp in fixture.employees:
        employee_id = await _seed_one_employee(
            session,
            org_id=org.id,
            user_id=user.id,
            employee=emp,
            office_ids=office_ids,
            post_id=post.id,
            component_ids=component_ids,
        )
        employee_ids[emp.fixture_id] = employee_id

    period = PayrollPeriod(
        organization_id=org.id,
        period_year=PERIOD_YEAR,
        period_month=PERIOD_MONTH,
        status="open",
    )
    session.add(period)
    await session.flush()
    run = PayrollRun(
        organization_id=org.id,
        period_id=period.id,
        status="draft",
    )
    session.add(run)
    await session.flush()
    session.add_all(
        initialize_run_roster(
            organization_id=org.id,
            run=run,
            employee_ids=list(employee_ids.values()),
            period_year=period.period_year,
            period_month=period.period_month,
        )
    )
    await session.commit()

    await _bind(session, org.id, user.id)
    calc = await calculate_run_command(
        session,
        organization_id=org.id,
        run_id=run.id,
        user_id=user.id,
    )

    await _bind(session, org.id, user.id)
    run_row = await session.get(PayrollRun, run.id)
    assert run_row is not None
    run_row.status = "approved"
    session.add(
        PayrollApproval(
            organization_id=org.id,
            run_id=run.id,
            run_version_id=calc["version_id"],
            content_hash=calc["content_hash"],
            action="approve",
            actor_user_id=user.id,
            reason="June golden approve",
        )
    )
    await session.commit()

    await _bind(session, org.id, user.id)
    posted = await post_run(
        session,
        organization_id=org.id,
        run_id=run.id,
        user_id=user.id,
    )
    assert posted["status"] == "posted"

    return {
        "org_id": org.id,
        "org_name": org.name,
        "user_id": user.id,
        "run_id": run.id,
        "version_id": UUID(str(calc["version_id"])),
        "employee_ids": employee_ids,
        "expected": fixture.expected.aggregates,
        "engine_version": calc.get("engine_version") or "0.1.0",
        "fixture": fixture,
    }


async def _june_world(session: AsyncSession) -> dict:
    global _CACHED_WORLD
    if _CACHED_WORLD is None:
        _CACHED_WORLD = await _seed_posted_june(session)
    return _CACHED_WORLD


def _ctx(world: dict, *, run_id: UUID | None = None) -> ReportContext:
    return ReportContext(
        organization_id=world["org_id"],
        posted_run_id=run_id or world["run_id"],
        template_version=TEMPLATE_VERSION,
        generated_at=datetime.now(UTC),
        engine_version=str(world["engine_version"]),
    )


def _section_by_title(dto, title: str):
    for section in dto.sections:
        if section.title == title:
            return section
    raise AssertionError(f"section {title!r} not found")


def _col_index(section, key: str) -> int:
    for idx, col in enumerate(section.columns):
        if col.key == key:
            return idx
    raise AssertionError(f"column {key!r} not found")


async def _posted_lines_by_employee(
    session: AsyncSession,
    *,
    org_id: UUID,
    version_id: UUID,
    component_code: str,
) -> dict[str, Decimal]:
    """Map employee_number → posted line amount for ``component_code``."""
    results = (
        (
            await session.execute(
                sa.select(payroll_employee_results).where(
                    payroll_employee_results.c.organization_id == org_id,
                    payroll_employee_results.c.run_version_id == version_id,
                )
            )
        )
        .mappings()
        .all()
    )
    by_id = {row["id"]: row for row in results}
    if not by_id:
        return {}

    lines = (
        (
            await session.execute(
                sa.select(payroll_result_lines).where(
                    payroll_result_lines.c.organization_id == org_id,
                    payroll_result_lines.c.employee_result_id.in_(list(by_id.keys())),
                    payroll_result_lines.c.component_code == component_code,
                )
            )
        )
        .mappings()
        .all()
    )
    out: dict[str, Decimal] = {}
    for line in lines:
        emp_no = by_id[line["employee_result_id"]]["employee_number"]
        out[emp_no] = out.get(emp_no, _ZERO) + _dec(line["amount"])
    return out


async def _posted_gross_by_employee(
    session: AsyncSession,
    *,
    org_id: UUID,
    version_id: UUID,
) -> dict[str, Decimal]:
    rows = (
        (
            await session.execute(
                sa.select(payroll_employee_results).where(
                    payroll_employee_results.c.organization_id == org_id,
                    payroll_employee_results.c.run_version_id == version_id,
                )
            )
        )
        .mappings()
        .all()
    )
    return {row["employee_number"]: _dec(row["gross_total"]) for row in rows}


@pytest.mark.asyncio
async def test_income_tax_schedule_golden(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    dto = await income_tax_builder.build(session, _ctx(world))

    assert dto.report_type == REPORT_TYPE_INCOME_TAX
    section = _section_by_title(dto, "Income Tax")
    assert len(section.rows) == 23
    assert section.totals is not None
    tax_idx = _col_index(section, "income_tax")
    assert _dec(section.totals[tax_idx]) == _dec("550700.00")

    db_tax = await _posted_lines_by_employee(
        session,
        org_id=world["org_id"],
        version_id=world["version_id"],
        component_code="INCOME_TAX",
    )
    db_gross = await _posted_gross_by_employee(
        session,
        org_id=world["org_id"],
        version_id=world["version_id"],
    )
    assert len(db_tax) == 23

    pan_idx = _col_index(section, "pan")
    gross_idx = _col_index(section, "gross")
    emp_idx = _col_index(section, "employee_number")

    schedule_numbers = {row[emp_idx] for row in section.rows}
    assert schedule_numbers == set(db_tax.keys())
    # Employees without an INCOME_TAX line are absent (9 of 32).
    assert len(db_gross) == 32
    assert len(set(db_gross.keys()) - schedule_numbers) == 9

    for row in section.rows:
        emp_no = row[emp_idx]
        assert _PAN_RE.match(str(row[pan_idx])), row[pan_idx]
        assert _dec(row[tax_idx]) == db_tax[emp_no]
        assert _dec(row[gross_idx]) == db_gross[emp_no]


@pytest.mark.asyncio
async def test_professional_tax_schedule_golden(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    dto = await professional_tax_builder.build(session, _ctx(world))

    assert dto.report_type == REPORT_TYPE_PROFESSIONAL_TAX
    section = _section_by_title(dto, "Professional Tax")
    assert len(section.rows) == 28
    assert section.totals is not None
    pt_idx = _col_index(section, "professional_tax")
    assert _dec(section.totals[pt_idx]) == _dec("5600.00")

    db_pt = await _posted_lines_by_employee(
        session,
        org_id=world["org_id"],
        version_id=world["version_id"],
        component_code="PROFESSIONAL_TAX",
    )
    assert len(db_pt) == 28

    emp_idx = _col_index(section, "employee_number")
    schedule_numbers = {row[emp_idx] for row in section.rows}
    assert schedule_numbers == set(db_pt.keys())

    # 4 non-liable employees absent.
    all_emps = set(world["employee_ids"].keys())
    assert len(all_emps - schedule_numbers) == 4

    for row in section.rows:
        emp_no = row[emp_idx]
        assert _dec(row[pt_idx]) == db_pt[emp_no]

    slab = _section_by_title(dto, "Slab note")
    assert len(slab.rows) == 1
    note = str(slab.rows[0][0])
    assert "slab" in note.lower() or "Professional tax" in note
    assert "28" in note
    # Fixture happens to post 200 for every PT line; note must reflect posted data.
    assert "200" in note


@pytest.mark.asyncio
async def test_gis_schedule_golden(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    dto = await gis_builder.build(session, _ctx(world))

    assert dto.report_type == REPORT_TYPE_GIS
    section = _section_by_title(dto, "GIS")
    assert len(section.rows) == 32
    assert section.totals is not None
    gis_idx = _col_index(section, "gis")
    assert _dec(section.totals[gis_idx]) == _dec("22440.00")

    db_gis = await _posted_lines_by_employee(
        session,
        org_id=world["org_id"],
        version_id=world["version_id"],
        component_code="GIS",
    )
    assert len(db_gis) == 32

    emp_idx = _col_index(section, "employee_number")
    for row in section.rows:
        emp_no = row[emp_idx]
        assert _dec(row[gis_idx]) == db_gis[emp_no]

    counts = Counter(_dec(row[gis_idx]) for row in section.rows)
    assert counts[_dec("960.00")] == 14
    assert counts[_dec("480.00")] == 10
    assert counts[_dec("240.00")] == 4
    assert counts[_dec("720.00")] == 2
    assert counts[_dec("800.00")] == 1
    assert counts[_dec("1000.00")] == 1


@pytest.mark.asyncio
async def test_statutory_excel_and_pdf_formatters(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])

    it_dto = await income_tax_builder.build(session, _ctx(world))
    xlsx = statutory_to_excel(it_dto)
    wb = load_workbook(BytesIO(xlsx))
    ws = wb.active
    assert ws is not None
    # Header row 5; first data row 6; income_tax is column 5.
    money_cell = ws.cell(row=6, column=5)
    assert money_cell.number_format == MONEY_FORMAT
    section = _section_by_title(it_dto, "Income Tax")
    totals_excel_row = 6 + len(section.rows)
    total_cell = ws.cell(row=totals_excel_row, column=5)
    assert _dec(total_cell.value) == _dec("550700.00")

    pdf = statutory_to_pdf(it_dto)
    assert pdf.startswith(b"%PDF")
    text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(pdf)).pages)
    assert "Income Tax" in text
    assert format_inr(_dec("550700.00")) in text

    pt_dto = await professional_tax_builder.build(session, _ctx(world))
    pt_xlsx = statutory_to_excel(pt_dto)
    pt_wb = load_workbook(BytesIO(pt_xlsx))
    assert pt_wb.active is not None
    pt_pdf = statutory_to_pdf(pt_dto)
    pt_text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(pt_pdf)).pages)
    assert "Professional Tax" in pt_text
    assert format_inr(_dec("5600.00")) in pt_text

    gis_dto = await gis_builder.build(session, _ctx(world))
    gis_pdf = statutory_to_pdf(gis_dto)
    gis_text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(gis_pdf)).pages)
    assert "GIS" in gis_text
    assert format_inr(_dec("22440.00")) in gis_text


@pytest.mark.asyncio
async def test_unposted_run_raises_conflict(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])

    period = PayrollPeriod(
        organization_id=world["org_id"],
        period_year=2026,
        period_month=7,
        status="open",
    )
    session.add(period)
    await session.flush()
    draft = PayrollRun(
        organization_id=world["org_id"],
        period_id=period.id,
        status="draft",
    )
    session.add(draft)
    await session.flush()
    draft_id = draft.id

    calculated_period = PayrollPeriod(
        organization_id=world["org_id"],
        period_year=2026,
        period_month=8,
        status="open",
    )
    session.add(calculated_period)
    await session.flush()
    calculated = PayrollRun(
        organization_id=world["org_id"],
        period_id=calculated_period.id,
        status="calculated",
    )
    session.add(calculated)
    await session.flush()
    calculated_id = calculated.id
    await session.commit()

    for builder in (income_tax_builder, professional_tax_builder, gis_builder):
        await _bind(session, world["org_id"], world["user_id"])
        with pytest.raises(ConflictError, match="must be posted"):
            await builder.build(session, _ctx(world, run_id=draft_id))

        await _bind(session, world["org_id"], world["user_id"])
        with pytest.raises(ConflictError, match="must be posted"):
            await builder.build(session, _ctx(world, run_id=calculated_id))


def test_register_statutory_reports():
    registry = ReportRegistry()
    register(registry)
    assert REPORT_TYPE_INCOME_TAX in registry
    assert REPORT_TYPE_PROFESSIONAL_TAX in registry
    assert REPORT_TYPE_GIS in registry
    for report_type in (
        REPORT_TYPE_INCOME_TAX,
        REPORT_TYPE_PROFESSIONAL_TAX,
        REPORT_TYPE_GIS,
    ):
        entry = registry.get(report_type)
        assert entry.report_type == report_type
        assert entry.builder is not None
        assert entry.formatters.to_json is not None
        assert entry.formatters.to_excel is not None
        assert entry.formatters.to_pdf is not None
