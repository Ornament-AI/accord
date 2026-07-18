"""Tests for the generic safe Excel report formatter."""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.reports.base import ColumnKind, ReportColumn, ReportDTO, TableSection
from app.reports.excel import MONEY_FORMAT, to_excel


def _sample_register_dto() -> ReportDTO:
    return ReportDTO(
        report_type="pay_bill",
        template_version="v1",
        title="Payroll Register — Pay Bill",
        organization_name="Accord Demo Org",
        subtitle="June 2026",
        sections=(
            TableSection(
                title="Register",
                columns=(
                    ReportColumn(key="employee", header="Employee", kind=ColumnKind.TEXT),
                    ReportColumn(key="note", header="Note", kind=ColumnKind.TEXT),
                    ReportColumn(key="gross", header="Gross", kind=ColumnKind.MONEY),
                ),
                rows=(
                    (
                        "Ada Lovelace",
                        '=HYPERLINK("http://evil.example","x")',
                        Decimal("5102985.00"),
                    ),
                    (
                        "Grace Hopper",
                        "=2+2",
                        Decimal("1000.00"),
                    ),
                    (
                        "Alan Turing",
                        "+cmd|'/c calc'!A0",
                        Decimal("200.50"),
                    ),
                    (
                        "Katherine Johnson",
                        "@SUM(A1:A2)",
                        Decimal("50.00"),
                    ),
                ),
            ),
        ),
    )


def test_to_excel_bytes_reload_and_money_format() -> None:
    raw = to_excel(_sample_register_dto())
    assert isinstance(raw, (bytes, bytearray))
    assert len(raw) > 0

    wb = load_workbook(BytesIO(raw))
    ws = wb.active
    assert ws is not None

    # Header row is row 5; freeze panes keep it visible.
    assert ws.freeze_panes == "A6"

    # Data rows start at row 6; Gross is column 3.
    money_cell = ws.cell(row=6, column=3)
    assert money_cell.value == 5102985.0
    assert money_cell.number_format == MONEY_FORMAT
    assert MONEY_FORMAT == "#,##,##0.00"


def test_to_excel_neutralizes_formula_injection() -> None:
    wb = load_workbook(BytesIO(to_excel(_sample_register_dto())))
    ws = wb.active
    assert ws is not None

    # Note column is column 2; four malicious rows at data rows 6-9.
    for row in range(6, 10):
        cell = ws.cell(row=row, column=2)
        assert cell.data_type != "f", f"row {row} stored as formula: {cell.value!r}"
        text = str(cell.value)
        # openpyxl may keep the leading apostrophe or store as plain text.
        assert text.startswith("'") or text.startswith(("=", "+", "@"))
        if text.startswith(("=", "+", "@")):
            # Must not be executable formula type.
            assert cell.data_type in {"s", "str", "inlineStr"}
