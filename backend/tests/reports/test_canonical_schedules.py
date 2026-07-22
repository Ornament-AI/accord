"""Canonical v3 schedule layout and privacy contract tests."""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO
import json
from pathlib import Path
import re
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pytest

from app.reports.base import ColumnKind, ReportColumn, ReportDTO, TableSection
from app.reports.canonical_schedules import (
    REPORT_SHEET_NAMES,
    TEMPLATE_PATH,
    canonical_schedule_to_excel,
)
from app.reports.families.recovery import recovery_to_excel
from app.reports.families.retirement import retirement_to_excel
from app.reports.families.statutory import statutory_to_excel


ROOT = Path(__file__).resolve().parents[3]
CONTRACT_PATH = ROOT / "fixtures" / "sanitized" / "june-2026" / "canonical_export_contract.json"


@pytest.fixture
def clean_identity_tables():
    """This formatter-only module has no database state to reset."""
    yield


def _dto(report_type: str, keys: tuple[str, ...], row: tuple[object, ...]) -> ReportDTO:
    money_keys = {
        "income_tax",
        "professional_tax",
        "gis",
        "basic_pay",
        "subscription",
        "advance_recovery",
        "scheduled_installment_amount",
        "installment_amount",
        "dearness_allowance",
        "employee_contribution",
        "employer_contribution",
        "informational_foregone_hra",
        "house_rent",
        "service_charge",
        "parking_charge",
        "additional_parking_charge",
        "license_fee_actual",
    }
    columns = tuple(
        ReportColumn(key, key, ColumnKind.MONEY if key in money_keys else ColumnKind.TEXT)
        for key in keys
    )
    return ReportDTO(
        report_type=report_type,
        template_version="v3",
        title="Schedule",
        organization_name="Example Public Corporation",
        subtitle="June 2026",
        sections=(TableSection("Schedule", columns, (row,)),),
        metadata={
            "report_profile": {
                "legal_name": "Ornament Legal Ltd.",
                "office_name": "Payroll Office",
                "nps_employee_account_head": "8342 Employee contribution",
                "nps_employer_account_head": "8342 Employer contribution",
                "head_of_account": {"major_head": "2070"},
                "gpf_remittance_profiles": {
                    "mumbai": {
                        "office_name": "Accountant General Maharashtra I Mumbai",
                        "address_lines": ["101 Mumbai Road", "Mumbai 400001"],
                        "account_code": "GPF-MUM-001",
                        "authority_text": "Mumbai Authority",
                    },
                    "nagpur": {
                        "office_name": "Accountant General Maharashtra II Nagpur",
                        "address_lines": ["202 Nagpur Road", "Nagpur 440001"],
                        "account_code": "GPF-NGP-002",
                        "authority_text": "Nagpur Authority",
                    },
                },
                "signatories": [
                    {
                        "role": "approving_officer",
                        "name": "Asha Approver",
                        "designation": "Chief Administrative Officer",
                    }
                ],
            }
        },
    )


@pytest.mark.parametrize("report_type", tuple(REPORT_SHEET_NAMES))
def test_individual_v3_schedule_routes_use_canonical_renderer(
    report_type: str, monkeypatch
) -> None:
    calls: list[str] = []

    def render(dto: ReportDTO) -> bytes:
        calls.append(dto.report_type)
        return b"canonical"

    monkeypatch.setattr("app.reports.canonical_schedules.canonical_schedule_to_excel", render)
    dto = ReportDTO(
        report_type=report_type,
        template_version="v3",
        title="Schedule",
        organization_name="Example",
        subtitle="June 2026",
        sections=(),
    )
    if report_type in {"income_tax_schedule", "professional_tax_schedule", "gis_schedule"}:
        formatter = statutory_to_excel
    elif report_type in {
        "gpf_nagpur_schedule",
        "gpf_mumbai_schedule",
        "nps_contribution_schedule",
    }:
        formatter = retirement_to_excel
    else:
        formatter = recovery_to_excel
    assert formatter(dto) == b"canonical"
    assert calls == [report_type]


def _dto_with_row_count(
    report_type: str,
    keys: tuple[str, ...],
    row: tuple[object, ...],
    row_count: int,
) -> ReportDTO:
    dto = _dto(report_type, keys, row)
    section = dto.sections[0]
    return ReportDTO(
        report_type=dto.report_type,
        template_version=dto.template_version,
        title=dto.title,
        organization_name=dto.organization_name,
        subtitle=dto.subtitle,
        sections=(
            TableSection(
                section.title,
                section.columns,
                tuple(row for _ in range(row_count)),
            ),
        ),
    )


@pytest.mark.parametrize(
    ("report_type", "keys", "row", "formula_cell"),
    (
        (
            "income_tax_schedule",
            ("employee_number", "name", "designation", "pan", "financial_year", "income_tax"),
            ("E001", "Employee One", "Officer", "ABCDE1234F", "2026-27", Decimal("100")),
            "F30",
        ),
        (
            "professional_tax_schedule",
            ("employee_number", "name", "designation", "professional_tax"),
            ("E001", "Employee One", "Officer", Decimal("200")),
            "E33",
        ),
        (
            "gis_schedule",
            ("employee_number", "name", "designation", "gis"),
            ("E001", "Employee One", "Officer", Decimal("960")),
            "E33",
        ),
        (
            "gpf_nagpur_schedule",
            (
                "employee_number",
                "name",
                "designation",
                "gpf_account_number",
                "basic_pay",
                "subscription",
                "advance_recovery",
            ),
            (
                "E001",
                "Employee One",
                "Officer",
                "GPF-1",
                Decimal("50000"),
                Decimal("5000"),
                Decimal("500"),
            ),
            "I41",
        ),
        (
            "gpf_mumbai_schedule",
            (
                "employee_number",
                "name",
                "designation",
                "gpf_account_number",
                "basic_pay",
                "subscription",
                "advance_recovery",
            ),
            (
                "E001",
                "Employee One",
                "Officer",
                "GPF-1",
                Decimal("50000"),
                Decimal("5000"),
                Decimal("500"),
            ),
            "I46",
        ),
        (
            "gpf_advance_schedule",
            (
                "employee_number",
                "name",
                "designation",
                "advance_reference",
                "scheduled_installment_amount",
                "installment_amount",
                "installments_progress",
            ),
            ("E001", "Employee One", "Officer", "ADV-1", Decimal("500"), Decimal("500"), "2/10"),
            "G23",
        ),
        (
            "hba_schedule",
            (
                "employee_number",
                "name",
                "designation",
                "installment_amount",
                "installments_progress",
            ),
            ("E001", "Employee One", "Officer", Decimal("1000"), "2/10"),
            "E12",
        ),
        (
            "motor_car_advance_schedule",
            (
                "employee_number",
                "name",
                "designation",
                "installment_amount",
                "installments_progress",
            ),
            ("E001", "Employee One", "Officer", Decimal("1000"), "2/10"),
            "E7",
        ),
        (
            "motorcycle_advance_schedule",
            (
                "employee_number",
                "name",
                "designation",
                "installment_amount",
                "installments_progress",
            ),
            ("E001", "Employee One", "Officer", Decimal("1000"), "2/10"),
            "E7",
        ),
        (
            "festival_advance_schedule",
            (
                "employee_number",
                "name",
                "designation",
                "installment_amount",
                "installments_progress",
            ),
            ("E001", "Employee One", "Officer", Decimal("1000"), "2/10"),
            "E30",
        ),
        (
            "nps_contribution_schedule",
            (
                "employee_number",
                "pension_account",
                "name",
                "sevarth_id",
                "pran",
                "month",
                "basic_pay",
                "dearness_allowance",
                "employee_contribution",
                "employer_contribution",
                "remarks",
            ),
            (
                "E001",
                "PEN-1",
                "Employee One",
                "SEV-1",
                "PRAN-1",
                "June-26",
                Decimal("50000"),
                Decimal("29000"),
                Decimal("7900"),
                Decimal("11060"),
                "",
            ),
            "I46",
        ),
        (
            "accommodation_worli_schedule",
            (
                "employee_number",
                "name",
                "designation",
                "quarters_address",
                "informational_foregone_hra",
                "house_rent",
                "service_charge",
                "parking_charge",
                "additional_parking_charge",
                "license_fee_actual",
            ),
            (
                "E001",
                "Employee One",
                "Officer",
                "Quarter 1",
                Decimal("5000"),
                Decimal("800"),
                Decimal("300"),
                Decimal("0"),
                Decimal("0"),
                Decimal("1100"),
            ),
            "H8",
        ),
        (
            "accommodation_mumbai_schedule",
            (
                "employee_number",
                "name",
                "designation",
                "quarters_address",
                "informational_foregone_hra",
                "house_rent",
                "service_charge",
                "parking_charge",
                "additional_parking_charge",
                "license_fee_actual",
            ),
            (
                "E001",
                "Employee One",
                "Officer",
                "Quarter 1",
                Decimal("5000"),
                Decimal("800"),
                Decimal("300"),
                Decimal("100"),
                Decimal("50"),
                Decimal("1250"),
            ),
            "J8",
        ),
    ),
)
def test_each_schedule_has_standalone_canonical_renderer(
    report_type: str,
    keys: tuple[str, ...],
    row: tuple[object, ...],
    formula_cell: str,
) -> None:
    content = canonical_schedule_to_excel(_dto(report_type, keys, row))
    workbook = load_workbook(BytesIO(content), data_only=False)
    assert workbook.sheetnames == [REPORT_SHEET_NAMES[report_type]]
    assert workbook.active[formula_cell].data_type == "f"


def test_nps_allocation_rows_match_canonical_source_semantics() -> None:
    keys = (
        "employee_number",
        "pension_account",
        "name",
        "sevarth_id",
        "pran",
        "month",
        "basic_pay",
        "dearness_allowance",
        "employee_contribution",
        "employer_contribution",
        "remarks",
    )
    row = (
        "E001",
        "PEN-1",
        "Employee One",
        "SEV-1",
        "PRAN-1",
        "June-26",
        Decimal("50000"),
        Decimal("29000"),
        Decimal("7900"),
        Decimal("11060"),
        "",
    )
    sheet = load_workbook(
        BytesIO(canonical_schedule_to_excel(_dto("nps_contribution_schedule", keys, row))),
        data_only=False,
    ).active
    assert sheet["B42"].value == sheet["B44"].value == "8342 Employee contribution"
    assert sheet["B43"].value == sheet["B45"].value == "8342 Employer contribution"
    assert sheet["I42"].value == sheet["I43"].value == 0
    assert sheet["I44"].value == "=H41"
    assert sheet["I45"].value == "=I41"
    assert sheet["I46"].value == "=SUM(I42:I45)"
    assert sheet["C47"].value == "=I46"
    assert sheet["G47"].value == "Major Head: 2070"
    assert "Asha Approver" in sheet["G55"].value


@pytest.mark.parametrize(
    ("report_type", "row_formula", "office_name", "address", "account_code", "authority"),
    (
        (
            "gpf_nagpur_schedule",
            "=SUM(F36:G36)",
            "Accountant General Maharashtra II Nagpur",
            "202 Nagpur Road, Nagpur 440001",
            "GPF-NGP-002",
            "Nagpur Authority",
        ),
        (
            "gpf_mumbai_schedule",
            "=SUM(F36:G36)",
            "Accountant General Maharashtra I Mumbai",
            "101 Mumbai Road, Mumbai 400001",
            "GPF-MUM-001",
            "Mumbai Authority",
        ),
    ),
)
def test_gpf_uses_canonical_formula_and_lowercase_remittance_profile(
    report_type: str,
    row_formula: str,
    office_name: str,
    address: str,
    account_code: str,
    authority: str,
) -> None:
    keys = (
        "employee_number",
        "name",
        "designation",
        "gpf_account_number",
        "basic_pay",
        "subscription",
        "advance_recovery",
    )
    row = (
        "E001",
        "Employee One",
        "Officer",
        "GPF-1",
        Decimal("50000"),
        Decimal("5000"),
        Decimal("500"),
    )
    sheet = load_workbook(
        BytesIO(canonical_schedule_to_excel(_dto(report_type, keys, row))),
        data_only=False,
    ).active

    assert sheet["I36"].value == row_formula
    assert sheet["B16"].value == "Drawing and disbursing office: Ornament Legal Ltd."
    assert sheet["A23"].value == "Ornament Legal Ltd."
    assert sheet["B20"].value == office_name
    assert address in sheet["B21"].value
    assert account_code in sheet["B21"].value
    assert authority in sheet["B21"].value


def test_hba_total_formula_keeps_canonical_placeholder_row() -> None:
    keys = (
        "employee_number",
        "name",
        "designation",
        "installment_amount",
        "installments_progress",
    )
    row = ("E001", "Employee One", "Officer", Decimal("1000"), "2/10")
    sheet = load_workbook(
        BytesIO(canonical_schedule_to_excel(_dto("hba_schedule", keys, row))),
        data_only=False,
    ).active

    assert sheet["E11"].value is None
    assert sheet["E12"].value == "=SUM(E5:E11)"


@pytest.mark.parametrize(
    ("report_type", "keys", "row", "used_range"),
    (
        (
            "professional_tax_schedule",
            ("employee_number", "name", "designation", "professional_tax"),
            ("E001", "Employee One", "Officer", Decimal("200")),
            "B1:G35",
        ),
        (
            "gis_schedule",
            ("employee_number", "name", "designation", "gis"),
            ("E001", "Employee One", "Officer", Decimal("960")),
            "B1:F35",
        ),
        (
            "nps_contribution_schedule",
            (
                "employee_number",
                "pension_account",
                "name",
                "sevarth_id",
                "pran",
                "month",
                "basic_pay",
                "dearness_allowance",
                "employee_contribution",
                "employer_contribution",
                "remarks",
            ),
            (
                "E001",
                "PEN-1",
                "Employee One",
                "SEV-1",
                "PRAN-1",
                "June-26",
                Decimal("50000"),
                Decimal("29000"),
                Decimal("7900"),
                Decimal("11060"),
                "",
            ),
            "B1:J57",
        ),
    ),
)
def test_schedule_renderer_preserves_canonical_used_range_start(
    report_type: str,
    keys: tuple[str, ...],
    row: tuple[object, ...],
    used_range: str,
) -> None:
    sheet = load_workbook(
        BytesIO(canonical_schedule_to_excel(_dto(report_type, keys, row))),
        data_only=False,
    ).active

    assert sheet.calculate_dimension() == used_range
    assert not any(cell.value is not None for cell in sheet["A"])


def test_accommodation_rejects_breakdown_that_does_not_match_actual_recovery() -> None:
    keys = (
        "employee_number",
        "name",
        "designation",
        "quarters_address",
        "informational_foregone_hra",
        "house_rent",
        "service_charge",
        "parking_charge",
        "additional_parking_charge",
        "license_fee_actual",
    )
    row = (
        "E001",
        "Employee One",
        "Officer",
        "Quarter 1",
        Decimal("5000"),
        Decimal("800"),
        Decimal("300"),
        Decimal("100"),
        Decimal("50"),
        Decimal("999"),
    )
    with pytest.raises(ValueError, match="do not reconcile"):
        canonical_schedule_to_excel(_dto("accommodation_mumbai_schedule", keys, row))


def test_accommodation_total_excludes_informational_foregone_hra() -> None:
    keys = (
        "employee_number",
        "name",
        "designation",
        "quarters_address",
        "informational_foregone_hra",
        "house_rent",
        "service_charge",
        "parking_charge",
        "additional_parking_charge",
        "license_fee_actual",
    )
    row = (
        "E001",
        "Employee One",
        "Officer",
        "Quarter 1",
        Decimal("5000"),
        Decimal("800"),
        Decimal("300"),
        Decimal("100"),
        Decimal("50"),
        Decimal("1250"),
    )
    workbook = load_workbook(
        BytesIO(canonical_schedule_to_excel(_dto("accommodation_mumbai_schedule", keys, row))),
        data_only=False,
    )
    assert workbook.active["J8"].value == "=SUM(F8:I8)"
    assert workbook.active["E8"].value == Decimal("5000")


@pytest.mark.parametrize(
    ("report_type", "keys", "row", "row_count", "total_cell", "print_titles"),
    (
        (
            "income_tax_schedule",
            ("employee_number", "name", "designation", "pan", "financial_year", "income_tax"),
            ("E001", "Employee One", "Officer", "ABCDE1234F", "2026-27", Decimal("100")),
            30,
            "F35",
            "$1:$4",
        ),
        (
            "professional_tax_schedule",
            ("employee_number", "name", "designation", "professional_tax"),
            ("E001", "Employee One", "Officer", Decimal("200")),
            30,
            "E35",
            "$1:$4",
        ),
        (
            "gis_schedule",
            ("employee_number", "name", "designation", "gis"),
            ("E001", "Employee One", "Officer", Decimal("960")),
            28,
            "E35",
            "$1:$6",
        ),
        (
            "gpf_nagpur_schedule",
            (
                "employee_number",
                "name",
                "designation",
                "gpf_account_number",
                "basic_pay",
                "subscription",
                "advance_recovery",
            ),
            (
                "E001",
                "Employee One",
                "Officer",
                "GPF-1",
                Decimal("50000"),
                Decimal("5000"),
                Decimal("500"),
            ),
            7,
            "I43",
            "$27:$34",
        ),
        (
            "nps_contribution_schedule",
            (
                "employee_number",
                "pension_account",
                "name",
                "sevarth_id",
                "pran",
                "month",
                "basic_pay",
                "dearness_allowance",
                "employee_contribution",
                "employer_contribution",
                "remarks",
            ),
            (
                "E001",
                "PEN-1",
                "Employee One",
                "SEV-1",
                "PRAN-1",
                "June-26",
                Decimal("50000"),
                Decimal("29000"),
                Decimal("7900"),
                Decimal("11060"),
                "",
            ),
            12,
            "I47",
            "$1:$10",
        ),
        (
            "hba_schedule",
            (
                "employee_number",
                "name",
                "designation",
                "installment_amount",
                "installments_progress",
            ),
            ("E001", "Employee One", "Officer", Decimal("1000"), "2/10"),
            8,
            "E14",
            "$1:$4",
        ),
        (
            "festival_advance_schedule",
            (
                "employee_number",
                "name",
                "designation",
                "installment_amount",
                "installments_progress",
            ),
            ("E001", "Employee One", "Officer", Decimal("1000"), "2/10"),
            27,
            "E32",
            "$1:$4",
        ),
    ),
)
def test_overflow_uses_same_sheet_continuation_pages(
    report_type: str,
    keys: tuple[str, ...],
    row: tuple[object, ...],
    row_count: int,
    total_cell: str,
    print_titles: str,
) -> None:
    dto = _dto_with_row_count(report_type, keys, row, row_count)
    workbook = load_workbook(BytesIO(canonical_schedule_to_excel(dto)), data_only=False)
    sheet = workbook.active
    assert sheet[total_cell].data_type == "f"
    assert sheet.print_title_rows == print_titles
    assert sheet.row_breaks.count == 1


def test_structural_template_matches_contract_and_contains_no_values() -> None:
    contract = json.loads(CONTRACT_PATH.read_text(encoding="utf-8"))
    workbook = load_workbook(TEMPLATE_PATH, data_only=False)
    assert workbook.sheetnames == [sheet["name"] for sheet in contract["sheets"]]
    for expected in contract["sheets"]:
        sheet = workbook[expected["name"]]
        assert sheet.sheet_state == expected["state"]
        assert sorted(str(item) for item in sheet.merged_cells.ranges) == sorted(
            expected["merged_cells"]
        )
        assert sheet.calculate_dimension() == expected["used_range"]
        actual_print_area = str(sheet.print_area) if sheet.print_area else None
        expected_print_area = expected.get("print_area")
        assert (actual_print_area.replace("'", "") if actual_print_area else None) == (
            expected_print_area.replace("'", "") if expected_print_area else None
        )
        actual_columns = []
        for dimension in sheet.column_dimensions.values():
            item = {
                "min": dimension.min,
                "max": dimension.max,
                "width": str(dimension.width),
            }
            if dimension.hidden:
                item["hidden"] = True
            actual_columns.append(item)
        assert len(actual_columns) == len(expected["column_dimensions"])
        for actual, wanted in zip(actual_columns, expected["column_dimensions"], strict=True):
            assert actual["min"] == wanted["min"]
            assert actual["max"] == wanted["max"]
            assert Decimal(actual["width"]) == Decimal(wanted["width"])
            assert actual.get("hidden") == wanted.get("hidden")
        actual_rows = [
            (index, Decimal(str(dimension.height)))
            for index, dimension in sheet.row_dimensions.items()
            if dimension.height is not None
        ]
        expected_rows = [
            (item["row"], Decimal(item["height"])) for item in expected["row_dimensions"]
        ]
        assert actual_rows == expected_rows
        assert [
            index for index, dimension in sheet.row_dimensions.items() if dimension.hidden
        ] == expected["hidden_rows"]
        assert [
            [dimension.min, dimension.max]
            for dimension in sheet.column_dimensions.values()
            if dimension.hidden
        ] == expected["hidden_columns"]
        for key, value in expected["page_setup"].items():
            assert str(getattr(sheet.page_setup, key)) == value
        for key, value in expected["page_margins"].items():
            assert getattr(sheet.page_margins, key) == pytest.approx(float(value), abs=1e-15)
        for key, value in expected["print_options"].items():
            assert str(int(bool(getattr(sheet.print_options, key)))) == value
        assert [item.id for item in sheet.row_breaks.brk] == expected["manual_row_breaks"]
        assert [item.id for item in sheet.col_breaks.brk] == expected["manual_column_breaks"]
        for row in sheet.iter_rows():
            for cell in row:
                assert cell.value is None
                assert cell.comment is None
                assert cell.hyperlink is None

    # Ensure max-ranged dimensions survive the clean-room copy.
    assert workbook["Income Tax"].column_dimensions[get_column_letter(7)].max == 16384


def test_structural_template_ooxml_has_no_payload_or_external_relationships() -> None:
    with ZipFile(TEMPLATE_PATH) as archive:
        names = archive.namelist()
        assert "xl/sharedStrings.xml" not in names
        assert not any(
            forbidden in name.casefold()
            for name in names
            for forbidden in ("externallink", "comment", "person", "connection")
        )
        xml = b"\n".join(archive.read(name) for name in names if name.endswith(".xml"))
        worksheet_xml = b"\n".join(
            archive.read(name)
            for name in names
            if name.startswith("xl/worksheets/") and name.endswith(".xml")
        )
    assert re.search(rb"\b[A-Z]{5}[0-9]{4}[A-Z]\b", xml) is None
    assert re.search(rb"\b[A-Z]{4}0[A-Z0-9]{6}\b", xml) is None
    assert b"<v>" not in worksheet_xml
    assert b"<is>" not in worksheet_xml
