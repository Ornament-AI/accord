"""Golden tests for Pay Bill and Treasury Face report builders/formatters."""

from __future__ import annotations

import json
import importlib.util
from dataclasses import replace
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from uuid import UUID, uuid4

import pytest
import sqlalchemy as sa
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader
from sqlalchemy.ext.asyncio import AsyncSession

from app.exceptions import ConflictError
from app.models.accommodation import AccommodationAssignment, accommodation_charge_versions
from app.models.advances import AdvanceAccount, advance_installment_versions
from app.models.employees import (
    Employee,
    employee_pay_versions,
    employee_posting_versions,
    employee_profile_versions,
)
from app.models.identity import Organization
from app.models.org_structure import Office, Post
from app.models.pay_components import PayComponent, component_rate_versions
from app.models.payroll_runs import (
    PayrollPeriod,
    PayrollRun,
    payroll_employee_results,
    payroll_report_snapshots,
    payroll_result_lines,
    payroll_run_versions,
)
from app.models.platform import PayrollApproval
from app.models.recurring_instructions import (
    RecurringInstruction,
    recurring_instruction_versions,
)
from app.models.reports import ReportConfiguration
from app.reports.amount_in_words import amount_in_words
from app.reports.base import ReportContext, ReportDTO, TableSection
from app.reports.excel import MONEY_FORMAT
from app.reports.families.payroll_register import (
    REPORT_TYPE_PAY_BILL,
    REPORT_TYPE_TREASURY_FACE,
    pay_bill_builder,
    pay_bill_to_excel,
    pay_bill_to_json,
    pay_bill_to_pdf,
    treasury_face_builder,
    treasury_face_to_excel,
    treasury_face_to_json,
    treasury_face_to_pdf,
)
from app.reports.canonical_excel import _text_preserving_zero, pay_bill_v3_to_excel
from app.reports.canonical_pay_bill_allocation import post_metadata
from app.services import versioning
from app.services.run_calculation import calculate_run_command
from app.services.run_posting import post_run
from app.tenancy import bind_tenant_context
from tests.roster_helpers import initialize_run_roster
from tests.e2e.fixture_loader import (
    ACCOMMODATION_COMPONENT_CODES,
    ADVANCE_COMPONENT_CODES,
    BASIC_CODE,
    RECURRING_COMPONENT_CODES,
    EmployeeSeed,
    line_amount,
    load_june_fixture,
    map_quarters_location,
    map_regime,
    money,
)
from tests.identity_helpers import seed_organization, seed_user

EFFECTIVE_FROM = date(2026, 1, 1)
PERIOD_YEAR = 2026
PERIOD_MONTH = 6
TEMPLATE_VERSION = "v1"

_CLASSIFICATION_DB = {
    "earning": "earning",
    "employer_contribution": "employer_contribution",
    "AG_deduction": "ag_deduction",
    "treasury_deduction": "treasury_deduction",
    "gross_adjustment": "gross_adjustment",
    "external_recovery": "external_recovery",
}


def test_v3_post_metadata_preserves_zero_display_order() -> None:
    metadata = post_metadata(
        {
            "pay_bill_post": {
                "id": "first-group",
                "heading": "First Group",
                "display_order": 0,
            }
        }
    )

    assert metadata[5] == 0


def test_v3_post_metadata_preserves_zero_strength_in_group_key() -> None:
    metadata = post_metadata(
        {
            "pay_bill_post": {
                "heading": "Zero Strength Group",
                "sanctioned_posts": 0,
                "vacant_posts": 0,
            }
        }
    )

    assert metadata[0] == "1000000|Zero Strength Group|0|0|"
    assert metadata[2:4] == (0, 0)
    assert _text_preserving_zero(metadata[2]) == "0"
    assert _text_preserving_zero(metadata[3]) == "0"
    assert _text_preserving_zero(None) == ""


_TWO = Decimal("0.01")
_ZERO = Decimal("0.00")

# Module-level cache so the expensive June seed runs once per process.
_CACHED_WORLD: dict | None = None


def _dec(value: object) -> Decimal:
    return Decimal(str(value)).quantize(_TWO)


async def _bind(session: AsyncSession, org_id: UUID, user_id: UUID) -> None:
    if session.in_transaction():
        await session.rollback()
    await session.begin()
    await bind_tenant_context(session, organization_id=org_id, user_id=user_id)


async def _seed_posted_june(session: AsyncSession) -> dict:
    """Seed fixture_loader June world, calculate, approve directly, and post."""
    fixture = load_june_fixture()
    assert len(fixture.employees) == 32

    if session.in_transaction():
        await session.rollback()

    org = await seed_organization(
        session,
        name=fixture.organization.name,
        slug=f"paybill-{uuid4().hex[:10]}",
    )
    user = await seed_user(session, workos_user_id=f"paybill_{uuid4().hex[:10]}")
    await session.commit()

    await _bind(session, org.id, user.id)

    office_ids: dict[str, UUID] = {}
    for office in fixture.organization.offices:
        row = Office(
            organization_id=org.id,
            name=office.name,
            jurisdiction=office.jurisdiction,
        )
        session.add(row)
        await session.flush()
        office_ids[office.fixture_id] = row.id

    post = Post(
        organization_id=org.id,
        designation="Synthetic Clerk",
        class_="III",
        sanctioned_strength=40,
        vacant_count=8,
        pay_scale="S-10",
        display_order=10,
    )
    session.add(post)
    await session.flush()

    component_ids: dict[str, UUID] = {}
    register_columns = {
        "BASIC": "basic_pay",
        "DA": "dearness_allowance",
        "CLA": "city_compensatory_allowance",
        "HRA": "house_rent_allowance",
        "TRANSPORT": "transport_pta_honorarium",
        "WASH_ALLOWANCE": "wash_child_other_charges",
        "OTHER_ALLOWANCE": "other_reimbursement_salary_increment_difference",
        "GPF_SUBSCRIPTION": "gpf_subscription_refund_arrears",
        "NPS_EMPLOYEE": "pension_employee_share",
        "NPS_EMPLOYER_TRANSFER": "pension_employer_share",
        "EPF_EMPLOYEE": "pension_employee_share",
        "EPF_EMPLOYER": "employer_share",
        "EPF_EMPLOYER_TRANSFER": "pension_employer_share",
        "INCOME_TAX": "income_tax",
        "PROFESSIONAL_TAX": "professional_tax",
        "GIS": "insurance",
        "HBA_INSTALLMENT": "advances",
        "ACCOMMODATION_LICENSE_FEE": "house_rent_service_charge_arrears",
    }
    display_order = 0
    for comp in fixture.components:
        display_order += 1
        api_cls = comp.api_classification
        if api_cls is None:
            assert comp.code == "FOREGONE_HRA"
            continue
        db_cls = _CLASSIFICATION_DB[comp.fixture_classification]
        component = PayComponent(
            organization_id=org.id,
            code=comp.code,
            name=comp.name,
            classification=db_cls,
            display_order=display_order,
            employer_transfer=comp.employer_transfer,
            transfer_of=comp.transfer_of,
            register_column=register_columns.get(comp.code),
        )
        session.add(component)
        await session.flush()
        component_ids[comp.code] = component.id

        if comp.code in RECURRING_COMPONENT_CODES or comp.code == BASIC_CODE:
            await versioning.insert_version(
                session,
                component_rate_versions,
                organization_id=org.id,
                header_id=component.id,
                effective_from=EFFECTIVE_FROM,
                values={
                    "calc_kind": "fixed_recurring_amount",
                    "amount": Decimal("0.00"),
                    "rate": None,
                    "basis": None,
                    "rounding_rule": "ROUND_HALF_UP_RUPEE",
                },
                change_reason=None,
                created_by=user.id,
            )

    employee_ids: dict[str, UUID] = {}
    for emp in fixture.employees:
        employee_id = await _seed_one_employee(
            session,
            org_id=org.id,
            user_id=user.id,
            employee=emp,
            office_ids=office_ids,
            post_id=post.id,
            component_ids=component_ids,
        )
        employee_ids[emp.fixture_id] = employee_id

    period = PayrollPeriod(
        organization_id=org.id,
        period_year=PERIOD_YEAR,
        period_month=PERIOD_MONTH,
        status="open",
    )
    session.add(period)
    await session.flush()
    run = PayrollRun(
        organization_id=org.id,
        period_id=period.id,
        status="draft",
    )
    session.add(run)
    await session.flush()
    session.add_all(
        initialize_run_roster(
            organization_id=org.id,
            run=run,
            employee_ids=list(employee_ids.values()),
            period_year=period.period_year,
            period_month=period.period_month,
        )
    )
    await session.commit()

    await _bind(session, org.id, user.id)
    calc = await calculate_run_command(
        session,
        organization_id=org.id,
        run_id=run.id,
        user_id=user.id,
    )

    await _bind(session, org.id, user.id)
    run_row = await session.get(PayrollRun, run.id)
    assert run_row is not None
    run_row.status = "approved"
    session.add(
        PayrollApproval(
            organization_id=org.id,
            run_id=run.id,
            run_version_id=calc["version_id"],
            content_hash=calc["content_hash"],
            action="approve",
            actor_user_id=user.id,
            reason="June golden approve",
        )
    )
    await session.commit()

    await _bind(session, org.id, user.id)
    posted = await post_run(
        session,
        organization_id=org.id,
        run_id=run.id,
        user_id=user.id,
    )
    assert posted["status"] == "posted"

    return {
        "org_id": org.id,
        "org_name": org.name,
        "user_id": user.id,
        "run_id": run.id,
        "version_id": UUID(str(calc["version_id"])),
        "employee_ids": employee_ids,
        "expected": fixture.expected.aggregates,
        "engine_version": calc.get("engine_version") or "0.1.0",
    }


async def _seed_one_employee(
    session: AsyncSession,
    *,
    org_id: UUID,
    user_id: UUID,
    employee: EmployeeSeed,
    office_ids: dict[str, UUID],
    post_id: UUID,
    component_ids: dict[str, UUID],
) -> UUID:
    basic = line_amount(employee, BASIC_CODE)
    assert basic is not None, f"{employee.fixture_id} missing BASIC"
    regime, gpf_jurisdiction = map_regime(employee.regime)

    header = Employee(organization_id=org_id, employee_number=employee.fixture_id)
    session.add(header)
    await session.flush()

    profile_values: dict = {
        "name": employee.name,
        "sevarth_id": employee.sevarth_id,
        "pan": employee.pan,
        "date_of_birth": date(1985, 1, 15),
        "date_of_joining": date(2010, 6, 1),
        "retirement_regime": regime,
        "gpf_jurisdiction": gpf_jurisdiction,
        "pran": None,
        "gpf_account_number": None,
        "epf_number": None,
        "pension_account": None,
    }
    if regime == "gpf":
        profile_values["gpf_account_number"] = employee.gpf_account
        if employee.pran:
            profile_values["pran"] = employee.pran
    elif regime == "nps":
        profile_values["pran"] = employee.pran or f"9000{employee.fixture_id[-4:].zfill(8)}"
    elif regime == "epf":
        profile_values["epf_number"] = employee.epf_number or f"SYNTEPF/{employee.fixture_id}/UAN"

    await versioning.insert_version(
        session,
        employee_profile_versions,
        organization_id=org_id,
        header_id=header.id,
        effective_from=EFFECTIVE_FROM,
        values=profile_values,
        change_reason=None,
        created_by=user_id,
    )
    await versioning.insert_version(
        session,
        employee_posting_versions,
        organization_id=org_id,
        header_id=header.id,
        effective_from=EFFECTIVE_FROM,
        values={
            "office_id": office_ids[employee.office_id],
            "post_id": post_id,
        },
        change_reason=None,
        created_by=user_id,
    )
    await versioning.insert_version(
        session,
        employee_pay_versions,
        organization_id=org_id,
        header_id=header.id,
        effective_from=EFFECTIVE_FROM,
        values={"pay_matrix_level": "L10", "basic_pay": money(basic).quantize(_TWO)},
        change_reason=None,
        created_by=user_id,
    )

    for line in employee.lines:
        code = line.component_code
        if code == BASIC_CODE:
            continue
        if code in RECURRING_COMPONENT_CODES:
            instruction = RecurringInstruction(
                organization_id=org_id,
                employee_id=header.id,
                component_id=component_ids[code],
            )
            session.add(instruction)
            await session.flush()
            await versioning.insert_version(
                session,
                recurring_instruction_versions,
                organization_id=org_id,
                header_id=instruction.id,
                effective_from=EFFECTIVE_FROM,
                values={
                    "amount": money(line.amount).quantize(_TWO),
                    "rate": None,
                    "reason": f"June fixture {code}",
                },
                change_reason=None,
                created_by=user_id,
            )
            continue
        if code in ADVANCE_COMPONENT_CODES:
            principal = max(line.amount * Decimal("24"), line.amount)
            advance = AdvanceAccount(
                organization_id=org_id,
                employee_id=header.id,
                advance_type="hba",
                principal=money(principal).quantize(_TWO),
                sanctioned_on=EFFECTIVE_FROM,
                reference=f"HBA-{employee.fixture_id}",
            )
            session.add(advance)
            await session.flush()
            await versioning.insert_version(
                session,
                advance_installment_versions,
                organization_id=org_id,
                header_id=advance.id,
                effective_from=EFFECTIVE_FROM,
                values={
                    "installment_amount": money(line.amount).quantize(_TWO),
                    "installments_total": 24,
                    "installments_recovered_opening": 0,
                },
                change_reason=None,
                created_by=user_id,
            )
            continue
        if code == "ACCOMMODATION_LICENSE_FEE":
            assert employee.accommodation is not None
            foregone = line_amount(employee, "FOREGONE_HRA")
            assignment = AccommodationAssignment(
                organization_id=org_id,
                employee_id=header.id,
                quarters_location=map_quarters_location(employee.accommodation.location),
                quarters_identifier=f"Q-{employee.fixture_id}",
            )
            session.add(assignment)
            await session.flush()
            await versioning.insert_version(
                session,
                accommodation_charge_versions,
                organization_id=org_id,
                header_id=assignment.id,
                effective_from=EFFECTIVE_FROM,
                values={
                    "license_fee": money(line.amount).quantize(_TWO),
                    "informational_hra_foregone": (
                        money(foregone).quantize(_TWO) if foregone is not None else None
                    ),
                },
                change_reason=None,
                created_by=user_id,
            )
            continue
        if code in ACCOMMODATION_COMPONENT_CODES:
            continue
        raise AssertionError(f"No seeding strategy for {employee.fixture_id} line {code}")

    return header.id


async def _june_world(session: AsyncSession) -> dict:
    global _CACHED_WORLD
    # Re-seed when the cached org was truncated away (singleton-org isolation).
    if _CACHED_WORLD is not None:
        await _bind(session, _CACHED_WORLD["org_id"], _CACHED_WORLD["user_id"])
        org = await session.get(Organization, _CACHED_WORLD["org_id"])
        if org is not None:
            return _CACHED_WORLD
        _CACHED_WORLD = None
    _CACHED_WORLD = await _seed_posted_june(session)
    return _CACHED_WORLD


def _ctx(
    world: dict,
    *,
    run_id: UUID | None = None,
    template_version: str = TEMPLATE_VERSION,
) -> ReportContext:
    return ReportContext(
        organization_id=world["org_id"],
        posted_run_id=run_id or world["run_id"],
        template_version=template_version,
        generated_at=datetime.now(UTC),
        engine_version=str(world["engine_version"]),
    )


def _section_by_title(dto, title: str):
    for section in dto.sections:
        if section.title == title:
            return section
    raise AssertionError(f"section {title!r} not found")


def _col_index(section, key: str) -> int:
    for idx, col in enumerate(section.columns):
        if col.key == key:
            return idx
    raise AssertionError(f"column {key!r} not found")


@pytest.mark.asyncio
async def test_pay_bill_footer_totals_and_headcount(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    dto = await pay_bill_builder.build(session, _ctx(world))

    assert dto.report_type == REPORT_TYPE_PAY_BILL
    register = _section_by_title(dto, "Register")
    assert len(register.rows) == 32
    assert register.totals is not None

    earnings_idx = _col_index(register, "earnings_total")
    deductions_idx = _col_index(register, "deductions_total")
    net_idx = _col_index(register, "net_payable")

    assert _dec(register.totals[earnings_idx]) == _dec("5073200.00")
    assert _dec(register.totals[deductions_idx]) == _dec("1264890.00")
    assert _dec(register.totals[net_idx]) == _dec("3838095.00")

    # Cross-check expected_totals.json aggregates.
    assert _dec(world["expected"]["salary_earnings"]) == _dec("5073200")
    assert _dec(world["expected"]["total_deductions"]) == _dec("1264890")
    assert _dec(world["expected"]["net_payable"]) == _dec("3838095")

    employee_numbers = [row[0] for row in register.rows]
    assert len(employee_numbers) == len(set(employee_numbers)) == 32


@pytest.mark.asyncio
async def test_pay_bill_rows_reconcile_to_db_results(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    dto = await pay_bill_builder.build(session, _ctx(world))
    register = _section_by_title(dto, "Register")

    db_rows = (
        (
            await session.execute(
                sa.select(payroll_employee_results).where(
                    payroll_employee_results.c.organization_id == world["org_id"],
                    payroll_employee_results.c.run_version_id == world["version_id"],
                )
            )
        )
        .mappings()
        .all()
    )
    by_number = {row["employee_number"]: row for row in db_rows}
    assert len(by_number) == 32

    earnings_idx = _col_index(register, "earnings_total")
    deductions_idx = _col_index(register, "deductions_total")
    net_idx = _col_index(register, "net_payable")

    for row in register.rows:
        emp_no = row[0]
        db = by_number[emp_no]
        assert _dec(row[earnings_idx]) == _dec(db["earnings_total"])
        assert _dec(row[deductions_idx]) == _dec(db["deductions_total"])
        assert _dec(row[net_idx]) == _dec(db["net_payable"])

        # Component columns reconcile to result lines.
        line_rows = (
            (
                await session.execute(
                    sa.select(payroll_result_lines).where(
                        payroll_result_lines.c.organization_id == world["org_id"],
                        payroll_result_lines.c.employee_result_id == db["id"],
                    )
                )
            )
            .mappings()
            .all()
        )
        amounts: dict[str, Decimal] = {}
        for line in line_rows:
            code = line["component_code"]
            if code == "FOREGONE_HRA":
                continue
            amounts[code] = amounts.get(code, _ZERO) + _dec(line["amount"])

        assert _dec(row[_col_index(register, "basic")]) == amounts.get("BASIC", _ZERO)
        assert _dec(row[_col_index(register, "gpf")]) == amounts.get("GPF_SUBSCRIPTION", _ZERO)
        assert _dec(row[_col_index(register, "pt")]) == amounts.get("PROFESSIONAL_TAX", _ZERO)
        assert _dec(row[_col_index(register, "hba")]) == amounts.get("HBA_INSTALLMENT", _ZERO)
        assert _dec(row[_col_index(register, "accommodation")]) == amounts.get(
            "ACCOMMODATION_LICENSE_FEE", _ZERO
        )
        transfers = amounts.get("NPS_EMPLOYER_TRANSFER", _ZERO) + amounts.get(
            "EPF_EMPLOYER_TRANSFER", _ZERO
        )
        assert _dec(row[_col_index(register, "transfers")]) == transfers


@pytest.mark.asyncio
async def test_pay_bill_v2_dynamic_columns_reconcile_and_excel_uses_formulas(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    dto = await pay_bill_builder.build(session, _ctx(world, template_version="v2"))
    register = _section_by_title(dto, "Register")

    keys = [column.key for column in register.columns]
    assert len(keys) == len(set(keys))
    assert "component:EPF_EMPLOYER" in keys
    assert "pan" in keys
    assert "gpf_account_number" in keys
    assert register.formulas

    earning_keys = [
        column.key
        for column in register.columns
        if column.key.startswith("component:")
        and column.key
        in {
            "component:BASIC",
            "component:DA",
            "component:HRA",
            "component:TRANSPORT",
            "component:OTHER_ALLOWANCE",
        }
    ]
    for row in register.rows:
        visible_earnings = sum(
            (row[_col_index(register, key)] for key in earning_keys),
            _ZERO,
        )
        assert _dec(visible_earnings) == _dec(row[_col_index(register, "earnings_total")])
        assert _dec(row[_col_index(register, "gross_bill")]) - _dec(
            row[_col_index(register, "deductions_total")]
        ) == _dec(row[_col_index(register, "net_payable")])

    workbook = load_workbook(BytesIO(pay_bill_to_excel(dto)), data_only=False)
    sheet = workbook["Register"]
    header_to_col = {sheet.cell(5, col).value: col for col in range(1, sheet.max_column + 1)}
    first_data_row = 6
    assert str(sheet.cell(first_data_row, header_to_col["Earnings Total"]).value).startswith("=")
    assert str(sheet.cell(first_data_row, header_to_col["Gross Bill"]).value).startswith("=")
    assert str(sheet.cell(first_data_row, header_to_col["Net Payable"]).value).startswith("=")


@pytest.mark.asyncio
async def test_pay_bill_v3_matches_canonical_contract_and_pdf_layout(session):
    contract_path = (
        Path(__file__).resolve().parents[3]
        / "fixtures/sanitized/june-2026/canonical_export_contract.json"
    )
    full_contract = json.loads(contract_path.read_text())
    contract = full_contract["pay_bill"]
    sheet_contract = next(item for item in full_contract["sheets"] if item["name"] == "Pay Bill")
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    dto = await pay_bill_builder.build(session, _ctx(world, template_version="v3"))
    register = _section_by_title(dto, "Register")
    assert len(register.rows) == 32

    workbook = load_workbook(BytesIO(pay_bill_to_excel(dto)), data_only=False)
    assert workbook.sheetnames == ["Pay Bill"]
    sheet = workbook["Pay Bill"]
    merged = {str(item) for item in sheet.merged_cells.ranges}
    assert set(contract["header_groups"].values()) <= merged
    assert sheet["A3"].value == "Sr. No."
    assert sheet["B3"].value == "Employee Name"
    assert sheet["M3"].value == "Festival Advance / Other Recovery"
    assert "Adjuesable" not in " ".join(
        str(sheet.cell(row=3, column=column).value or "") for column in range(1, 29)
    )
    assert [sheet.cell(8, column).value for column in range(1, 29)] == list(range(1, 29))
    assert sheet["B9"].value == ("Post of Synthetic Clerk (Total Posts 40. Vacant 8) - Scale S-10")
    assert sheet["K15"].value == "=SUM(C15:J15)"
    assert sheet["N15"].value == "=K15+L15-M15"
    assert sheet["Z15"].value == "=SUM(P15:Y15)"
    assert sheet["AA15"].value == "=N15-Z15"
    assert sheet.print_title_rows == "$2:$7"
    assert sheet.page_setup.orientation == "landscape"
    assert str(sheet.page_setup.paperSize) == sheet.PAPERSIZE_A4
    assert sheet.page_setup.scale == 33
    assert sheet.print_area.startswith("'Pay Bill'!$A$1:$AB$")
    assert int(sheet.print_area.rsplit("$", 1)[-1]) >= 208
    assert [item.id for item in sheet.row_breaks.brk] == sheet_contract["manual_row_breaks"]
    assert [item.id for item in sheet.col_breaks.brk] == sheet_contract["manual_column_breaks"]
    expected_widths = {
        column: float(item["width"])
        for item in sheet_contract["column_dimensions"]
        if item["min"] <= 28 and "width" in item
        for column in range(item["min"], min(item["max"], 28) + 1)
    }
    assert {
        column: sheet.column_dimensions[get_column_letter(column)].width for column in range(1, 29)
    } == expected_widths

    post_index = _col_index(register, "post_title")
    post_group_index = _col_index(register, "post_group_key")
    designation_index = _col_index(register, "designation")
    planned_rows = []
    post_groups = ("Post 1", "Post 2", "Post 3", "Post 4", "Post 5", "Post 6")
    for index, source_row in enumerate(register.rows[:28], start=1):
        values = list(source_row)
        group_index = (
            0
            if index == 1
            else 1
            if index == 2
            else 2
            if index < 16
            else 3
            if index < 25
            else 4
            if index == 25
            else 5
        )
        values[post_index] = post_groups[group_index]
        values[post_group_index] = f"group-{group_index}"
        if index in {24, 25}:
            values[designation_index] = "Accounts Officer"
        if group_index == 5:
            # Distinct group IDs can intentionally share a printed heading.
            values[post_index] = "Post 5"
        planned_rows.append(tuple(values))
    employee_numbers = {str(row[_col_index(register, "employee_number")]) for row in planned_rows}
    detail = _section_by_title(dto, "Component detail lines")
    detail_employee_index = _col_index(detail, "employee_number")
    detail_reason_index = _col_index(detail, "reason")
    assert any(str(row[detail_reason_index]).startswith("June fixture ") for row in detail.rows)
    employee_25_number = str(planned_rows[24][_col_index(register, "employee_number")])
    planned_detail_rows = []
    narration_added = False
    for source_row in detail.rows:
        if str(source_row[detail_employee_index]) not in employee_numbers:
            continue
        values = list(source_row)
        if str(values[detail_employee_index]) == employee_25_number and not narration_added:
            values[detail_reason_index] = "June fixture recovery"
            narration_added = True
        planned_detail_rows.append(tuple(values))
    planned_dto = ReportDTO(
        report_type=dto.report_type,
        template_version=dto.template_version,
        title=dto.title,
        organization_name=dto.organization_name,
        subtitle=dto.subtitle,
        sections=(
            TableSection(register.title, register.columns, tuple(planned_rows)),
            TableSection(
                detail.title,
                detail.columns,
                tuple(planned_detail_rows),
            ),
        ),
        metadata={
            "report_profile": {
                "legal_name": "Canonical Legal Organization",
                "office_name": "Canonical Payroll Office",
            },
            "run_metadata": {"payment_date": "2026-07-01"},
        },
    )
    planned_sheet = load_workbook(BytesIO(pay_bill_v3_to_excel(planned_dto)), data_only=False)[
        "Pay Bill"
    ]
    for serial, start, total in (
        (1, 10, 15),
        (8, 55, 60),
        (9, 68, 73),
        (16, 111, 116),
        (18, 123, 128),
        (19, 135, 140),
        (24, 165, 170),
        (26, 179, 184),
        (28, 191, 196),
    ):
        assert planned_sheet.cell(start, 1).value == serial
        assert planned_sheet.cell(total, 2).value == "Total Rs."
    assert planned_sheet["A171"].value == 25
    assert planned_sheet["A172"].value is None
    assert "June fixture" in str(planned_sheet["B174"].value)
    assert str(planned_sheet["B175"].value).startswith("Basic @ Rs.")
    planned_merges = {str(item) for item in planned_sheet.merged_cells.ranges}
    assert planned_merges == set(sheet_contract["merged_cells"])
    assert {
        "A25:A30",
        "A31:A36",
        "A37:A42",
        "A111:A116",
        "A165:A169",
        "A171:A177",
        "A179:A183",
        "A62:A67",
        "B62:B66",
        "A129:A133",
        "A197:A201",
        "B197:B201",
        "A203:A207",
        "B203:B207",
        "AB57:AB59",
        "AB76:AB78",
        "V211:Y211",
    } <= planned_merges
    assert {"B9:AB9", "B17:AB17", "B24:AB24"} <= planned_merges
    assert {"B110:AA110", "B171:AA171", "B178:AA178"} <= planned_merges
    assert str(planned_sheet["B171"].value).startswith("Post of Post 5")
    assert str(planned_sheet["B178"].value).startswith("Post of Post 5")
    assert planned_sheet["B197"].value == "Total of Page No. 3"
    assert planned_sheet["B203"].value == "Total of All Pages"
    assert planned_sheet["B208"].value == "Grand Total Rs."
    assert planned_sheet["C177"].value == "=SUM(C171:C176)"
    assert "C171" in planned_sheet["C197"].value
    assert "C172" in planned_sheet["C198"].value
    assert planned_sheet["O1"].value == datetime(2026, 6, 1)
    assert planned_sheet["S1"].value == datetime(2026, 7, 1)
    assert planned_sheet.page_margins.left == pytest.approx(0.07874015748031496)
    assert planned_sheet.page_margins.right == pytest.approx(0.07874015748031496)
    assert "Canonical Legal Organization" in planned_sheet.oddFooter.left.text
    validator_path = Path(__file__).resolve().parents[3] / "scripts/validate_canonical_export.py"
    validator_spec = importlib.util.spec_from_file_location("pay_bill_validator", validator_path)
    assert validator_spec is not None and validator_spec.loader is not None
    validator = importlib.util.module_from_spec(validator_spec)
    validator_spec.loader.exec_module(validator)
    semantic_issues = validator.validate_pay_bill_semantics(
        planned_sheet,
        full_contract,
        require_june_totals=False,
    )
    assert semantic_issues == []

    pdf = PdfReader(BytesIO(pay_bill_to_pdf(planned_dto)))
    assert len(pdf.pages) > 1
    for page in pdf.pages:
        assert float(page.mediabox.width) > float(page.mediabox.height)
        assert round(float(page.mediabox.width)) == 792
        assert round(float(page.mediabox.height)) == 612
        text = page.extract_text() or ""
        assert "Adjustable by Accountant General" in text
        assert "Adjustable by Treasury" in text
        assert "Employee Name" in text
        assert "Canonical Legal Organization" in text

    unsafe_profile_dto = replace(
        planned_dto,
        metadata={
            **planned_dto.metadata,
            "report_profile": {"legal_name": "=1+1"},
        },
    )
    unsafe_sheet = load_workbook(
        BytesIO(pay_bill_v3_to_excel(unsafe_profile_dto)), data_only=False
    )["Pay Bill"]
    assert unsafe_sheet["V211"].value == "'=1+1"
    assert unsafe_sheet["V211"].data_type != "f"


@pytest.mark.asyncio
async def test_pay_bill_v3_rejects_nonzero_component_without_explicit_column(session, monkeypatch):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    snapshot_row = (
        (
            await session.execute(
                sa.select(payroll_report_snapshots).where(
                    payroll_report_snapshots.c.organization_id == world["org_id"],
                    payroll_report_snapshots.c.run_version_id == world["version_id"],
                )
            )
        )
        .mappings()
        .one()
    )
    changed = json.loads(json.dumps(snapshot_row["snapshot"]))
    basic = next(item for item in changed["component_catalog"] if item["code"] == "BASIC")
    basic["register_column"] = None

    async def load_changed_snapshot(*_args, **_kwargs):
        return changed

    monkeypatch.setattr(
        "app.reports.families.payroll_register.load_report_snapshot",
        load_changed_snapshot,
    )
    with pytest.raises(ConflictError, match="not mapped to a canonical Pay Bill column"):
        await pay_bill_builder.build(session, _ctx(world, template_version="v3"))


@pytest.mark.asyncio
async def test_pay_bill_amount_in_words_matches_numeric(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    dto = await pay_bill_builder.build(session, _ctx(world))
    register = _section_by_title(dto, "Register")
    words_section = _section_by_title(dto, "Amount in words")
    net = _dec(register.totals[_col_index(register, "net_payable")])
    assert words_section.rows[0][1] == amount_in_words(net)


@pytest.mark.asyncio
async def test_treasury_face_totals_and_reconciliation(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    dto = await treasury_face_builder.build(session, _ctx(world))

    assert dto.report_type == REPORT_TYPE_TREASURY_FACE
    summary = _section_by_title(dto, "Treasury Face Summary")
    by_label = {row[0]: _dec(row[1]) for row in summary.rows}

    assert by_label["Gross bill"] == _dec("5102985.00")
    assert by_label["Employer share"] == _dec("29785.00")
    assert by_label["Total deductions"] == _dec("1264890.00")
    assert by_label["Net payable"] == _dec("3838095.00")

    assert by_label["Gross bill"] - by_label["Total deductions"] == by_label["Net payable"]
    assert (
        by_label["AG deductions"]
        + by_label["Treasury deductions"]
        + by_label["External recoveries"]
        == by_label["Total deductions"]
    )

    words = _section_by_title(dto, "Amount in words")
    words_by_label = {row[0]: row[1] for row in words.rows}
    assert words_by_label["Gross bill"] == amount_in_words(by_label["Gross bill"])
    assert words_by_label["Net payable"] == amount_in_words(by_label["Net payable"])


async def _seed_minimal_posted_run_with_gross_adjustment(session: AsyncSession) -> dict:
    """Minimal posted run: one employee with earning + gross_adjustment + deduction.

    Inserted directly (posting-shaped rows) to prove Treasury Face includes
    gross_adjustment lines in Gross bill — the June fixture has none.
    """
    org = await seed_organization(session, name="Gross Adj Org", slug=f"ga-{uuid4().hex[:10]}")
    user = await seed_user(session, workos_user_id=f"ga_{uuid4().hex[:10]}")
    await session.commit()
    await _bind(session, org.id, user.id)

    employee = Employee(organization_id=org.id, employee_number="GA-01")
    session.add(employee)
    period = PayrollPeriod(
        organization_id=org.id,
        period_year=PERIOD_YEAR,
        period_month=PERIOD_MONTH,
        status="open",
    )
    session.add(period)
    await session.flush()
    run = PayrollRun(organization_id=org.id, period_id=period.id, status="posted")
    session.add(run)
    await session.flush()

    version_id = (
        await session.execute(
            sa.insert(payroll_run_versions)
            .values(
                organization_id=org.id,
                run_id=run.id,
                version_number=1,
                engine_version="0.1.0",
                content_hash="ga-test-hash",
                calculated_at=datetime.now(UTC),
                calculated_by=user.id,
                inputs_snapshot={},
                totals={
                    "earnings_total": "1000.00",
                    "employer_contribution_total": "0.00",
                    "gross_adjustment_total": "50.00",
                    "gross_total": "1050.00",
                    "deductions_total": "125.00",
                    "net_payable": "925.00",
                    "offbill_employer_remittance": "0.00",
                    "disbursement": "925.00",
                },
            )
            .returning(payroll_run_versions.c.id)
        )
    ).scalar_one()
    run.current_version_id = version_id

    result_id = (
        await session.execute(
            sa.insert(payroll_employee_results)
            .values(
                organization_id=org.id,
                run_version_id=version_id,
                employee_id=employee.id,
                employee_number="GA-01",
                earnings_total=Decimal("1000.00"),
                employer_contribution_total=Decimal("0.00"),
                gross_total=Decimal("1050.00"),
                deductions_total=Decimal("125.00"),
                net_payable=Decimal("925.00"),
                offbill_employer_remittance=Decimal("0.00"),
                disbursement=Decimal("925.00"),
            )
            .returning(payroll_employee_results.c.id)
        )
    ).scalar_one()

    lines = (
        ("BASIC", "earning", "earning", Decimal("1000.00"), 1),
        ("DA_DIFFERENCE", "gross_adjustment", "gross_adjustment", Decimal("50.00"), 2),
        ("INCOME_TAX", "treasury_deduction", "treasury_deduction", Decimal("100.00"), 3),
        ("RECOVERY", "external_recovery", "external_recovery", Decimal("25.00"), 4),
    )
    for code, db_class, trace_class, amount, sequence in lines:
        await session.execute(
            sa.insert(payroll_result_lines).values(
                organization_id=org.id,
                employee_result_id=result_id,
                component_code=code,
                classification=db_class,
                calc_kind="fixed_recurring_amount",
                amount=amount,
                sequence=sequence,
                trace={"classification": trace_class},
            )
        )
    await session.commit()
    return {"org_id": org.id, "user_id": user.id, "run_id": run.id, "engine_version": "0.1.0"}


@pytest.mark.asyncio
async def test_treasury_face_includes_gross_adjustment_in_gross_bill(session):
    """Regression: gross_adjustment lines must be part of Gross bill so that
    gross − deductions = posted net payable (engine identity, ADR 0007)."""
    world = await _seed_minimal_posted_run_with_gross_adjustment(session)
    await _bind(session, world["org_id"], world["user_id"])
    dto = await treasury_face_builder.build(session, _ctx(world))

    summary = _section_by_title(dto, "Treasury Face Summary")
    by_label = {row[0]: _dec(row[1]) for row in summary.rows}
    assert by_label["Gross bill"] == _dec("1050.00")
    assert by_label["Gross adjustments (in gross bill)"] == _dec("50.00")
    assert by_label["Total deductions"] == _dec("125.00")
    assert by_label["Net payable"] == _dec("925.00")
    assert by_label["Gross bill"] - by_label["Total deductions"] == by_label["Net payable"]


@pytest.mark.asyncio
async def test_pay_bill_v3_reconciles_gross_adjustment_and_recovery(session, monkeypatch):
    world = await _seed_minimal_posted_run_with_gross_adjustment(session)
    await _bind(session, world["org_id"], world["user_id"])

    async def snapshot(*_args, **_kwargs):
        return {
            "organization": {"name": "Gross Adj Org"},
            "report_profile": {},
            "run_metadata": {},
            "employee_identity": {
                # The test seed has exactly one employee; obtain its UUID from result rows below.
            },
            "component_catalog": [
                {"code": "BASIC", "name": "Basic", "register_column": "C", "display_order": 1},
                {
                    "code": "DA_DIFFERENCE",
                    "name": "DA Difference",
                    "register_column": "D",
                    "display_order": 2,
                },
                {
                    "code": "INCOME_TAX",
                    "name": "Income Tax",
                    "register_column": "U",
                    "display_order": 3,
                },
                {
                    "code": "RECOVERY",
                    "name": "Recovery",
                    "register_column": "M",
                    "display_order": 4,
                },
            ],
        }

    result = (
        (
            await session.execute(
                sa.select(payroll_employee_results).where(
                    payroll_employee_results.c.organization_id == world["org_id"]
                )
            )
        )
        .mappings()
        .one()
    )
    packed_snapshot = await snapshot()
    packed_snapshot["employee_identity"][str(result["employee_id"])] = {
        "name": "Gross Adjustment Employee",
        "designation": "Officer",
        "post": {"id": "officer-1", "title": "Officer", "display_order": 1},
    }

    async def load_snapshot(*_args, **_kwargs):
        return packed_snapshot

    monkeypatch.setattr("app.reports.families.payroll_register.load_report_snapshot", load_snapshot)
    dto = await pay_bill_builder.build(session, _ctx(world, template_version="v3"))
    register = _section_by_title(dto, "Register")
    row = register.rows[0]
    assert _dec(row[_col_index(register, "d_da")]) == _dec("50.00")
    assert _dec(row[_col_index(register, "m_recovery")]) == _dec("25.00")
    assert _dec(row[_col_index(register, "u_income_tax")]) == _dec("100.00")

    face_dto = await treasury_face_builder.build(session, _ctx(world, template_version="v3"))
    from app.reports.canonical_front_sheets import treasury_face_to_excel as render_face

    standalone = load_workbook(BytesIO(render_face(face_dto)), data_only=False)[" Face "]
    consolidated = load_workbook(BytesIO(render_face(face_dto, dto)), data_only=False)[" Face "]
    for cell in (
        "I18",
        "I19",
        "I20",
        "I25",
        "I31",
        "I38",
        "I41",
        "I42",
        "I44",
        "I45",
        "I46",
        "I47",
    ):
        assert standalone[cell].value == consolidated[cell].value


@pytest.mark.asyncio
async def test_pay_bill_v3_detail_order_preserves_catalog_zero(session, monkeypatch):
    world = await _seed_minimal_posted_run_with_gross_adjustment(session)
    await _bind(session, world["org_id"], world["user_id"])
    result = (
        (
            await session.execute(
                sa.select(payroll_employee_results).where(
                    payroll_employee_results.c.organization_id == world["org_id"]
                )
            )
        )
        .mappings()
        .one()
    )
    snapshot = {
        "organization": {"name": "Detail Order Org"},
        "report_profile": {},
        "run_metadata": {},
        "employee_identity": {
            str(result["employee_id"]): {
                "name": "Detail Order Employee",
                "designation": "Officer",
                "post": {"id": "officer-1", "title": "Officer", "display_order": 1},
            }
        },
        "component_catalog": [
            {"code": "BASIC", "name": "Basic", "register_column": "C", "display_order": 1},
            {
                "code": "DA_DIFFERENCE",
                "name": "DA Difference",
                "register_column": "D",
                "display_order": 2,
            },
            {
                "code": "INCOME_TAX",
                "name": "Income Tax",
                "register_column": "M",
                "display_order": 0,
            },
            {
                "code": "RECOVERY",
                "name": "Recovery",
                "register_column": "M",
                "display_order": 10,
            },
        ],
    }

    async def load_snapshot(*_args, **_kwargs):
        return snapshot

    monkeypatch.setattr("app.reports.families.payroll_register.load_report_snapshot", load_snapshot)
    dto = await pay_bill_builder.build(session, _ctx(world, template_version="v3"))
    detail = _section_by_title(dto, "Component detail lines")
    bucket_index = _col_index(detail, "register_column")
    code_index = _col_index(detail, "component_code")
    order_index = _col_index(detail, "display_order")
    recovery_lines = [row for row in detail.rows if row[bucket_index] == "m_recovery"]

    assert [(row[code_index], row[order_index]) for row in recovery_lines] == [
        ("INCOME_TAX", 0),
        ("RECOVERY", 10),
    ]


@pytest.mark.asyncio
async def test_treasury_face_signatories_from_report_configurations(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])

    dto_empty = await treasury_face_builder.build(session, _ctx(world))
    assert _section_by_title(dto_empty, "Signatories").rows == ()

    session.add(
        ReportConfiguration(
            organization_id=world["org_id"],
            key="signatories",
            value={"chair": "Director", "dda": "Deputy Director"},
        )
    )
    await session.commit()

    await _bind(session, world["org_id"], world["user_id"])
    dto = await treasury_face_builder.build(session, _ctx(world))
    signatories = _section_by_title(dto, "Signatories")
    assert ("chair", "Director") in signatories.rows
    assert ("dda", "Deputy Director") in signatories.rows

    await session.execute(
        sa.delete(ReportConfiguration).where(
            ReportConfiguration.organization_id == world["org_id"],
            ReportConfiguration.key == "signatories",
        )
    )
    await session.commit()


@pytest.mark.asyncio
async def test_pay_bill_excel_and_pdf_formatters(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    dto = await pay_bill_builder.build(session, _ctx(world))

    payload = pay_bill_to_json(dto)
    assert payload["report_type"] == REPORT_TYPE_PAY_BILL
    assert payload["sections"][0]["totals"]["net_payable"] == "3838095.00"

    xlsx = pay_bill_to_excel(dto)
    wb = load_workbook(BytesIO(xlsx))
    ws = wb.active
    assert ws is not None
    money_cell = ws.cell(row=6, column=4)
    assert money_cell.number_format == MONEY_FORMAT
    register = _section_by_title(dto, "Register")
    totals_excel_row = 6 + len(register.rows)
    net_col = len(register.columns)
    net_cell = ws.cell(row=totals_excel_row, column=net_col)
    assert net_cell.number_format == MONEY_FORMAT
    assert _dec(net_cell.value) == _dec("3838095.00")

    pdf = pay_bill_to_pdf(dto)
    assert pdf.startswith(b"%PDF")
    text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(pdf)).pages)
    assert "Payroll Register — Pay Bill" in text or "Payroll Register" in text
    assert "38,38,095.00" in text


@pytest.mark.asyncio
async def test_treasury_face_excel_and_pdf_formatters(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    dto = await treasury_face_builder.build(session, _ctx(world))

    payload = treasury_face_to_json(dto)
    assert payload["report_type"] == REPORT_TYPE_TREASURY_FACE

    xlsx = treasury_face_to_excel(dto)
    wb = load_workbook(BytesIO(xlsx))
    ws = wb.active
    assert ws is not None
    money_cell = ws.cell(row=6, column=2)
    assert money_cell.number_format == MONEY_FORMAT
    assert _dec(money_cell.value) == _dec("5102985.00")

    pdf = treasury_face_to_pdf(dto)
    assert pdf.startswith(b"%PDF")
    text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(pdf)).pages)
    assert "Treasury Face" in text
    assert "38,38,095.00" in text


@pytest.mark.asyncio
async def test_unposted_run_raises_conflict(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])

    period = PayrollPeriod(
        organization_id=world["org_id"],
        period_year=2026,
        period_month=7,
        status="open",
    )
    session.add(period)
    await session.flush()
    draft = PayrollRun(
        organization_id=world["org_id"],
        period_id=period.id,
        status="draft",
    )
    session.add(draft)
    await session.flush()
    draft_id = draft.id
    await session.commit()

    await _bind(session, world["org_id"], world["user_id"])
    with pytest.raises(ConflictError, match="must be posted"):
        await pay_bill_builder.build(session, _ctx(world, run_id=draft_id))

    await _bind(session, world["org_id"], world["user_id"])
    with pytest.raises(ConflictError, match="must be posted"):
        await treasury_face_builder.build(session, _ctx(world, run_id=draft_id))
