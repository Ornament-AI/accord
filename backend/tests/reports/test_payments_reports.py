"""Golden tests for Bank/RTGS advice and payslip builders/formatters."""

from __future__ import annotations

from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from uuid import UUID, uuid4

import pytest
import sqlalchemy as sa
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy.ext.asyncio import AsyncSession

from app.exceptions import ConflictError
from app.models.accommodation import AccommodationAssignment, accommodation_charge_versions
from app.models.advances import AdvanceAccount, advance_installment_versions
from app.models.employees import (
    Employee,
    employee_bank_account_versions,
    employee_pay_versions,
    employee_posting_versions,
    employee_profile_versions,
)
from app.models.org_structure import Office, PayrollUnit, Post
from app.models.pay_components import PayComponent, component_rate_versions
from app.models.payroll_runs import (
    PayrollPeriod,
    PayrollRun,
    payroll_employee_results,
    payroll_result_lines,
)
from app.models.platform import PayrollApproval
from app.models.recurring_instructions import (
    RecurringInstruction,
    recurring_instruction_versions,
)
from app.reports.amount_in_words import amount_in_words
from app.reports.base import ReportContext, ReportRegistry
from app.reports.excel import MONEY_FORMAT
from app.reports.formatting import format_inr
from app.reports.families.payments import (
    REPORT_TYPE_BANK_ADVICE,
    REPORT_TYPE_PAYSLIPS,
    MissingPrimarySalaryAccountError,
    bank_advice_builder,
    bank_advice_to_excel,
    bank_advice_to_json,
    payslip_bundle_builder,
    payslip_to_excel,
    payslip_to_json,
    payslip_to_pdf,
    register_payment_reports,
)
from app.schemas.employees import mask_value
from app.services import versioning
from app.services.run_calculation import calculate_run_command
from app.services.run_posting import post_run
from app.tenancy import bind_tenant_context
from tests.e2e.fixture_loader import (
    ACCOMMODATION_COMPONENT_CODES,
    ADVANCE_COMPONENT_CODES,
    BASIC_CODE,
    RECURRING_COMPONENT_CODES,
    EmployeeSeed,
    line_amount,
    load_june_fixture,
    map_quarters_location,
    map_regime,
    money,
)
from tests.identity_helpers import seed_organization, seed_user

EFFECTIVE_FROM = date(2026, 1, 1)
PERIOD_YEAR = 2026
PERIOD_MONTH = 6
TEMPLATE_VERSION = "v1"
EXPECTED_NET = Decimal("3838095.00")

_CLASSIFICATION_DB = {
    "earning": "earning",
    "employer_contribution": "employer_contribution",
    "AG_deduction": "ag_deduction",
    "treasury_deduction": "treasury_deduction",
    "gross_adjustment": "gross_adjustment",
    "external_recovery": "external_recovery",
}

_TWO = Decimal("0.01")
_ZERO = Decimal("0.00")

# Module-level cache so the expensive June seed runs once per process.
_CACHED_WORLD: dict | None = None


def _dec(value: object) -> Decimal:
    return Decimal(str(value)).quantize(_TWO)


async def _bind(session: AsyncSession, org_id: UUID, user_id: UUID) -> None:
    if session.in_transaction():
        await session.rollback()
    await session.begin()
    await bind_tenant_context(session, organization_id=org_id, user_id=user_id)


async def _seed_posted_june(session: AsyncSession) -> dict:
    """Seed fixture_loader June world with bank accounts, calculate, approve, post."""
    fixture = load_june_fixture()
    assert len(fixture.employees) == 32

    if session.in_transaction():
        await session.rollback()

    org = await seed_organization(
        session,
        name=fixture.organization.name,
        slug=f"payments-{uuid4().hex[:10]}",
    )
    user = await seed_user(session, workos_user_id=f"payments_{uuid4().hex[:10]}")
    await session.commit()

    await _bind(session, org.id, user.id)

    office_ids: dict[str, UUID] = {}
    for office in fixture.organization.offices:
        row = Office(
            organization_id=org.id,
            name=office.name,
            code=office.code,
            jurisdiction=office.jurisdiction,
        )
        session.add(row)
        await session.flush()
        office_ids[office.fixture_id] = row.id

    unit = PayrollUnit(
        organization_id=org.id,
        name=fixture.organization.pay_unit_name,
        code=fixture.organization.pay_unit_code,
    )
    session.add(unit)
    await session.flush()

    post = Post(
        organization_id=org.id,
        designation="Synthetic Clerk",
        class_="III",
    )
    session.add(post)
    await session.flush()

    component_ids: dict[str, UUID] = {}
    display_order = 0
    for comp in fixture.components:
        display_order += 1
        api_cls = comp.api_classification
        if api_cls is None:
            assert comp.code == "FOREGONE_HRA"
            continue
        db_cls = _CLASSIFICATION_DB[comp.fixture_classification]
        component = PayComponent(
            organization_id=org.id,
            code=comp.code,
            name=comp.name,
            classification=db_cls,
            display_order=display_order,
        )
        session.add(component)
        await session.flush()
        component_ids[comp.code] = component.id

        if comp.code in RECURRING_COMPONENT_CODES or comp.code == BASIC_CODE:
            await versioning.insert_version(
                session,
                component_rate_versions,
                organization_id=org.id,
                header_id=component.id,
                effective_from=EFFECTIVE_FROM,
                values={
                    "calc_kind": "fixed_recurring_amount",
                    "amount": Decimal("0.00"),
                    "rate": None,
                    "basis": None,
                    "rounding_rule": "ROUND_HALF_UP_RUPEE",
                },
                change_reason=None,
                created_by=user.id,
            )

    employee_ids: dict[str, UUID] = {}
    bank_by_fixture: dict[str, str] = {}
    for emp in fixture.employees:
        employee_id = await _seed_one_employee(
            session,
            org_id=org.id,
            user_id=user.id,
            employee=emp,
            office_ids=office_ids,
            unit_id=unit.id,
            post_id=post.id,
            component_ids=component_ids,
            seed_bank=True,
        )
        employee_ids[emp.fixture_id] = employee_id
        bank_by_fixture[emp.fixture_id] = emp.bank_account

    period = PayrollPeriod(
        organization_id=org.id,
        period_year=PERIOD_YEAR,
        period_month=PERIOD_MONTH,
        status="open",
    )
    session.add(period)
    await session.flush()
    run = PayrollRun(
        organization_id=org.id,
        period_id=period.id,
        run_type="regular",
        status="draft",
    )
    session.add(run)
    await session.commit()

    await _bind(session, org.id, user.id)
    calc = await calculate_run_command(
        session,
        organization_id=org.id,
        run_id=run.id,
        user_id=user.id,
    )

    await _bind(session, org.id, user.id)
    run_row = await session.get(PayrollRun, run.id)
    assert run_row is not None
    run_row.status = "approved"
    session.add(
        PayrollApproval(
            organization_id=org.id,
            run_id=run.id,
            run_version_id=calc["version_id"],
            content_hash=calc["content_hash"],
            action="approve",
            actor_user_id=user.id,
            reason="June golden approve",
        )
    )
    await session.commit()

    await _bind(session, org.id, user.id)
    posted = await post_run(
        session,
        organization_id=org.id,
        run_id=run.id,
        user_id=user.id,
    )
    assert posted["status"] == "posted"

    return {
        "org_id": org.id,
        "org_name": org.name,
        "user_id": user.id,
        "run_id": run.id,
        "version_id": UUID(str(calc["version_id"])),
        "employee_ids": employee_ids,
        "bank_by_fixture": bank_by_fixture,
        "fixture_employees": {e.fixture_id: e for e in fixture.employees},
        "expected": fixture.expected.aggregates,
        "engine_version": calc.get("engine_version") or "0.1.0",
    }


async def _seed_one_employee(
    session: AsyncSession,
    *,
    org_id: UUID,
    user_id: UUID,
    employee: EmployeeSeed,
    office_ids: dict[str, UUID],
    unit_id: UUID,
    post_id: UUID,
    component_ids: dict[str, UUID],
    seed_bank: bool = True,
) -> UUID:
    basic = line_amount(employee, BASIC_CODE)
    assert basic is not None, f"{employee.fixture_id} missing BASIC"
    regime, gpf_jurisdiction = map_regime(employee.regime)

    header = Employee(organization_id=org_id, employee_number=employee.fixture_id)
    session.add(header)
    await session.flush()

    profile_values: dict = {
        "name": employee.name,
        "sevarth_id": employee.sevarth_id,
        "pan": employee.pan,
        "date_of_birth": date(1985, 1, 15),
        "date_of_joining": date(2010, 6, 1),
        "retirement_regime": regime,
        "gpf_jurisdiction": gpf_jurisdiction,
        "pran": None,
        "gpf_account_number": None,
        "epf_number": None,
        "pension_account": None,
    }
    if regime == "gpf":
        profile_values["gpf_account_number"] = employee.gpf_account
        if employee.pran:
            profile_values["pran"] = employee.pran
    elif regime == "nps":
        profile_values["pran"] = employee.pran or f"9000{employee.fixture_id[-4:].zfill(8)}"
    elif regime == "epf":
        profile_values["epf_number"] = employee.epf_number or f"SYNTEPF/{employee.fixture_id}/UAN"

    await versioning.insert_version(
        session,
        employee_profile_versions,
        organization_id=org_id,
        header_id=header.id,
        effective_from=EFFECTIVE_FROM,
        values=profile_values,
        change_reason=None,
        created_by=user_id,
    )
    await versioning.insert_version(
        session,
        employee_posting_versions,
        organization_id=org_id,
        header_id=header.id,
        effective_from=EFFECTIVE_FROM,
        values={
            "office_id": office_ids[employee.office_id],
            "payroll_unit_id": unit_id,
            "post_id": post_id,
            "employee_group_id": None,
        },
        change_reason=None,
        created_by=user_id,
    )
    await versioning.insert_version(
        session,
        employee_pay_versions,
        organization_id=org_id,
        header_id=header.id,
        effective_from=EFFECTIVE_FROM,
        values={"pay_matrix_level": "L10", "basic_pay": money(basic).quantize(_TWO)},
        change_reason=None,
        created_by=user_id,
    )

    if seed_bank:
        await versioning.insert_version(
            session,
            employee_bank_account_versions,
            organization_id=org_id,
            header_id=header.id,
            effective_from=EFFECTIVE_FROM,
            values={
                "account_number": employee.bank_account,
                "ifsc": employee.ifsc,
                "bank_name": "Synthetic Bank",
                "branch": "Synthetic Branch",
                "is_primary_salary": True,
            },
            change_reason=None,
            created_by=user_id,
        )

    for line in employee.lines:
        code = line.component_code
        if code == BASIC_CODE:
            continue
        if code in RECURRING_COMPONENT_CODES:
            instruction = RecurringInstruction(
                organization_id=org_id,
                employee_id=header.id,
                component_id=component_ids[code],
            )
            session.add(instruction)
            await session.flush()
            await versioning.insert_version(
                session,
                recurring_instruction_versions,
                organization_id=org_id,
                header_id=instruction.id,
                effective_from=EFFECTIVE_FROM,
                values={
                    "amount": money(line.amount).quantize(_TWO),
                    "rate": None,
                    "reason": f"June fixture {code}",
                },
                change_reason=None,
                created_by=user_id,
            )
            continue
        if code in ADVANCE_COMPONENT_CODES:
            principal = max(line.amount * Decimal("24"), line.amount)
            advance = AdvanceAccount(
                organization_id=org_id,
                employee_id=header.id,
                advance_type="hba",
                principal=money(principal).quantize(_TWO),
                sanctioned_on=EFFECTIVE_FROM,
                reference=f"HBA-{employee.fixture_id}",
            )
            session.add(advance)
            await session.flush()
            await versioning.insert_version(
                session,
                advance_installment_versions,
                organization_id=org_id,
                header_id=advance.id,
                effective_from=EFFECTIVE_FROM,
                values={
                    "installment_amount": money(line.amount).quantize(_TWO),
                    "installments_total": 24,
                    "installments_recovered_opening": 0,
                },
                change_reason=None,
                created_by=user_id,
            )
            continue
        if code == "ACCOMMODATION_LICENSE_FEE":
            assert employee.accommodation is not None
            foregone = line_amount(employee, "FOREGONE_HRA")
            assignment = AccommodationAssignment(
                organization_id=org_id,
                employee_id=header.id,
                quarters_location=map_quarters_location(employee.accommodation.location),
                quarters_identifier=f"Q-{employee.fixture_id}",
            )
            session.add(assignment)
            await session.flush()
            await versioning.insert_version(
                session,
                accommodation_charge_versions,
                organization_id=org_id,
                header_id=assignment.id,
                effective_from=EFFECTIVE_FROM,
                values={
                    "license_fee": money(line.amount).quantize(_TWO),
                    "informational_hra_foregone": (
                        money(foregone).quantize(_TWO) if foregone is not None else None
                    ),
                },
                change_reason=None,
                created_by=user_id,
            )
            continue
        if code in ACCOMMODATION_COMPONENT_CODES:
            continue
        raise AssertionError(f"No seeding strategy for {employee.fixture_id} line {code}")

    return header.id


async def _june_world(session: AsyncSession) -> dict:
    global _CACHED_WORLD
    if _CACHED_WORLD is None:
        _CACHED_WORLD = await _seed_posted_june(session)
    return _CACHED_WORLD


def _ctx(world: dict, *, run_id: UUID | None = None) -> ReportContext:
    return ReportContext(
        organization_id=world["org_id"],
        posted_run_id=run_id or world["run_id"],
        template_version=TEMPLATE_VERSION,
        generated_at=datetime.now(UTC),
        engine_version=str(world["engine_version"]),
    )


def _section_by_title(dto, title: str):
    for section in dto.sections:
        if section.title == title:
            return section
    raise AssertionError(f"section {title!r} not found")


def _col_index(section, key: str) -> int:
    for idx, col in enumerate(section.columns):
        if col.key == key:
            return idx
    raise AssertionError(f"column {key!r} not found")


@pytest.mark.asyncio
async def test_bank_advice_total_and_row_count(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    dto = await bank_advice_builder.build(session, _ctx(world))

    assert dto.report_type == REPORT_TYPE_BANK_ADVICE
    credits = _section_by_title(dto, "Payment credits")
    assert credits.totals is not None
    net_idx = _col_index(credits, "net_payable")
    assert _dec(credits.totals[net_idx]) == EXPECTED_NET

    db_rows = (
        (
            await session.execute(
                sa.select(payroll_employee_results).where(
                    payroll_employee_results.c.organization_id == world["org_id"],
                    payroll_employee_results.c.run_version_id == world["version_id"],
                )
            )
        )
        .mappings()
        .all()
    )
    paid_count = sum(1 for row in db_rows if _dec(row["net_payable"]) > _ZERO)
    assert len(credits.rows) == paid_count
    assert paid_count == 32


@pytest.mark.asyncio
async def test_bank_advice_full_account_numbers(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    dto = await bank_advice_builder.build(session, _ctx(world))
    credits = _section_by_title(dto, "Payment credits")
    acct_idx = _col_index(credits, "account_number")
    emp_idx = _col_index(credits, "employee_number")

    for row in credits.rows:
        emp_no = str(row[emp_idx])
        expected_acct = world["bank_by_fixture"][emp_no]
        assert row[acct_idx] == expected_acct
        assert "••••" not in str(row[acct_idx])


@pytest.mark.asyncio
async def test_missing_primary_account_error_then_recovers(session):
    world = await _june_world(session)
    target_number = "E001"
    target_id = world["employee_ids"][target_number]
    fixture_emp = world["fixture_employees"][target_number]

    await _bind(session, world["org_id"], world["user_id"])
    await session.execute(
        sa.delete(employee_bank_account_versions).where(
            employee_bank_account_versions.c.organization_id == world["org_id"],
            employee_bank_account_versions.c.header_id == target_id,
        )
    )
    await session.commit()

    try:
        await _bind(session, world["org_id"], world["user_id"])
        with pytest.raises(MissingPrimarySalaryAccountError) as exc_info:
            await bank_advice_builder.build(session, _ctx(world))
        assert target_number in exc_info.value.employee_numbers
        assert target_number in str(exc_info.value)

        await _bind(session, world["org_id"], world["user_id"])
        await versioning.insert_version(
            session,
            employee_bank_account_versions,
            organization_id=world["org_id"],
            header_id=target_id,
            effective_from=EFFECTIVE_FROM,
            values={
                "account_number": fixture_emp.bank_account,
                "ifsc": fixture_emp.ifsc,
                "bank_name": "Synthetic Bank",
                "branch": "Synthetic Branch",
                "is_primary_salary": True,
            },
            change_reason="restore for payments lane test",
            created_by=world["user_id"],
        )
        await session.commit()

        await _bind(session, world["org_id"], world["user_id"])
        dto = await bank_advice_builder.build(session, _ctx(world))
        credits = _section_by_title(dto, "Payment credits")
        assert _dec(credits.totals[_col_index(credits, "net_payable")]) == EXPECTED_NET
    finally:
        # Ensure cached world keeps a bank account for later tests even if assert fails.
        await _bind(session, world["org_id"], world["user_id"])
        existing = (
            await session.execute(
                sa.select(employee_bank_account_versions.c.id).where(
                    employee_bank_account_versions.c.organization_id == world["org_id"],
                    employee_bank_account_versions.c.header_id == target_id,
                    employee_bank_account_versions.c.is_primary_salary.is_(True),
                )
            )
        ).first()
        if existing is None:
            await versioning.insert_version(
                session,
                employee_bank_account_versions,
                organization_id=world["org_id"],
                header_id=target_id,
                effective_from=EFFECTIVE_FROM,
                values={
                    "account_number": fixture_emp.bank_account,
                    "ifsc": fixture_emp.ifsc,
                    "bank_name": "Synthetic Bank",
                    "branch": "Synthetic Branch",
                    "is_primary_salary": True,
                },
                change_reason="finally restore bank account",
                created_by=world["user_id"],
            )
            await session.commit()


@pytest.mark.asyncio
async def test_payslip_nets_sum_and_line_reconciliation(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    dto = await payslip_bundle_builder.build(session, _ctx(world))

    assert dto.report_type == REPORT_TYPE_PAYSLIPS
    assert len(dto.sections) == 32

    nets_sum = _ZERO
    kind_idx = 0
    code_idx = 1
    amount_idx = 3

    db_results = (
        (
            await session.execute(
                sa.select(payroll_employee_results).where(
                    payroll_employee_results.c.organization_id == world["org_id"],
                    payroll_employee_results.c.run_version_id == world["version_id"],
                )
            )
        )
        .mappings()
        .all()
    )
    by_number = {row["employee_number"]: row for row in db_results}
    assert len(by_number) == 32

    for section in dto.sections:
        emp_no = section.title.removeprefix("Payslip — ")
        db = by_number[emp_no]

        net_row = next(row for row in section.rows if row[code_idx] == "net_payable")
        net = _dec(net_row[amount_idx])
        assert net == _dec(db["net_payable"])
        nets_sum += net

        words_row = next(row for row in section.rows if row[code_idx] == "amount_in_words")
        assert words_row[2] == amount_in_words(net)

        line_rows = (
            (
                await session.execute(
                    sa.select(payroll_result_lines)
                    .where(
                        payroll_result_lines.c.organization_id == world["org_id"],
                        payroll_result_lines.c.employee_result_id == db["id"],
                    )
                    .order_by(payroll_result_lines.c.sequence)
                )
            )
            .mappings()
            .all()
        )

        payslip_component_rows = [
            row
            for row in section.rows
            if row[kind_idx] in {"earning", "deduction", "employer_contribution", "informational"}
        ]
        assert len(payslip_component_rows) == len(line_rows)
        for payslip_row, db_line in zip(payslip_component_rows, line_rows, strict=True):
            assert payslip_row[code_idx] == db_line["component_code"]
            assert _dec(payslip_row[amount_idx]) == _dec(db_line["amount"])

    assert _dec(nets_sum) == EXPECTED_NET


@pytest.mark.asyncio
async def test_payslip_masked_ids(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    dto = await payslip_bundle_builder.build(session, _ctx(world))

    saw_pan_mask = False
    for section in dto.sections:
        for row in section.rows:
            if row[1] == "pan":
                detail = str(row[2])
                assert detail.startswith("••••")
                saw_pan_mask = True
                emp_no = section.title.removeprefix("Payslip — ")
                raw_pan = world["fixture_employees"][emp_no].pan
                assert detail == mask_value(raw_pan)
            if row[1] == "pran" and row[2]:
                assert str(row[2]).startswith("••••")
    assert saw_pan_mask


@pytest.mark.asyncio
async def test_bank_advice_excel_reload(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    dto = await bank_advice_builder.build(session, _ctx(world))
    payload = bank_advice_to_json(dto)
    assert payload["report_type"] == REPORT_TYPE_BANK_ADVICE
    assert payload["sections"][0]["totals"]["net_payable"] == "3838095.00"

    xlsx = bank_advice_to_excel(dto)
    wb = load_workbook(BytesIO(xlsx))
    ws = wb.active
    assert ws is not None
    credits = _section_by_title(dto, "Payment credits")
    # Header row 5; data starts at 6; totals after data rows.
    assert ws.cell(row=5, column=1).value == "Employee No."
    first_emp = ws.cell(row=6, column=1).value
    assert first_emp == credits.rows[0][0]
    # Full account number in column 3.
    assert ws.cell(row=6, column=3).value == credits.rows[0][2]
    assert "••••" not in str(ws.cell(row=6, column=3).value)

    totals_row = 6 + len(credits.rows)
    net_cell = ws.cell(row=totals_row, column=5)
    assert net_cell.number_format == MONEY_FORMAT
    assert _dec(net_cell.value) == EXPECTED_NET


@pytest.mark.asyncio
async def test_payslip_pdf_contains_employee_and_net(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    dto = await payslip_bundle_builder.build(session, _ctx(world))

    pdf = payslip_to_pdf(dto)
    assert pdf.startswith(b"%PDF")
    reader = PdfReader(BytesIO(pdf))
    assert len(reader.pages) == 32

    # Spot-check first page has employee number and that employee's net amount.
    page0 = reader.pages[0].extract_text() or ""
    first_emp = dto.sections[0].title.removeprefix("Payslip — ")
    first_net = next(row[3] for row in dto.sections[0].rows if row[1] == "net_payable")
    assert first_emp in page0
    assert format_inr(_dec(first_net)) in page0

    payload = payslip_to_json(dto)
    assert payload["report_type"] == REPORT_TYPE_PAYSLIPS
    assert len(payload["sections"]) == 32

    with pytest.raises(NotImplementedError, match="intentionally skipped"):
        payslip_to_excel(dto)


@pytest.mark.asyncio
async def test_unposted_run_raises_conflict(session):
    world = await _june_world(session)
    await _bind(session, world["org_id"], world["user_id"])

    period = PayrollPeriod(
        organization_id=world["org_id"],
        period_year=2026,
        period_month=7,
        status="open",
    )
    session.add(period)
    await session.flush()
    draft = PayrollRun(
        organization_id=world["org_id"],
        period_id=period.id,
        run_type="supplemental",
        status="draft",
    )
    session.add(draft)
    await session.flush()
    draft_id = draft.id
    await session.commit()

    await _bind(session, world["org_id"], world["user_id"])
    with pytest.raises(ConflictError, match="must be posted"):
        await bank_advice_builder.build(session, _ctx(world, run_id=draft_id))

    await _bind(session, world["org_id"], world["user_id"])
    with pytest.raises(ConflictError, match="must be posted"):
        await payslip_bundle_builder.build(session, _ctx(world, run_id=draft_id))


def test_register_payment_reports_entries() -> None:
    registry = ReportRegistry()
    register_payment_reports(registry)
    assert REPORT_TYPE_BANK_ADVICE in registry
    assert REPORT_TYPE_PAYSLIPS in registry
    bank = registry.get(REPORT_TYPE_BANK_ADVICE)
    slips = registry.get(REPORT_TYPE_PAYSLIPS)
    assert bank.builder is bank_advice_builder
    assert slips.builder is payslip_bundle_builder
