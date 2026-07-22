"""Allowlist consistency + consolidated ZIP export for product report sheets."""

from __future__ import annotations

import io
import json
import zipfile
from decimal import Decimal
from pathlib import Path
from typing import Any
from uuid import UUID, uuid4

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.jobs.memory import InMemoryJobQueue
from app.models.payroll_runs import PayrollPeriod, PayrollRun
from app.models.platform import ExportArtifact
from app.reports.base import (
    ColumnKind,
    ReportColumn,
    ReportContext,
    ReportDTO,
    ReportRegistry,
    TableSection,
    to_json,
)
from app.reports.canonical_excel import CANONICAL_PRODUCT_SHEETS, consolidate_v3_workbooks
from app.reports.canonical_front_sheets import bank_tip_to_excel, payslip_to_excel
from app.reports.canonical_schedules import clone_canonical_sheet_structure
from app.reports.registry_setup import (
    PRODUCT_REPORT_SHEET_TITLES,
    PRODUCT_REPORT_SHEETS,
    build_report_registry,
)
from app.reports.excel import to_excel
from app.services.artifacts import create_artifact
from app.services.report_generation import (
    CANONICAL_RENDERER_REVISION,
    execute_consolidated_xlsx,
    list_registered_reports,
    manifest_hash,
    product_sheet_manifest,
    request_consolidated_export,
)
from app.storage.memory import InMemoryObjectStorage
from app.tenancy import bind_tenant_context
from tests.identity_helpers import seed_membership, seed_organization, seed_user

CONTENT_TYPES = {
    "json": "application/json",
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}


@pytest.mark.parametrize("payload", ("=1+1", "+cmd", "-2+3", "@SUM(A1:A2)"))
def test_canonical_front_sheets_escape_untrusted_text_without_escaping_formulas(
    payload: str,
) -> None:
    label_columns = (
        ReportColumn(key="label", header="Label"),
        ReportColumn(key="value", header="Value"),
    )
    bank_dto = ReportDTO(
        report_type="bank_rtgs_advice",
        template_version="v3",
        title="Bank advice",
        organization_name="Fallback organization",
        subtitle="June 2026",
        sections=(
            TableSection(
                title="Advice recipient",
                columns=label_columns,
                rows=(("Bank", payload),),
            ),
            TableSection(
                title="Payment credits",
                columns=(ReportColumn(key="name", header="Name"),),
                rows=(),
            ),
        ),
        metadata={
            "report_profile": {
                "legal_name": payload,
                "address_lines": [payload],
                "signatories": [
                    {
                        "role": "maker",
                        "name": payload,
                        "designation": "Financial Officer",
                    }
                ],
            }
        },
    )
    bank = load_workbook(io.BytesIO(bank_tip_to_excel(bank_dto)), data_only=False)["Bank Tip"]
    for coordinate in ("D1", "D4", "C9", "D20"):
        assert bank[coordinate].value.startswith("'")
        assert bank[coordinate].data_type == "s"
    assert bank["G14"].value == "=SUM(G14:G13)"
    assert bank["G14"].data_type == "f"

    payslip_columns = (
        ReportColumn(key="line_kind", header="Line kind"),
        ReportColumn(key="code", header="Code"),
        ReportColumn(key="detail", header="Detail"),
        ReportColumn(key="amount", header="Amount", kind=ColumnKind.MONEY),
        ReportColumn(key="employer_transfer", header="Employer transfer"),
    )
    payslip_dto = ReportDTO(
        report_type="payslips",
        template_version="v3",
        title="Payslips",
        organization_name="Organization",
        subtitle="June 2026",
        sections=(
            TableSection(
                title="Employee",
                columns=payslip_columns,
                rows=(
                    ("identity", "name", payload, None, None),
                    ("identity", "designation", payload, None, None),
                    ("identity", "pran", payload, None, None),
                    ("earning", payload, "earning", Decimal("10.00"), None),
                    ("net", "disbursement", None, Decimal("10.00"), None),
                    ("net", "amount_in_words", payload, None, None),
                ),
            ),
        ),
    )
    payslip = load_workbook(io.BytesIO(payslip_to_excel(payslip_dto)), data_only=False)["PaySlip"]
    for coordinate in ("E2", "K2", "T2", "B6", "H16"):
        assert payslip[coordinate].value.startswith("'")
        assert payslip[coordinate].data_type == "s"
    assert payslip["F15"].value == "=SUM(F6:F14)"
    assert payslip["F15"].data_type == "f"


def test_consolidated_v3_copies_dynamic_overflow_without_stale_template_merges() -> None:
    rendered: list[tuple[str, str, bytes]] = []
    for name, _report_type, state in CANONICAL_PRODUCT_SHEETS:
        source = Workbook()
        sheet = source.active
        clone_canonical_sheet_structure(sheet_name=name, target=sheet)
        sheet.sheet_state = "visible"
        if name == "Bank Tip":
            sheet.merge_cells("B200:D200")
            sheet["B200"] = "Dynamic overflow row"
            sheet.print_area = "A1:G200"
        output = io.BytesIO()
        source.save(output)
        rendered.append((name, state, output.getvalue()))

    workbook = load_workbook(io.BytesIO(consolidate_v3_workbooks(rendered)))
    bank = workbook["Bank Tip"]
    assert bank["B200"].value == "Dynamic overflow row"
    assert "B200:D200" in {str(item) for item in bank.merged_cells.ranges}
    assert bank.print_area == "'Bank Tip'!$A$1:$G$200"


class _FakeBuilder:
    async def build(self, session: Any, ctx: ReportContext) -> ReportDTO:
        return ReportDTO(
            report_type="placeholder",
            template_version=ctx.template_version,
            title="Sheet",
            organization_name="Org",
            subtitle="June 2026",
            sections=(
                TableSection(
                    title="Rows",
                    columns=(
                        ReportColumn(key="name", header="Name", kind=ColumnKind.TEXT),
                        ReportColumn(key="amount", header="Amount", kind=ColumnKind.MONEY),
                    ),
                    rows=(("Ada", Decimal("1.00")),),
                ),
            ),
            metadata={
                "report_profile": {
                    "legal_name": "Ornament Legal Ltd.",
                    "office_name": "Payroll Office",
                    "address_lines": ["1 Canonical Way", "Mumbai"],
                    "cin": "CIN-123",
                    "phone": "+91-22-5555-0100",
                    "website": "https://example.test",
                    "ddo_name": "Dana DDO",
                    "ddo_code": "DDO-42",
                    "department_code": "DEPT-7",
                    "administrative_department": "Finance Department",
                    "treasury_code": "TRY-9",
                    "fund_source": "Consolidated Fund",
                    "plan_status": "Non-plan",
                    "nps_employee_account_head": "8342 Employee NPS",
                    "nps_employer_account_head": "8342 Employer NPS",
                    "salary_reference_prefix": "SAL-REF",
                    "pay_bill_footer_text": "Certified for treasury payment.",
                    "head_of_account": {
                        "major_head": "2070",
                        "sub_head": "01",
                        "detailed_head": "001",
                    },
                    "signatories": [
                        {"role": "maker", "name": "Mira Maker", "designation": "Accountant"},
                        {"role": "checker", "name": "Chen Checker", "designation": "Officer"},
                        {
                            "role": "approving_officer",
                            "name": "Asha Approver",
                            "designation": "Chief Administrative Officer",
                        },
                    ],
                },
                "run_metadata": {
                    "bank_advice_number": "BANK-2026-06",
                    "bank_advice_date": "2026-06-30",
                    "bill_date": "2026-06-30",
                },
            },
        )


def _minimal_dto(report_type: str) -> ReportDTO:
    return ReportDTO(
        report_type=report_type,
        template_version="v2",
        title=PRODUCT_REPORT_SHEET_TITLES.get(report_type, report_type),
        organization_name="Test Org",
        subtitle="June 2026",
        sections=(
            TableSection(
                title="Rows",
                columns=(
                    ReportColumn(key="name", header="Name", kind=ColumnKind.TEXT),
                    ReportColumn(key="amount", header="Amount", kind=ColumnKind.MONEY),
                ),
                rows=(("Ada", Decimal("1.00")),),
                totals=(None, Decimal("1.00")),
            ),
        ),
    )


def _product_fake_registry() -> ReportRegistry:
    registry = ReportRegistry()
    builder = _FakeBuilder()
    for report_type in PRODUCT_REPORT_SHEETS:
        registry.register(
            report_type,
            builder=builder,
            to_json=to_json,
            to_excel=lambda dto: b"PK\x03\x04fake-xlsx",
            to_pdf=lambda dto: b"%PDF-fake",
            content_types=CONTENT_TYPES,
            filename_pattern="{report_type}_{posted_run_id}.{ext}",
        )
    return registry


def _product_valid_xlsx_registry() -> ReportRegistry:
    registry = ReportRegistry()
    builder = _FakeBuilder()
    for report_type in PRODUCT_REPORT_SHEETS:
        registry.register(
            report_type,
            builder=builder,
            to_json=to_json,
            to_excel=to_excel,
            to_pdf=lambda dto: b"%PDF-fake",
            content_types=CONTENT_TYPES,
            filename_pattern="{report_type}_{posted_run_id}.{ext}",
        )
    return registry


async def _bind(session: AsyncSession, org_id: UUID, user_id: UUID) -> None:
    if session.in_transaction():
        await session.rollback()
    await session.begin()
    await bind_tenant_context(session, organization_id=org_id, user_id=user_id)


async def _seed_posted_run(session: AsyncSession) -> dict[str, UUID]:
    if session.in_transaction():
        await session.rollback()
    org = await seed_organization(session, name="Product Sheets Org", slug=f"ps-{uuid4().hex[:10]}")
    user = await seed_user(session, workos_user_id=f"ps_{uuid4().hex[:10]}")
    await seed_membership(
        session,
        organization_id=org.id,
        user_id=user.id,
        role="organization_administrator",
    )
    period = PayrollPeriod(
        organization_id=org.id,
        period_year=2026,
        period_month=6,
        status="open",
    )
    session.add(period)
    await session.flush()
    run = PayrollRun(
        organization_id=org.id,
        period_id=period.id,
        status="posted",
    )
    session.add(run)
    await session.commit()
    await _bind(session, org.id, user.id)
    return {"organization_id": org.id, "user_id": user.id, "posted_run_id": run.id}


def test_product_report_sheets_allowlist_consistency() -> None:
    registry = build_report_registry()
    assert len(PRODUCT_REPORT_SHEETS) == 18
    assert "payslips" in PRODUCT_REPORT_SHEETS
    assert "advance_schedule" not in PRODUCT_REPORT_SHEETS
    # Generic custom variant infrastructure stays outside the fixed workbook pack.
    assert "advance_schedule" in registry

    for report_type in PRODUCT_REPORT_SHEETS:
        assert report_type in registry, f"{report_type} not registered"
        assert report_type in PRODUCT_REPORT_SHEET_TITLES
        registration = registry.get(report_type)
        assert "excel" in registration.formatters.content_types
        assert "xlsx" not in registration.formatters.content_types


def test_v3_manifest_carries_immutable_renderer_revision() -> None:
    manifest = product_sheet_manifest(template_version="v3")
    assert len(manifest) == 18
    assert {version for _, version in manifest} == {f"v3+{CANONICAL_RENDERER_REVISION}"}
    legacy_manifest = tuple(sorted((report_type, "v3") for report_type in PRODUCT_REPORT_SHEETS))
    assert manifest_hash(manifest) != manifest_hash(legacy_manifest)


@pytest.mark.asyncio
async def test_consolidated_request_defaults_to_revisioned_v3(session, monkeypatch) -> None:
    async def ready(*_args, **_kwargs):
        return None

    monkeypatch.setattr("app.services.report_generation.require_v3_report_readiness", ready)
    world = await _seed_posted_run(session)
    job = await request_consolidated_export(
        session,
        InMemoryJobQueue(),
        organization_id=world["organization_id"],
        posted_run_id=world["posted_run_id"],
        requested_by=world["user_id"],
        registry=_product_fake_registry(),
    )

    assert job.payload["template_version"] == "v3"
    assert job.payload["renderer_revision"] == CANONICAL_RENDERER_REVISION
    assert job.payload["manifest_hash"] == manifest_hash(
        product_sheet_manifest(template_version="v3")
    )


@pytest.mark.asyncio
async def test_revisioned_consolidated_export_does_not_reuse_pre_fix_artifact(
    session, monkeypatch
) -> None:
    async def ready(*_args, **_kwargs):
        return None

    monkeypatch.setattr("app.services.report_generation.require_v3_report_readiness", ready)
    monkeypatch.setattr("app.services.report_generation.CANONICAL_PRODUCT_SHEETS", ())
    monkeypatch.setattr(
        "app.services.report_generation.consolidate_v3_workbooks",
        lambda _sources: b"revisioned-canonical-xlsx",
    )
    world = await _seed_posted_run(session)
    storage = InMemoryObjectStorage()
    legacy_manifest = tuple(sorted((report_type, "v3") for report_type in PRODUCT_REPORT_SHEETS))
    legacy_hash = manifest_hash(legacy_manifest)
    legacy = await create_artifact(
        session,
        storage,
        organization_id=world["organization_id"],
        report_type="consolidated_xlsx",
        template_version=f"v3+{legacy_hash}",
        content=b"pre-fix-xlsx",
        content_type=CONTENT_TYPES["excel"],
        requested_by=world["user_id"],
        posted_run_id=world["posted_run_id"],
    )
    legacy_id = legacy.id
    await _bind(session, world["organization_id"], world["user_id"])
    job = await request_consolidated_export(
        session,
        InMemoryJobQueue(),
        organization_id=world["organization_id"],
        posted_run_id=world["posted_run_id"],
        requested_by=world["user_id"],
        registry=_product_fake_registry(),
    )
    job.payload["manifest_hash"] = legacy_hash

    await _bind(session, world["organization_id"], world["user_id"])
    first = await execute_consolidated_xlsx(
        session,
        storage,
        job,
        registry=_product_fake_registry(),
    )
    assert first["artifact_id"] != str(legacy_id)
    assert first["manifest_hash"] == manifest_hash(product_sheet_manifest(template_version="v3"))

    await _bind(session, world["organization_id"], world["user_id"])
    current = await session.get(ExportArtifact, UUID(first["artifact_id"]))
    assert current is not None
    assert current.template_version.startswith(f"v3+{CANONICAL_RENDERER_REVISION}+")

    await _bind(session, world["organization_id"], world["user_id"])
    second = await execute_consolidated_xlsx(
        session,
        storage,
        job,
        registry=_product_fake_registry(),
    )
    assert second["artifact_id"] == first["artifact_id"]
    assert second["reused"] is True


def test_family_content_types_use_excel_key() -> None:
    registry = build_report_registry()
    for report_type, registration in registry._entries.items():  # noqa: SLF001
        keys = set(registration.formatters.content_types)
        assert "xlsx" not in keys, report_type
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if mime in registration.formatters.content_types.values():
            assert "excel" in keys, report_type


def test_product_sheets_to_excel_callable() -> None:
    """Each allowlisted sheet's to_excel accepts a minimal DTO without raising."""
    registry = build_report_registry()
    for report_type in PRODUCT_REPORT_SHEETS:
        registration = registry.get(report_type)
        xlsx = registration.formatters.to_excel(_minimal_dto(report_type))
        assert isinstance(xlsx, (bytes, bytearray))
        assert xlsx[:2] == b"PK"
        wb = load_workbook(io.BytesIO(xlsx))
        assert wb.sheetnames


def test_list_reports_marks_product_sheets() -> None:
    registry = build_report_registry()
    items = {item.report_type: item for item in list_registered_reports(registry)}
    product = [rt for rt, item in items.items() if item.product_sheet]
    assert sorted(product) == sorted(PRODUCT_REPORT_SHEETS)
    assert items["payslips"].product_sheet is True
    assert items["advance_schedule"].product_sheet is False
    assert items["pay_bill"].title == "Pay Bill"
    assert "excel" in items["pay_bill"].formats


@pytest.mark.asyncio
async def test_consolidated_export_zip_has_eighteen_entries(session) -> None:
    world = await _seed_posted_run(session)
    registry = _product_fake_registry()
    storage = InMemoryObjectStorage()
    queue = InMemoryJobQueue()

    job = await request_consolidated_export(
        session,
        queue,
        organization_id=world["organization_id"],
        posted_run_id=world["posted_run_id"],
        requested_by=world["user_id"],
        registry=registry,
        template_version="v2",
    )
    assert job.job_type == "consolidated_xlsx"
    assert job.dedupe_key.startswith(f"consolidated_xlsx:{world['posted_run_id']}:")
    expected_hash = manifest_hash(product_sheet_manifest(template_version="v2"))
    assert job.dedupe_key == f"consolidated_xlsx:{world['posted_run_id']}:{expected_hash}"

    result = await execute_consolidated_xlsx(
        session,
        storage,
        job,
        registry=registry,
    )
    assert result["artifact_id"]
    assert result["manifest_hash"] == expected_hash
    assert "June 2026" in result["filename"]
    assert result["filename"].endswith(".zip")

    artifact = await session.get(ExportArtifact, UUID(result["artifact_id"]))
    assert artifact is not None
    assert artifact.content_type == "application/zip"
    assert artifact.report_type == "consolidated_xlsx"

    blob = await storage.get(artifact.object_key)
    with zipfile.ZipFile(io.BytesIO(blob), "r") as archive:
        names = sorted(archive.namelist())
    assert len(names) == 18
    assert any("payslip" in n.lower() for n in names)
    assert any("advance" in n.lower() for n in names)
    for title in PRODUCT_REPORT_SHEET_TITLES.values():
        assert any(name.startswith(f"{title} - ") and name.endswith(".xlsx") for name in names)

    reused = await execute_consolidated_xlsx(
        session,
        storage,
        job,
        registry=registry,
    )
    assert reused["reused"] is True
    assert reused["artifact_id"] == result["artifact_id"]


@pytest.mark.asyncio
async def test_consolidated_v3_is_one_workbook_with_canonical_topology(
    session, monkeypatch
) -> None:
    async def ready(*_args, **_kwargs):
        return None

    monkeypatch.setattr("app.services.report_generation.require_v3_report_readiness", ready)
    contract_path = (
        Path(__file__).resolve().parents[3]
        / "fixtures/sanitized/june-2026/canonical_export_contract.json"
    )
    contract = json.loads(contract_path.read_text())
    expected_names = [sheet["name"] for sheet in contract["sheets"]]
    expected_states = {sheet["name"]: sheet["state"] for sheet in contract["sheets"]}

    world = await _seed_posted_run(session)
    registry = _product_valid_xlsx_registry()
    storage = InMemoryObjectStorage()
    queue = InMemoryJobQueue()
    job = await request_consolidated_export(
        session,
        queue,
        organization_id=world["organization_id"],
        posted_run_id=world["posted_run_id"],
        requested_by=world["user_id"],
        registry=registry,
        template_version="v3",
    )
    result = await execute_consolidated_xlsx(session, storage, job, registry=registry)
    assert result["filename"].endswith(".xlsx")

    artifact = await session.get(ExportArtifact, UUID(result["artifact_id"]))
    assert artifact is not None
    assert artifact.content_type == CONTENT_TYPES["excel"]
    workbook = load_workbook(io.BytesIO(await storage.get(artifact.object_key)))
    assert workbook.sheetnames == expected_names
    assert {name: workbook[name].sheet_state for name in expected_names} == expected_states

    office = workbook["office tip"]
    bank = workbook["Bank Tip"]
    payslip = workbook["PaySlip"]
    face = workbook[" Face "]
    assert office.print_area == "'office tip'!$A$1:$F$64"
    assert office.page_setup.orientation == "portrait"
    assert office.page_setup.scale == 90
    assert {"D4:E4", "C5:E5", "C6:E6"} <= {str(item) for item in office.merged_cells.ranges}
    assert bank.page_setup.orientation == "portrait"
    assert bank.page_setup.scale == 61
    assert {"B1:C5", "D1:G1", "E10:G10"} <= {str(item) for item in bank.merged_cells.ranges}
    assert bank["D1"].value == "Ornament Legal Ltd."
    assert "1 Canonical Way" in bank["D4"].value
    assert "CIN-123" in bank["D4"].value
    assert payslip.page_setup.orientation == "landscape"
    assert payslip.page_setup.scale == 52
    assert payslip.print_area.startswith("'PaySlip'!$A$1:$X$")
    assert payslip.calculate_dimension() == "A2:X18"
    assert face.print_area == "' Face '!$A$1:$U$121"
    assert face.page_setup.orientation == "landscape"
    assert face.page_setup.scale == 51
    assert face["I22"].value == "=I18-I19+I20"
    assert face["I51"].value == "=I36+I49"
    assert face["I52"].value == "=I22-I51"
    assert face["H5"].value == "Finance Department"
    assert "DEPT-7" in face["I7"].value
    assert "Consolidated Fund" in face["B15"].value
    assert "TRY-9" in face["B16"].value
    nps = workbook["Pension Sub (2)"]
    assert "Dana DDO" in nps["F4"].value
    assert "DEPT-7" in nps["H4"].value
    assert "TRY-9" in nps["I4"].value
    assert nps["B42"].value == "8342 Employee NPS"
    assert nps["B43"].value == "8342 Employer NPS"
    assert workbook["P.T."].calculate_dimension() == "B1:G35"
    assert workbook["GIS"].calculate_dimension() == "B1:F35"
    assert nps.calculate_dimension() == "B1:J57"
    for sheet in (office, bank, payslip, face):
        formulas = (
            cell.value
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        )
        assert all("#REF!" not in formula for formula in formulas)
