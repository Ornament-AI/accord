"""Golden tests for GPF Mumbai / Nagpur and NPS contribution schedules."""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from uuid import UUID, uuid4

import pytest
import sqlalchemy as sa
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy.ext.asyncio import AsyncSession

from app.exceptions import ConflictError
from app.models.employees import employee_profile_versions
from app.models.identity import Organization
from app.models.payroll_runs import PayrollPeriod, PayrollRun
from app.reports.base import ReportContext, ReportRegistry
from app.reports.excel import MONEY_FORMAT
from app.reports.families.retirement import (
    REPORT_TYPE_GPF_MUMBAI,
    REPORT_TYPE_GPF_NAGPUR,
    REPORT_TYPE_NPS,
    gpf_mumbai_builder,
    gpf_nagpur_builder,
    nps_contribution_builder,
    register_retirement_reports,
    retirement_to_excel,
    retirement_to_pdf,
)
from tests.e2e.fixture_loader import load_june_fixture
from tests.identity_helpers import seed_organization, seed_user
from tests.reports.test_payroll_register import (
    TEMPLATE_VERSION,
    _bind,
    _dec,
    _june_world,
    _seed_posted_june,
)

# Lane-local cache; invalidate if shared DB truncate removes the seeded org.
_CACHED_WORLD: dict | None = None


async def _retirement_june_world(session: AsyncSession) -> dict:
    """Posted June world for this lane (re-seed if shared-DB cache is stale)."""
    global _CACHED_WORLD
    if _CACHED_WORLD is not None:
        await _bind(session, _CACHED_WORLD["org_id"], _CACHED_WORLD["user_id"])
        org = await session.get(Organization, _CACHED_WORLD["org_id"])
        if org is not None:
            return _CACHED_WORLD
        _CACHED_WORLD = None

    # Prefer the payroll-register module cache when still valid (one June seed).
    shared = await _june_world(session)
    await _bind(session, shared["org_id"], shared["user_id"])
    org = await session.get(Organization, shared["org_id"])
    if org is not None:
        _CACHED_WORLD = shared
        return _CACHED_WORLD

    _CACHED_WORLD = await _seed_posted_june(session)
    return _CACHED_WORLD


def _ctx(world: dict, *, run_id: UUID | None = None) -> ReportContext:
    return ReportContext(
        organization_id=world["org_id"],
        posted_run_id=run_id or world["run_id"],
        template_version=TEMPLATE_VERSION,
        generated_at=datetime.now(UTC),
        engine_version=str(world["engine_version"]),
    )


def _section(dto, *, index: int = 0):
    assert dto.sections, "expected at least one section"
    return dto.sections[index]


def _col_index(section, key: str) -> int:
    for idx, col in enumerate(section.columns):
        if col.key == key:
            return idx
    raise AssertionError(f"column {key!r} not found")


async def _profile_by_employee_number(
    session: AsyncSession,
    *,
    org_id: UUID,
    employee_ids: dict[str, UUID],
) -> dict[str, dict]:
    """Map employee_number → active profile fields for June as-of."""
    out: dict[str, dict] = {}
    for emp_no, emp_id in employee_ids.items():
        row = (
            (
                await session.execute(
                    sa.select(employee_profile_versions).where(
                        employee_profile_versions.c.organization_id == org_id,
                        employee_profile_versions.c.header_id == emp_id,
                    )
                )
            )
            .mappings()
            .one()
        )
        out[emp_no] = dict(row)
    return out


@pytest.mark.asyncio
async def test_gpf_mumbai_and_nagpur_schedules_golden(session):
    world = await _retirement_june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    fixture = load_june_fixture()
    profiles = await _profile_by_employee_number(
        session,
        org_id=world["org_id"],
        employee_ids=world["employee_ids"],
    )

    mumbai = await gpf_mumbai_builder.build(session, _ctx(world))
    await _bind(session, world["org_id"], world["user_id"])
    nagpur = await gpf_nagpur_builder.build(session, _ctx(world))

    assert mumbai.report_type == REPORT_TYPE_GPF_MUMBAI
    assert nagpur.report_type == REPORT_TYPE_GPF_NAGPUR
    assert "Mumbai" in mumbai.title
    assert "Nagpur" in nagpur.title

    m_section = _section(mumbai)
    n_section = _section(nagpur)
    assert len(m_section.rows) == 9
    assert len(n_section.rows) == 7

    sub_idx = _col_index(m_section, "subscription")
    acct_idx = _col_index(m_section, "gpf_account_number")
    assert m_section.totals is not None and n_section.totals is not None
    mumbai_total = _dec(m_section.totals[sub_idx])
    nagpur_total = _dec(n_section.totals[sub_idx])
    assert mumbai_total == _dec("165000.00")
    assert nagpur_total == _dec("115000.00")
    assert mumbai_total + nagpur_total == _dec("280000.00")
    assert mumbai_total + nagpur_total == _dec(fixture.expected.aggregates["gpf_total"])

    mumbai_ids = {row[0] for row in m_section.rows}
    nagpur_ids = {row[0] for row in n_section.rows}
    assert mumbai_ids.isdisjoint(nagpur_ids)

    for row in m_section.rows:
        emp_no = row[0]
        assert profiles[emp_no]["gpf_jurisdiction"] == "mumbai"
        assert profiles[emp_no]["retirement_regime"] == "gpf"
        assert row[acct_idx], f"missing GPF account for {emp_no}"
        assert str(row[acct_idx]).startswith("SYNGPF/MUM/")

    n_acct_idx = _col_index(n_section, "gpf_account_number")
    for row in n_section.rows:
        emp_no = row[0]
        assert profiles[emp_no]["gpf_jurisdiction"] == "nagpur"
        assert profiles[emp_no]["retirement_regime"] == "gpf"
        assert row[n_acct_idx], f"missing GPF account for {emp_no}"
        assert str(row[n_acct_idx]).startswith("SYNGPF/NGP/")


@pytest.mark.asyncio
async def test_nps_schedule_golden_excludes_epf(session):
    world = await _retirement_june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    fixture = load_june_fixture()
    profiles = await _profile_by_employee_number(
        session,
        org_id=world["org_id"],
        employee_ids=world["employee_ids"],
    )

    epf_numbers = {emp.fixture_id for emp in fixture.employees if emp.regime == "epf"}
    assert len(epf_numbers) == 4

    dto = await nps_contribution_builder.build(session, _ctx(world))
    assert dto.report_type == REPORT_TYPE_NPS
    section = _section(dto)
    assert len(section.rows) == 12

    emp_idx = _col_index(section, "employee_contribution")
    er_idx = _col_index(section, "employer_contribution")
    total_idx = _col_index(section, "total")
    pran_idx = _col_index(section, "pran")
    assert section.totals is not None

    employee_total = _dec(section.totals[emp_idx])
    employer_total = _dec(section.totals[er_idx])
    combined_total = _dec(section.totals[total_idx])
    assert employee_total == _dec("109245.00")
    assert employer_total == _dec("152943.00")
    assert combined_total == _dec("262188.00")
    assert employee_total + employer_total == combined_total

    nps_ids = {row[0] for row in section.rows}
    assert nps_ids.isdisjoint(epf_numbers)

    for row in section.rows:
        emp_no = row[0]
        assert profiles[emp_no]["retirement_regime"] == "nps"
        assert row[pran_idx], f"missing PRAN for {emp_no}"
        assert str(row[pran_idx]).startswith("9000")
        assert _dec(row[emp_idx]) + _dec(row[er_idx]) == _dec(row[total_idx])


@pytest.mark.asyncio
async def test_gpf_excel_round_trip_and_pdf_contains_title_total(session):
    world = await _retirement_june_world(session)
    await _bind(session, world["org_id"], world["user_id"])
    dto = await gpf_mumbai_builder.build(session, _ctx(world))
    section = _section(dto)
    sub_idx = _col_index(section, "subscription")
    total = _dec(section.totals[sub_idx])
    assert total == _dec("165000.00")

    xlsx = retirement_to_excel(dto)
    wb = load_workbook(BytesIO(xlsx))
    ws = wb.active
    assert ws is not None
    # Header row 5; subscription is column 4; first data row 6.
    money_cell = ws.cell(row=6, column=4)
    assert money_cell.number_format == MONEY_FORMAT
    totals_excel_row = 6 + len(section.rows)
    total_cell = ws.cell(row=totals_excel_row, column=4)
    assert total_cell.number_format == MONEY_FORMAT
    assert _dec(total_cell.value) == _dec("165000.00")

    pdf = retirement_to_pdf(dto)
    assert pdf.startswith(b"%PDF")
    text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(pdf)).pages)
    assert "Mumbai" in text
    assert "1,65,000.00" in text


@pytest.mark.asyncio
async def test_unposted_run_raises_conflict(session):
    """Unposted run uses a lane-local org so shared-DB truncates cannot orphan FKs."""
    if session.in_transaction():
        await session.rollback()

    org = await seed_organization(
        session,
        name="Retirement Unposted Org",
        slug=f"ret-unposted-{uuid4().hex[:10]}",
    )
    user = await seed_user(session, workos_user_id=f"ret_unposted_{uuid4().hex[:10]}")
    org_id = org.id
    user_id = user.id
    await session.commit()

    await _bind(session, org_id, user_id)
    period = PayrollPeriod(
        organization_id=org_id,
        period_year=2026,
        period_month=8,
        status="open",
    )
    session.add(period)
    await session.flush()
    draft = PayrollRun(
        organization_id=org_id,
        period_id=period.id,
        status="draft",
    )
    session.add(draft)
    await session.flush()
    draft_id = draft.id
    await session.commit()

    ctx = ReportContext(
        organization_id=org_id,
        posted_run_id=draft_id,
        template_version=TEMPLATE_VERSION,
        generated_at=datetime.now(UTC),
        engine_version="0.1.0",
    )

    await _bind(session, org_id, user_id)
    with pytest.raises(ConflictError, match="must be posted"):
        await gpf_mumbai_builder.build(session, ctx)

    await _bind(session, org_id, user_id)
    with pytest.raises(ConflictError, match="must be posted"):
        await gpf_nagpur_builder.build(session, ctx)

    await _bind(session, org_id, user_id)
    with pytest.raises(ConflictError, match="must be posted"):
        await nps_contribution_builder.build(session, ctx)


def test_register_retirement_reports_entries() -> None:
    registry = ReportRegistry()
    register_retirement_reports(registry)
    assert REPORT_TYPE_GPF_MUMBAI in registry
    assert REPORT_TYPE_GPF_NAGPUR in registry
    assert REPORT_TYPE_NPS in registry
    assert registry.get(REPORT_TYPE_GPF_MUMBAI).builder is gpf_mumbai_builder
    assert registry.get(REPORT_TYPE_GPF_NAGPUR).builder is gpf_nagpur_builder
    assert registry.get(REPORT_TYPE_NPS).builder is nps_contribution_builder
