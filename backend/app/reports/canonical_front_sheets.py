"""Canonical v3 renderers for the four payroll workbook front sheets.

These sheets deliberately consume normalized report DTOs rather than copying
cross-sheet formulas from the legacy workbook.  The legacy PaySlip sheet in
particular contains deleted references; rebuilding the formulas here keeps the
accepted visual roles without preserving its calculation errors.
"""

from __future__ import annotations

from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
from typing import Any, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.worksheet import Worksheet

from app.reports.amount_in_words import amount_in_words
from app.reports.base import ReportDTO, TableSection
from app.reports.excel import MONEY_FORMAT, sanitize_excel_text
from app.reports.canonical_schedules import clone_canonical_sheet_structure

_THIN = Side(style="thin", color="000000")
_MEDIUM = Side(style="medium", color="000000")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _section(dto: ReportDTO, title: str) -> TableSection | None:
    return next((item for item in dto.sections if item.title == title), None)


def _rows_as_dicts(section: TableSection | None) -> list[dict[str, Any]]:
    if section is None:
        return []
    keys = [column.key for column in section.columns]
    return [dict(zip(keys, row, strict=True)) for row in section.rows]


def _label_values(section: TableSection | None) -> dict[str, Any]:
    rows = _rows_as_dicts(section)
    result: dict[str, Any] = {}
    for row in rows:
        values = list(row.values())
        if len(values) >= 2:
            result[str(values[0])] = values[1]
    return result


def _money(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    return Decimal(str(value))


def _set_text(sheet: Worksheet, coordinate: str | tuple[int, int], value: Any) -> None:
    """Write untrusted presentation text without allowing Excel formulas."""
    cell = sheet[coordinate] if isinstance(coordinate, str) else sheet.cell(*coordinate)
    cell.value = sanitize_excel_text(str(value))


def _profile(dto: ReportDTO) -> Mapping[str, Any]:
    value = dto.metadata.get("report_profile", {})
    return value if isinstance(value, Mapping) else {}


def _run_metadata(dto: ReportDTO) -> Mapping[str, Any]:
    value = dto.metadata.get("run_metadata", {})
    return value if isinstance(value, Mapping) else {}


def _organization_label(dto: ReportDTO) -> str:
    profile = _profile(dto)
    return str(profile.get("legal_name") or profile.get("office_name") or dto.organization_name)


def _contact_block(dto: ReportDTO) -> str:
    profile = _profile(dto)
    address = profile.get("address_lines") or []
    lines = [str(item) for item in address if str(item).strip()]
    contacts = " | ".join(str(profile.get(key)) for key in ("phone", "website") if profile.get(key))
    if contacts:
        lines.append(contacts)
    if profile.get("cin"):
        lines.append(f"CIN: {profile['cin']}")
    return "\n".join(lines)


def _signatory(dto: ReportDTO, role: str) -> Mapping[str, Any]:
    aliases = {
        "final_approver": ("final_approver", "approving_officer"),
        "approving_officer": ("approving_officer", "final_approver"),
    }
    wanted = aliases.get(role, (role,))
    for item in _profile(dto).get("signatories") or []:
        if isinstance(item, Mapping) and str(item.get("role") or "") in wanted:
            return item
    return {}


def _signatory_text(dto: ReportDTO, role: str, fallback: str = "") -> str:
    item = _signatory(dto, role)
    name = str(item.get("name") or "").strip()
    designation = str(item.get("designation") or fallback).strip()
    return "\n".join(value for value in (name, designation) if value)


def _save(workbook: Workbook, title: str) -> bytes:
    workbook.properties.title = title
    workbook.properties.created = datetime.now(timezone.utc).replace(tzinfo=None)
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _new_sheet(title: str) -> tuple[Workbook, Worksheet]:
    workbook = Workbook()
    sheet = workbook.active
    clone_canonical_sheet_structure(sheet_name=title, target=sheet)
    return workbook, sheet


def _append_row_break_once(sheet: Worksheet, row: int) -> None:
    if row not in {int(item.id) for item in sheet.row_breaks.brk}:
        sheet.row_breaks.append(Break(id=row))


def _set_dimensions(
    sheet: Worksheet,
    widths: tuple[float, ...],
    heights: Mapping[int, float],
) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row, height in heights.items():
        sheet.row_dimensions[row].height = height


def _border_range(sheet: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _pay_bill_rows(pay_bill: ReportDTO) -> list[dict[str, Any]]:
    return _rows_as_dicts(_section(pay_bill, "Register"))


def _employee_net(row: Mapping[str, Any]) -> Decimal:
    gross = sum(
        (
            _money(row.get(key))
            for key in (
                "c_basic",
                "d_da",
                "e_cla",
                "f_hra",
                "g_wash_other",
                "h_other_reimbursement",
                "i_additional_allowance",
                "j_ta",
            )
        ),
        Decimal("0"),
    )
    gross += _money(row.get("l_employer_share")) - _money(row.get("m_recovery"))
    deductions = sum(
        (
            _money(row.get(key))
            for key in (
                "p_gpf",
                "q_pension_employer",
                "r_pension_employee",
                "s_advance",
                "t_flood",
                "u_income_tax",
                "v_insurance_gis",
                "w_hrr",
                "x_professional_tax",
                "y_co_op",
            )
        ),
        Decimal("0"),
    )
    return gross - deductions


def office_tip_to_excel(approval: ReportDTO, pay_bill: ReportDTO | None = None) -> bytes:
    """Render the canonical office-note role with a dynamic beneficiary table."""
    workbook, sheet = _new_sheet("office tip")
    if pay_bill is None:
        rows = _rows_as_dicts(_section(approval, "Beneficiaries"))
    else:
        rows = _pay_bill_rows(pay_bill)
    bill_ref = _label_values(_section(approval, "Bill reference"))
    signatories = _rows_as_dicts(_section(approval, "Signatories"))
    widths = (8, 5.1640625, 41.5, 31.5, 15.6640625, 9.1640625)
    _set_dimensions(sheet, widths, {3: 7.5, 4: 39.75, 5: 90.75, 6: 5.25, 7: 39.75})
    sheet.sheet_format.defaultRowHeight = 18

    note_number = str(bill_ref.get("Approval note No.") or "")
    note_date = str(bill_ref.get("Approval note date") or "")
    _set_text(sheet, "E1", note_number or "Administration / Payroll")
    _set_text(sheet, "E2", f"Date: {note_date}" if note_date else "Date:")
    sheet["C4"] = "Subject:"
    sheet.merge_cells("D4:E4")
    _set_text(sheet, "D4", f"Salary for {approval.subtitle}")
    sheet.merge_cells("C5:E5")
    _set_text(
        sheet,
        "C5",
        f"The following officers of {approval.organization_name} are proposed for salary "
        f"payment for {approval.subtitle}. The employee-wise amounts are set out below.",
    )
    sheet.merge_cells("C6:E6")
    headers = ("Sr. No.", "Name of beneficiary", "Designation", "Amount")
    for column, value in enumerate(headers, start=2):
        sheet.cell(7, column, value)
    for index, row in enumerate(rows, start=1):
        target = 7 + index
        sheet.cell(target, 2, index)
        _set_text(sheet, (target, 3), row.get("name") or "")
        _set_text(sheet, (target, 4), row.get("designation") or "")
        amount = row.get("net_payable") if pay_bill is None else _employee_net(row)
        sheet.cell(target, 5, float(_money(amount))).number_format = MONEY_FORMAT
        sheet.row_dimensions[target].height = 19
    total_row = 8 + len(rows)
    sheet.cell(total_row, 4, "Total Rs.")
    sheet.cell(total_row, 5, f"=SUM(E8:E{total_row - 1})").number_format = MONEY_FORMAT
    sheet.row_dimensions[total_row].height = 18
    _border_range(sheet, 7, total_row, 2, 5)

    prose_start = total_row + 2
    profile = _profile(approval)
    funding = ", ".join(
        value
        for value in (
            f"fund source {profile.get('fund_source')}" if profile.get("fund_source") else "",
            f"plan status {profile.get('plan_status')}" if profile.get("plan_status") else "",
        )
        if value
    )
    prose = (
        "Details of salary of the above officers are kept below. The net salary shown above is proposed for payment.",
        "The expenditure will be met from "
        + (funding or "the funds available to the organization")
        + " and adjusted under the applicable head of account.",
        "The approving authority is requested to approve the salary and allowance payment.",
        "Submitted for approval.",
    )
    prose_heights = (40.5, 56.25, 51, 37.5)
    for offset, (text, height) in enumerate(zip(prose, prose_heights, strict=True)):
        row = prose_start + offset
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        _set_text(sheet, (row, 2), text)
        sheet.row_dimensions[row].height = height
        sheet.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")

    signature_start = prose_start + 7
    slots = [
        (3, "maker", "Assistant Manager"),
        (8, "checker", "Accounts Officer"),
        (13, "approving_officer", "Chief Administrative Officer"),
    ]
    by_slot = {str(item.get("slot") or ""): item for item in signatories}
    for offset, slot, fallback in slots:
        target = signature_start + offset
        item = by_slot.get(slot, {})
        text = _signatory_text(approval, slot, fallback) or "\n".join(
            value
            for value in (
                str(item.get("name") or "").strip(),
                str(item.get("designation") or fallback).strip(),
            )
            if value
        )
        sheet.merge_cells(start_row=target, start_column=1, end_row=target, end_column=3)
        _set_text(sheet, (target, 1), text)
        sheet.cell(target, 1).font = Font(name="Times New Roman", size=14, bold=True)

    for cell in sheet[7]:
        cell.font = Font(name="Times New Roman", size=14, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=6):
        for cell in row:
            if cell.font.name == "Calibri":
                cell.font = Font(name="Times New Roman", size=14)
    sheet.print_area = f"A1:F{max(64, signature_start + 18)}"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.scale = 90
    sheet.page_margins.left = sheet.page_margins.right = 0
    sheet.page_margins.top = 0.39370078740157483
    sheet.page_margins.bottom = 0.1968503937007874
    sheet.page_margins.header = sheet.page_margins.footer = 0.11811023622047245
    sheet.print_options.horizontalCentered = True
    _append_row_break_once(sheet, prose_start + 1)
    return _save(workbook, approval.title)


def bank_tip_to_excel(dto: ReportDTO) -> bytes:
    """Render canonical bank-advice stationery and payment rows."""
    workbook, sheet = _new_sheet("Bank Tip")
    recipient = _label_values(_section(dto, "Advice recipient"))
    credits = _rows_as_dicts(_section(dto, "Payment credits"))
    widths = (2.6640625, 5.1640625, 40, 37, 28.5, 23.1640625, 24.1640625)
    heights = {
        1: 23.25,
        2: 20.25,
        3: 22,
        4: 34.5,
        5: 8.25,
        6: 24,
        7: 23,
        8: 23,
        9: 80,
        10: 23,
        11: 84,
        12: 12.75,
        13: 39.75,
    }
    _set_dimensions(sheet, widths, heights)
    sheet.sheet_format.defaultRowHeight = 13
    sheet.merge_cells("B1:C5")
    for target in ("D1:G1", "D2:G2", "D3:G3", "D4:G4"):
        sheet.merge_cells(target)
    _set_text(sheet, "D1", _organization_label(dto))
    sheet["D2"] = "(Payroll disbursement advice)"
    _set_text(sheet, "D3", dto.subtitle)
    _set_text(sheet, "D4", _contact_block(dto))
    sheet.merge_cells("B7:D7")
    profile = _profile(dto)
    run_metadata = _run_metadata(dto)
    reference_prefix = str(profile.get("salary_reference_prefix") or _organization_label(dto))
    advice_number = str(run_metadata.get("bank_advice_number") or "")
    advice_date = str(run_metadata.get("bank_advice_date") or "")
    _set_text(sheet, "B7", f"No. {advice_number or f'{reference_prefix}/Salary/{dto.subtitle}'}")
    _set_text(sheet, "F7", f"Date: {advice_date}" if advice_date else "Date:")
    sheet["B8"] = "To"
    _set_text(
        sheet,
        "C9",
        "\n".join(
            filter(
                None,
                (
                    str(recipient.get("Bank") or "The Branch Manager"),
                    str(recipient.get("Branch") or ""),
                    str(recipient.get("Address") or ""),
                ),
            )
        ),
    )
    sheet["D10"] = "Subject:"
    sheet.merge_cells("E10:G10")
    sheet["E10"] = "Authority for salary payment by electronic transfer"
    sheet.merge_cells("C11:G11")
    _set_text(
        sheet,
        "C11",
        f"Please credit the following employee accounts for {dto.subtitle}. "
        "The total is authorized from the organization's salary account.",
    )
    sheet.merge_cells("C12:G12")
    headers = (
        "Sr. No.",
        "Name of beneficiary",
        "Bank name and branch",
        "Account No.",
        "IFSC code",
        "Amount",
    )
    for column, value in enumerate(headers, start=2):
        sheet.cell(13, column, value)
    for index, row in enumerate(credits, start=1):
        target = 13 + index
        sheet.cell(target, 2, index)
        _set_text(sheet, (target, 3), row.get("name") or "")
        bank = ", ".join(
            filter(None, (str(row.get("bank_name") or ""), str(row.get("bank_branch") or "")))
        )
        _set_text(sheet, (target, 4), bank)
        _set_text(sheet, (target, 5), row.get("account_number") or "")
        _set_text(sheet, (target, 6), row.get("ifsc") or "")
        sheet.cell(target, 7, float(_money(row.get("disbursement")))).number_format = MONEY_FORMAT
    total_row = 14 + len(credits)
    sheet.cell(total_row, 5, "Total Rs.")
    sheet.cell(total_row, 7, f"=SUM(G14:G{total_row - 1})").number_format = MONEY_FORMAT
    sheet.row_dimensions[total_row].height = 25.5
    _border_range(sheet, 13, total_row, 2, 7)
    closing_row = total_row + 1
    sheet.cell(closing_row, 6, "Yours faithfully")
    sheet.merge_cells(
        start_row=closing_row + 1, start_column=5, end_row=closing_row + 1, end_column=7
    )
    _set_text(sheet, (closing_row + 1, 5), f"For {_organization_label(dto)}")
    signature_row = closing_row + 5
    sheet.merge_cells(start_row=signature_row, start_column=4, end_row=signature_row, end_column=5)
    sheet.merge_cells(start_row=signature_row, start_column=6, end_row=signature_row, end_column=8)
    _set_text(sheet, (signature_row, 4), _signatory_text(dto, "maker", "Financial Officer"))
    _set_text(sheet, (signature_row, 6), _signatory_text(dto, "approving_officer"))
    sheet.row_dimensions[signature_row].height = 87.75
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.font = Font(name="Arial", size=15, bold=cell.row in {1, 13, total_row})
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.print_area = f"A1:G{signature_row}"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.scale = 61
    sheet.page_setup.fitToHeight = 0
    sheet.page_margins.left = sheet.page_margins.right = 0.31496062992125984
    sheet.page_margins.top = 0.5511811023622047
    sheet.page_margins.bottom = 0.15748031496062992
    sheet.page_margins.header = sheet.page_margins.footer = 0.31496062992125984
    sheet.print_options.horizontalCentered = True
    _append_row_break_once(sheet, 33)
    return _save(workbook, dto.title)


_PAYSLIP_WIDTHS = (
    5.1640625,
    16.6640625,
    4.33203125,
    5.33203125,
    9.5,
    15.6640625,
    23,
    6.6640625,
    3.83203125,
    6.6640625,
    8.1640625,
    27.1640625,
    12.33203125,
    17.5,
    10.6640625,
    2.6640625,
    6.33203125,
    2.6640625,
    16.6640625,
    5,
    10.6640625,
    10.5,
    17.6640625,
    9.83203125,
)


def _merge_payslip_block(sheet: Worksheet, start: int) -> None:
    sheet.merge_cells(start_row=start, start_column=1, end_row=start + 16, end_column=1)
    for min_col, max_col in ((2, 4), (5, 7), (8, 10), (11, 12), (14, 15), (16, 19), (20, 24)):
        sheet.merge_cells(start_row=start, start_column=min_col, end_row=start, end_column=max_col)
    for min_col, max_col in ((2, 12), (13, 14), (15, 18), (19, 24)):
        sheet.merge_cells(
            start_row=start + 1, start_column=min_col, end_row=start + 1, end_column=max_col
        )
    for min_col, max_col in ((2, 7), (8, 16), (17, 24)):
        sheet.merge_cells(
            start_row=start + 2, start_column=min_col, end_row=start + 2, end_column=max_col
        )
    for row in range(start + 3, start + 14):
        for min_col, max_col in ((2, 5), (6, 7), (8, 11), (12, 13), (14, 16), (17, 21), (22, 24)):
            sheet.merge_cells(start_row=row, start_column=min_col, end_row=row, end_column=max_col)
    sheet.merge_cells(start_row=start + 14, start_column=2, end_row=start + 15, end_column=5)
    sheet.merge_cells(start_row=start + 14, start_column=6, end_row=start + 15, end_column=7)
    sheet.merge_cells(start_row=start + 14, start_column=8, end_row=start + 15, end_column=15)
    sheet.merge_cells(start_row=start + 14, start_column=16, end_row=start + 15, end_column=23)


def payslip_to_excel(dto: ReportDTO) -> bytes:
    """Render one canonical payslip block per employee with repaired formulas."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PaySlip"
    _set_dimensions(sheet, _PAYSLIP_WIDTHS, {})
    sheet.sheet_format.defaultRowHeight = 24
    for index, section in enumerate(dto.sections, start=1):
        start = 2 + (index - 1) * 19
        _merge_payslip_block(sheet, start)
        rows = _rows_as_dicts(section)
        identity = {
            str(row.get("code")): row.get("detail")
            for row in rows
            if row.get("line_kind") == "identity"
        }
        # The accepted PaySlip has emolument and recovery sections only.
        # Employer contributions remain visible on Pay Bill and their statutory
        # schedules; Amount Credited separately uses posted disbursement.
        earning_rows = [
            row for row in rows if row.get("line_kind") in {"earning", "gross_adjustment"}
        ]
        deduction_rows = [
            row
            for row in rows
            if row.get("line_kind") == "deduction" and not row.get("employer_transfer")
        ]
        government = [row for row in deduction_rows if row.get("detail") != "external_recovery"]
        nongovernment = [row for row in deduction_rows if row.get("detail") == "external_recovery"]
        overflow = {
            label: len(values)
            for label, values in (
                ("earnings", earning_rows),
                ("government recoveries", government),
                ("non-government recoveries", nongovernment),
            )
            if len(values) > 9
        }
        if overflow:
            raise ValueError(
                "Canonical payslip supports at most nine lines in each section; "
                f"employee {identity.get('employee_number') or index}: {overflow}."
            )
        net_rows = {str(row.get("code")): row for row in rows if row.get("line_kind") == "net"}
        sheet.cell(start, 1, index)
        sheet.cell(start, 2, "Employee Name:")
        _set_text(sheet, (start, 5), identity.get("name") or "")
        sheet.cell(start, 8, "Designation:")
        _set_text(sheet, (start, 11), identity.get("designation") or "")
        sheet.cell(start, 13, "Month:")
        _set_text(sheet, (start, 14), dto.subtitle)
        sheet.cell(start, 16, "GPF / Pension No.:")
        _set_text(sheet, (start, 20), identity.get("pran") or "")
        _set_text(sheet, (start + 1, 2), f"OFFICE OF {_organization_label(dto)}")
        _set_text(sheet, (start + 1, 13), f"Employee No. {identity.get('employee_number') or ''}")
        sheet.cell(start + 2, 2, "Emoluments")
        sheet.cell(start + 2, 8, "Government Recoveries")
        sheet.cell(start + 2, 17, "Non-Government Recoveries")
        for col, text in (
            (2, "Particulars"),
            (6, "Amount (Rs.)"),
            (8, "Particulars (Govt.)"),
            (12, "Amount (Rs.)"),
            (14, "Install No."),
            (17, "Particulars (Non Govt.)"),
            (22, "Amount (Rs.)"),
        ):
            sheet.cell(start + 3, col, text)
        detail_start = start + 4
        for offset in range(9):
            target = detail_start + offset
            earning = earning_rows[offset] if offset < len(earning_rows) else None
            govt = government[offset] if offset < len(government) else None
            non_govt = nongovernment[offset] if offset < len(nongovernment) else None
            if earning:
                _set_text(
                    sheet,
                    (target, 2),
                    str(earning.get("code") or "").replace("_", " ").title(),
                )
                sheet.cell(
                    target, 6, float(_money(earning.get("amount")))
                ).number_format = MONEY_FORMAT
            if govt:
                _set_text(
                    sheet,
                    (target, 8),
                    str(govt.get("code") or "").replace("_", " ").title(),
                )
                sheet.cell(
                    target, 12, float(_money(govt.get("amount")))
                ).number_format = MONEY_FORMAT
            if non_govt:
                _set_text(
                    sheet,
                    (target, 17),
                    str(non_govt.get("code") or "").replace("_", " ").title(),
                )
                sheet.cell(
                    target, 22, float(_money(non_govt.get("amount")))
                ).number_format = MONEY_FORMAT
        total_row = start + 13
        sheet.cell(total_row, 2, "Total Emoluments")
        sheet.cell(
            total_row, 6, f"=SUM(F{detail_start}:F{total_row - 1})"
        ).number_format = MONEY_FORMAT
        sheet.cell(total_row, 8, "Total Government Recovery")
        sheet.cell(
            total_row, 12, f"=SUM(L{detail_start}:L{total_row - 1})"
        ).number_format = MONEY_FORMAT
        sheet.cell(total_row, 17, "Total Non-Government Recovery")
        sheet.cell(
            total_row, 22, f"=SUM(V{detail_start}:V{total_row - 1})"
        ).number_format = MONEY_FORMAT
        net_row = start + 14
        sheet.cell(net_row, 2, "Amount Credited")
        sheet.cell(
            net_row, 6, float(_money((net_rows.get("disbursement") or {}).get("amount")))
        ).number_format = MONEY_FORMAT
        words = str((net_rows.get("amount_in_words") or {}).get("detail") or "")
        _set_text(sheet, (net_row, 8), words)
        _set_text(sheet, (net_row, 16), _signatory_text(dto, "approving_officer"))
        sheet.row_dimensions[start].height = 47.25
        for row in range(start + 1, start + 14):
            sheet.row_dimensions[row].height = 24
        sheet.row_dimensions[total_row].height = 36
        sheet.row_dimensions[net_row].height = 54
        _border_range(sheet, start, net_row + 1, 1, 24)
        for row in range(start, net_row + 2):
            for cell in sheet[row][:24]:
                cell.font = Font(
                    name="Tahoma", size=12, bold=row in {start + 1, start + 2, total_row, net_row}
                )
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        if index % 2 == 0 and index != len(dto.sections):
            _append_row_break_once(sheet, start + 17)
    last_row = max(1, 2 + (len(dto.sections) - 1) * 19 + 16)
    sheet.print_area = f"A1:X{last_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.scale = 52
    sheet.page_margins.left = sheet.page_margins.right = 0.11811023622047245
    sheet.page_margins.top = sheet.page_margins.bottom = 0.5118110236220472
    sheet.page_margins.header = sheet.page_margins.footer = 0.5118110236220472
    return _save(workbook, dto.title)


_FACE_MERGES = (
    "C2:I2",
    "B3:C3",
    "D3:H3",
    "B4:C4",
    "H5:I5",
    "C6:D6",
    "H6:I6",
    "B7:C7",
    "B8:C8",
    "B9:C9",
    "B10:C10",
    "B11:C11",
    "B12:C12",
    "B13:C13",
    "B14:C14",
    "B15:C15",
    "B16:C16",
    "C17:G17",
    "C18:G18",
    "C19:G19",
    "C20:D20",
    "C21:G21",
    "C22:G22",
    "C23:G23",
    "C24:G24",
    "D25:G25",
    "D26:G26",
    "D27:G27",
    "D28:G28",
    "D29:G29",
    "C30:G30",
    "C31:G31",
    "C32:G32",
    "C33:G33",
    "C34:G34",
    "C35:G35",
    "C36:G36",
    "C37:G37",
    "C38:G38",
    "C39:G39",
    "C40:G40",
    "C41:G41",
    "C42:G42",
    "C43:G43",
    "C44:G44",
    "C45:G45",
    "C46:G46",
    "C47:G47",
    "C48:D48",
    "C49:G49",
    "C50:G50",
    "C51:G51",
    "C52:G52",
    "J49:J51",
)


def treasury_face_to_excel(dto: ReportDTO, pay_bill: ReportDTO | None = None) -> bytes:
    """Render the canonical treasury face, recomputed from normalized Pay Bill rows."""
    workbook, sheet = _new_sheet(" Face ")
    # The cloned canonical template already carries the accepted dimensions.
    for merged in _FACE_MERGES:
        sheet.merge_cells(merged)
    rows = [] if pay_bill is None else _pay_bill_rows(pay_bill)
    sums = {
        key: sum((_money(row.get(key)) for row in rows), Decimal("0"))
        for key in (
            "c_basic",
            "d_da",
            "e_cla",
            "f_hra",
            "g_wash_other",
            "h_other_reimbursement",
            "i_additional_allowance",
            "j_ta",
            "l_employer_share",
            "m_recovery",
            "p_gpf",
            "q_pension_employer",
            "r_pension_employee",
            "s_advance",
            "t_flood",
            "u_income_tax",
            "v_insurance_gis",
            "w_hrr",
            "x_professional_tax",
            "y_co_op",
        )
    }
    if pay_bill is None:
        allocations = _label_values(_section(dto, "Canonical column allocations"))
        for key in sums:
            sums[key] = _money(allocations.get(key))
    earnings = sum(
        (
            sums[key]
            for key in (
                "c_basic",
                "d_da",
                "e_cla",
                "f_hra",
                "g_wash_other",
                "h_other_reimbursement",
                "i_additional_allowance",
                "j_ta",
            )
        ),
        Decimal("0"),
    )
    summary = _label_values(_section(dto, "Treasury Face Summary"))
    header = _label_values(_section(dto, "Bill header"))
    sheet["I1"] = "MTR 19 (Rule 29(1))"
    sheet["C2"] = "Government treasury salary bill"
    sheet["B3"] = "Name of office"
    _set_text(sheet, "D3", _organization_label(dto))
    sheet["B4"] = "Month"
    _set_text(sheet, "D4", dto.subtitle)
    _set_text(sheet, "H4", f"Bill No. {header.get('Bill No.', '')}")
    profile = _profile(dto)
    run_metadata = _run_metadata(dto)
    _set_text(sheet, "H5", profile.get("administrative_department") or "")
    _set_text(
        sheet,
        "B15",
        " | ".join(
            filter(
                None,
                (
                    f"Fund source: {profile.get('fund_source')}"
                    if profile.get("fund_source")
                    else "",
                    f"Plan status: {profile.get('plan_status')}"
                    if profile.get("plan_status")
                    else "",
                ),
            )
        ),
    )
    _set_text(
        sheet,
        "B16",
        " | ".join(
            filter(
                None,
                (
                    f"DDO: {profile.get('ddo_name')} ({profile.get('ddo_code')})"
                    if profile.get("ddo_name") or profile.get("ddo_code")
                    else "",
                    f"Treasury code: {profile.get('treasury_code')}"
                    if profile.get("treasury_code")
                    else "",
                ),
            )
        ),
    )
    sheet["C6"] = "FOR THE TREASURY"
    sheet["H6"] = "HEAD OF ACCOUNT"
    for row, label, value in (
        (
            7,
            "Administrative Department",
            " | ".join(
                filter(
                    None,
                    (
                        str(profile.get("administrative_department") or ""),
                        f"Code {profile.get('department_code')}"
                        if profile.get("department_code")
                        else "",
                    ),
                )
            ),
        ),
        (9, "Demand No.", header.get("Demand No.", "")),
        (11, "Major Head", header.get("Major head", "")),
        (13, "Sub Head", header.get("Sub head", "")),
        (15, "Detailed Head", header.get("Detailed head", "")),
    ):
        sheet.cell(row, 7, label)
        _set_text(sheet, (row, 9), value)
    sheet["B17"], sheet["C17"], sheet["H17"], sheet["I17"] = (
        "Sr. No.",
        "Detailed Heads",
        "Col. No.",
        "Amount",
    )
    detail_rows = (
        (18, 1, "Total Pay (Gross Salary)", 10, earnings),
        (19, 2, "Festival advance / overpayment recovery", "11 / 1&2", sums["m_recovery"]),
        (20, 3, "Employer Share", 12, sums["l_employer_share"]),
        (21, 4, "Honorarium", "11 / 3", Decimal("0")),
    )
    for row, serial, label, column_no, amount in detail_rows:
        sheet.cell(row, 2, serial)
        sheet.cell(row, 3, label)
        sheet.cell(row, 8, column_no)
        sheet.cell(row, 9, float(amount)).number_format = MONEY_FORMAT
    sheet["B22"], sheet["C22"], sheet["H22"], sheet["I22"] = (
        4,
        "Gross Total (10 - 11 + 12)",
        12,
        "=I18-I19+I20",
    )
    sheet["C23"] = "(A) Deductions adjustable by Accountant General"
    ag_lines = (
        (
            25,
            5,
            "8005 General Provident Fund subscription / refund / arrears",
            "14b",
            sums["p_gpf"],
        ),
        (31, 8, "7610 House Building / Motor / Other Advance", "16 / 1", sums["s_advance"]),
        (35, 12, "Other Accountant General recoveries", "16 / 4", Decimal("0")),
    )
    for row, serial, label, col, amount in ag_lines:
        sheet.cell(row, 2, serial)
        sheet.cell(row, 3, label)
        sheet.cell(row, 8, col)
        sheet.cell(row, 9, float(amount)).number_format = MONEY_FORMAT
    sheet["B36"], sheet["C36"], sheet["I36"] = 13, "Total (A)", "=SUM(I25:I35)"
    sheet["C37"] = "(B) Deductions adjustable by Treasury"
    treasury_lines = (
        (38, 14, "0021 Income Tax", 18, sums["u_income_tax"]),
        (41, 17, "8011 Insurance / GIS", "19 / 4", sums["v_insurance_gis"]),
        (42, 18, "0216 House Rent / Service Charges", 20, sums["w_hrr"]),
        (44, 20, "0028 Professional Tax", 21, sums["x_professional_tax"]),
        (45, 21, "Co-operative Recovery", 22, sums["y_co_op"]),
        (46, 22, "Flood-Affected Advance", "16 / 1&2", sums["t_flood"]),
        (
            47,
            23,
            "8342 Defined Contribution Pension Scheme",
            15,
            sums["q_pension_employer"] + sums["r_pension_employee"],
        ),
    )
    for row, serial, label, col, amount in treasury_lines:
        sheet.cell(row, 2, serial)
        sheet.cell(row, 3, label)
        sheet.cell(row, 8, col)
        sheet.cell(row, 9, float(amount)).number_format = MONEY_FORMAT
    sheet["B49"], sheet["C49"], sheet["I49"] = 24, "Total (B)", "=SUM(I37:I48)"
    sheet["B50"], sheet["C50"], sheet["I50"] = 25, "Undisbursed pay refund", 0
    sheet["B51"], sheet["C51"], sheet["H51"], sheet["I51"] = (
        26,
        "Total Deductions (A + B)",
        24,
        "=I36+I49",
    )
    sheet["B52"], sheet["C52"], sheet["H52"], sheet["I52"] = (
        27,
        "Net Payable Amount",
        25,
        "=I22-I51",
    )
    net = (
        earnings
        - sums["m_recovery"]
        + sums["l_employer_share"]
        - sum(
            (
                sums[key]
                for key in (
                    "p_gpf",
                    "q_pension_employer",
                    "r_pension_employee",
                    "s_advance",
                    "t_flood",
                    "u_income_tax",
                    "v_insurance_gis",
                    "w_hrr",
                    "x_professional_tax",
                    "y_co_op",
                )
            ),
            Decimal("0"),
        )
    )
    if pay_bill is None:
        net = _money(summary.get("Net payable"))
    sheet["B67"] = f"Rupees in words: {amount_in_words(net)}"
    sheet["G67"] = "=I22"
    sheet["B69"], sheet["G69"] = "Net Amount Required for Payment", "=I52"
    sheet["B70"] = "Brought forward from first page"
    sheet["B71"] = f"Rupees in words: {amount_in_words(net)}"
    for row, label, formula in (
        (74, "In Cash", "=G69"),
        (75, "By RTGS / Bank Draft", "=G69"),
        (76, "By Adjustment (AG)", "=I36"),
        (77, "By Adjustment (Treasury)", "=I49"),
        (78, "By Adjustment (Total)", "=SUM(D76:D77)"),
    ):
        sheet.cell(row, 2, label)
        sheet.cell(row, 4, formula)
    static = {
        80: "Certified that all emoluments included in this bill have been checked and are payable to the proper persons.",
        86: "Details of pay of absentee refunded",
        87: "Section / Establishment | Name | Period | Rupees | Financial Year",
        93: "FOR USE OF TREASURY",
        95: "Pay (in words)",
        99: "Paid by transfer credit",
        100: "As detailed below",
        101: "0021 Taxes on Income",
        102: "0028 Tax on Employment",
        103: "0216 Housing",
        104: "6216 Loans for Housing",
        105: "8011 Insurance and Pension Fund",
        107: "8342 Other Deposits (NPS)",
        109: "Total",
        114: "Date:",
        116: "Head Accountant",
        118: "Deputy Accountant",
        121: "Auditor / Section Officer / Accounts Officer",
    }
    for row, text in static.items():
        sheet.cell(row, 2, text)
    if run_metadata.get("bill_date"):
        _set_text(sheet, "B114", f"Date: {run_metadata['bill_date']}")
    _set_text(sheet, "B116", _signatory_text(dto, "maker"))
    _set_text(sheet, "B118", _signatory_text(dto, "checker"))
    _set_text(sheet, "B121", _signatory_text(dto, "approving_officer"))
    if profile.get("pay_bill_footer_text"):
        _set_text(sheet, "B80", profile["pay_bill_footer_text"])
    _border_range(sheet, 7, 52, 2, 9)
    for row in sheet.iter_rows(min_row=1, max_row=126, min_col=1, max_col=13):
        for cell in row:
            cell.font = Font(
                name="Times New Roman",
                size=14,
                bold=cell.row in {3, 6, 17, 22, 23, 36, 37, 49, 51, 52, 93},
            )
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.print_area = "A1:U121"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.scale = 51
    sheet.page_margins.left = sheet.page_margins.right = 0.7480314960629921
    sheet.page_margins.top = sheet.page_margins.bottom = 0.5118110236220472
    sheet.page_margins.header = sheet.page_margins.footer = 0.5118110236220472
    _append_row_break_once(sheet, 52)
    return _save(workbook, dto.title)


FRONT_SHEET_NAMES = frozenset({"office tip", "Bank Tip", "PaySlip", " Face "})


def render_front_sheet(sheet_name: str, reports: Mapping[str, ReportDTO]) -> bytes:
    """Dispatch a canonical front-sheet renderer from the complete v3 DTO map."""
    pay_bill = reports["pay_bill"]
    if sheet_name == "office tip":
        return office_tip_to_excel(reports["approval_note"], pay_bill)
    if sheet_name == "Bank Tip":
        return bank_tip_to_excel(reports["bank_rtgs_advice"])
    if sheet_name == "PaySlip":
        return payslip_to_excel(reports["payslips"])
    if sheet_name == " Face ":
        return treasury_face_to_excel(reports["treasury_face"], pay_bill)
    raise KeyError(sheet_name)
