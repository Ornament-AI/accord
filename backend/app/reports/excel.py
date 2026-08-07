"""Generic safe Excel writer for tabular :class:`~app.reports.base.ReportDTO` values.

Formula-injection neutralization for Excel cells:
strip openpyxl ``ILLEGAL_CHARACTERS_RE``, then escape leading ``=``, ``+``, ``-``,
``@`` (including after leading whitespace/control chars) with a leading apostrophe.

Money cells are written as numbers (``float`` conversion at the cell boundary
only, from :class:`~decimal.Decimal` quantized to 2 decimal places) with the
Indian-grouping number format ``#,##,##0.00``.

Western ``#,##0.00`` is common elsewhere. Accord deliberately uses
``#,##,##0.00`` because payroll amounts are INR and display grouping follows
:mod:`app.reports.formatting` (Indian lakh/crore). Excel applies the custom
group sizes only to the trailing groups of this pattern — it is not a perfect
full Indian numbering system for arbitrarily large values, but it is the
standard openpyxl/Excel custom-format approach for INR-oriented exports and
matches operator expectations better than western thousands grouping.
"""

from __future__ import annotations

from datetime import date, datetime, timezone
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.reports.base import CellValue, ColumnKind, FormulaScope, ReportDTO, TableSection

_FORMULA_PREFIXES = ("=", "+", "-", "@")
_UNSAFE_LEADING_CHARS = frozenset({"\t", "\r", "\n", "\0"})
_TWO_PLACES = Decimal("0.01")

# Indian-grouping money format (see module docstring).
MONEY_FORMAT = "#,##,##0.00"
COUNT_FORMAT = "#,##0"
DATE_FORMAT = "dd-mmm-yyyy"
TEXT_FORMAT = "@"

HEADER_FILL = PatternFill("solid", fgColor="E2E8F0")
TOTAL_FILL = PatternFill("solid", fgColor="F8FAFC")
_BORDER_SIDE = Side(style="thin", color="CBD5E1")
BORDER = Border(
    left=_BORDER_SIDE,
    right=_BORDER_SIDE,
    top=_BORDER_SIDE,
    bottom=_BORDER_SIDE,
)

_MIN_COL_WIDTH = 8.0
_MAX_COL_WIDTH = 60.0


def leading_content_starts_with_formula_prefix(value: str) -> bool:
    """Return True when visible content after leading whitespace/control is formula-like."""
    index = 0
    while index < len(value):
        char = value[index]
        if char.isspace() or char in _UNSAFE_LEADING_CHARS:
            index += 1
            continue
        return value[index:].startswith(_FORMULA_PREFIXES)
    return False


def sanitize_excel_text(value: str) -> str:
    """Strip illegal characters, then neutralize spreadsheet formula injection."""
    cleaned = ILLEGAL_CHARACTERS_RE.sub("", value)
    if leading_content_starts_with_formula_prefix(cleaned):
        return f"'{cleaned}"
    return cleaned


def _excel_money_value(value: Decimal) -> float:
    return float(value.quantize(_TWO_PLACES, rounding=ROUND_HALF_UP))


def _cell_value_and_format(value: CellValue, kind: ColumnKind) -> tuple[Any, str | None]:
    if value is None:
        return None, None
    if kind is ColumnKind.MONEY:
        if not isinstance(value, Decimal):
            raise TypeError(f"money column requires Decimal, got {type(value).__name__}")
        return _excel_money_value(value), MONEY_FORMAT
    if kind is ColumnKind.COUNT:
        if isinstance(value, bool) or not isinstance(value, int):
            raise TypeError(f"count column requires int, got {type(value).__name__}")
        return value, COUNT_FORMAT
    if kind is ColumnKind.DATE:
        if isinstance(value, datetime):
            return value, DATE_FORMAT
        if isinstance(value, date):
            return value, DATE_FORMAT
        return sanitize_excel_text(str(value)), TEXT_FORMAT
    # text (and any other kind treated as text)
    if isinstance(value, Decimal):
        # Defensive: money-like value in a text column still written as text.
        return sanitize_excel_text(format(value.quantize(_TWO_PLACES), "f")), TEXT_FORMAT
    return sanitize_excel_text(str(value)), TEXT_FORMAT


def _style_header_row(ws: Worksheet, row_index: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_index, column=col)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")


def _safe_sheet_title(title: str, used: set[str]) -> str:
    cleaned = "".join("_" if c in r"[]:*?/\\" else c for c in title).strip() or "Sheet"
    cleaned = cleaned[:31]
    candidate = cleaned
    suffix = 2
    while candidate in used:
        tail = f"_{suffix}"
        candidate = f"{cleaned[: 31 - len(tail)]}{tail}"
        suffix += 1
    used.add(candidate)
    return candidate


def _write_section(ws: Worksheet, dto: ReportDTO, section: TableSection) -> None:
    ws["A1"] = sanitize_excel_text(dto.title)
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = sanitize_excel_text(f"{dto.organization_name} — {dto.subtitle}")
    ws["A3"] = sanitize_excel_text(
        f"{section.title} | {dto.report_type} | template {dto.template_version}"
    )

    header_row = 5
    col_count = len(section.columns)
    for col_idx, column in enumerate(section.columns, start=1):
        cell = ws.cell(
            row=header_row,
            column=col_idx,
            value=sanitize_excel_text(column.header),
        )
        cell.number_format = TEXT_FORMAT
    _style_header_row(ws, header_row, col_count)

    # Freeze so the styled header row stays visible while scrolling data.
    ws.freeze_panes = f"A{header_row + 1}"

    widths = [max(_MIN_COL_WIDTH, min(_MAX_COL_WIDTH, len(c.header) + 2)) for c in section.columns]

    data_start = header_row + 1
    for row_offset, row in enumerate(section.rows):
        excel_row = data_start + row_offset
        for col_idx, (column, value) in enumerate(zip(section.columns, row, strict=True), start=1):
            cell_value, number_format = _cell_value_and_format(value, column.kind)
            cell = ws.cell(row=excel_row, column=col_idx, value=cell_value)
            if number_format is not None:
                cell.number_format = number_format
            if column.kind is ColumnKind.MONEY:
                cell.alignment = Alignment(horizontal="right")
            display_len = len(str(cell_value)) if cell_value is not None else 0
            widths[col_idx - 1] = max(
                widths[col_idx - 1],
                min(_MAX_COL_WIDTH, max(_MIN_COL_WIDTH, display_len + 2)),
            )

    if section.totals is not None:
        totals_row = data_start + len(section.rows)
        for col_idx, (column, value) in enumerate(
            zip(section.columns, section.totals, strict=True), start=1
        ):
            cell_value, number_format = _cell_value_and_format(value, column.kind)
            cell = ws.cell(row=totals_row, column=col_idx, value=cell_value)
            cell.font = Font(bold=True)
            cell.fill = TOTAL_FILL
            cell.border = BORDER
            if number_format is not None:
                cell.number_format = number_format

    column_by_key = {column.key: index for index, column in enumerate(section.columns, start=1)}

    # A populated money footer is a trusted column aggregate. Keep exported
    # totals live unless the builder supplied a more specific relationship.
    if dto.template_version == "v2" and section.totals is not None:
        explicit_total_targets = {
            spec.target_key
            for spec in section.formulas
            if spec.scope in {FormulaScope.TOTALS, FormulaScope.COLUMN_TOTAL}
        }
        totals_row = data_start + len(section.rows)
        for col_idx, (column, value) in enumerate(
            zip(section.columns, section.totals, strict=True), start=1
        ):
            if (
                column.kind is ColumnKind.MONEY
                and value is not None
                and column.key not in explicit_total_targets
            ):
                letter = get_column_letter(col_idx)
                formula = (
                    f"=SUM({letter}{data_start}:{letter}{totals_row - 1})" if section.rows else "=0"
                )
                ws.cell(row=totals_row, column=col_idx).value = formula

    def expression(
        row_index: int, add_keys: tuple[str, ...], subtract_keys: tuple[str, ...]
    ) -> str:
        add_refs = [f"{get_column_letter(column_by_key[key])}{row_index}" for key in add_keys]
        subtract_refs = [
            f"{get_column_letter(column_by_key[key])}{row_index}" for key in subtract_keys
        ]
        value = "+".join(add_refs) if add_refs else "0"
        if subtract_refs:
            value += "-" + "-".join(subtract_refs)
        return f"={value}"

    for spec in section.formulas:
        if spec.target_key not in column_by_key:
            raise ValueError(f"Unknown formula target column: {spec.target_key}")
        unknown = (set(spec.add_keys) | set(spec.subtract_keys)) - set(column_by_key)
        if unknown:
            raise ValueError(f"Unknown formula source columns: {sorted(unknown)}")
        target_col = column_by_key[spec.target_key]
        if spec.scope is FormulaScope.ROWS:
            for row_index in range(data_start, data_start + len(section.rows)):
                ws.cell(row=row_index, column=target_col).value = expression(
                    row_index, spec.add_keys, spec.subtract_keys
                )
        elif spec.scope is FormulaScope.TOTALS:
            if section.totals is None:
                raise ValueError("Totals formula requires a totals row")
            row_index = data_start + len(section.rows)
            ws.cell(row=row_index, column=target_col).value = expression(
                row_index, spec.add_keys, spec.subtract_keys
            )
        elif spec.scope is FormulaScope.COLUMN_TOTAL:
            if section.totals is None:
                raise ValueError("Column-total formula requires a totals row")
            totals_row = data_start + len(section.rows)
            letter = get_column_letter(target_col)
            if section.rows:
                formula = f"=SUM({letter}{data_start}:{letter}{totals_row - 1})"
            else:
                formula = "=0"
            ws.cell(row=totals_row, column=target_col).value = formula

    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def to_excel(dto: ReportDTO) -> bytes:
    """Render a tabular report DTO to ``.xlsx`` bytes via openpyxl."""
    wb = Workbook()
    # Remove the default sheet; recreate from sections (or one empty sheet).
    default = wb.active
    wb.remove(default)

    used_titles: set[str] = set()
    sections = dto.sections or (TableSection(title=dto.title, columns=(), rows=()),)
    for section in sections:
        ws = wb.create_sheet(title=_safe_sheet_title(section.title or dto.title, used_titles))
        _write_section(ws, dto, section)

    wb.properties.title = dto.title
    wb.properties.created = datetime.now(timezone.utc).replace(tzinfo=None)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
