"""Composition of rendered reports into the canonical 18-sheet workbook."""

from __future__ import annotations

from copy import copy
from datetime import datetime, timezone
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

CANONICAL_PRODUCT_SHEETS: tuple[tuple[str, str, str], ...] = (
    ("office tip", "approval_note", "visible"),
    ("Bank Tip", "bank_rtgs_advice", "visible"),
    ("PaySlip", "payslips", "visible"),
    (" Face ", "treasury_face", "visible"),
    ("Pay Bill", "pay_bill", "visible"),
    ("Income Tax", "income_tax_schedule", "visible"),
    ("GPF-Nagpur", "gpf_nagpur_schedule", "visible"),
    ("P.T.", "professional_tax_schedule", "visible"),
    ("GPF-Mumbai", "gpf_mumbai_schedule", "visible"),
    ("GPF-IV", "gpf_advance_schedule", "visible"),
    ("GIS", "gis_schedule", "visible"),
    ("HBA Ad", "hba_schedule", "visible"),
    ("Motor car Ad", "motor_car_advance_schedule", "hidden"),
    ("Motor cycale Ad (2)", "motorcycle_advance_schedule", "hidden"),
    ("Pension Sub (2)", "nps_contribution_schedule", "visible"),
    ("Festival", "festival_advance_schedule", "hidden"),
    ("WORLI", "accommodation_worli_schedule", "visible"),
    ("Mumbai", "accommodation_mumbai_schedule", "visible"),
)


def _translate_formula(value: str, *, origin: str, target: str) -> str:
    try:
        return Translator(value, origin=origin).translate_formula(target)
    except (TypeError, ValueError):
        return value


def _copy_workbook_into_sheet(source_bytes: bytes, target: Worksheet) -> None:
    source = load_workbook(BytesIO(source_bytes), data_only=False)
    row_offset = 0
    for source_sheet in source.worksheets:
        if row_offset:
            row_offset += 2
        min_col, min_row, max_col, max_row = range_boundaries(source_sheet.calculate_dimension())
        for row in source_sheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        ):
            for source_cell in row:
                if isinstance(source_cell, MergedCell):
                    continue
                target_cell = target.cell(source_cell.row + row_offset, source_cell.column)
                if isinstance(target_cell, MergedCell):
                    continue
                value = source_cell.value
                if isinstance(value, str) and value.startswith("=") and row_offset:
                    target_coord = target_cell.coordinate
                    value = _translate_formula(
                        value,
                        origin=source_cell.coordinate,
                        target=target_coord,
                    )
                target_cell.value = value
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.border = copy(source_cell.border)
                    target_cell.alignment = copy(source_cell.alignment)
                    target_cell.protection = copy(source_cell.protection)
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
        for merged in source_sheet.merged_cells.ranges:
            target.merge_cells(
                start_row=merged.min_row + row_offset,
                start_column=merged.min_col,
                end_row=merged.max_row + row_offset,
                end_column=merged.max_col,
            )
        for index, dimension in source_sheet.row_dimensions.items():
            target_dimension = target.row_dimensions[index + row_offset]
            target_dimension.height = dimension.height
            target_dimension.hidden = dimension.hidden
            target_dimension.outlineLevel = dimension.outlineLevel
            target_dimension.collapsed = dimension.collapsed
        for letter, dimension in source_sheet.column_dimensions.items():
            target_dimension = target.column_dimensions[letter]
            target_dimension.min = dimension.min
            target_dimension.max = dimension.max
            target_dimension.width = dimension.width
            target_dimension.hidden = dimension.hidden
            target_dimension.outlineLevel = dimension.outlineLevel
            target_dimension.collapsed = dimension.collapsed
            target_dimension.bestFit = dimension.bestFit
        if row_offset == 0:
            target.sheet_view.showGridLines = source_sheet.sheet_view.showGridLines
            target.sheet_view.zoomScale = source_sheet.sheet_view.zoomScale
            target.sheet_view.zoomScaleNormal = source_sheet.sheet_view.zoomScaleNormal
            target.freeze_panes = source_sheet.freeze_panes
            target.sheet_format.defaultColWidth = source_sheet.sheet_format.defaultColWidth
            target.sheet_format.defaultRowHeight = source_sheet.sheet_format.defaultRowHeight
            target.page_setup = copy(source_sheet.page_setup)
            target.page_margins = copy(source_sheet.page_margins)
            target.print_options = copy(source_sheet.print_options)
            target.sheet_properties.pageSetUpPr = copy(source_sheet.sheet_properties.pageSetUpPr)
            if source_sheet.print_area:
                target.print_area = source_sheet.print_area.rsplit("!", 1)[-1]
            if source_sheet.print_title_rows:
                target.print_title_rows = source_sheet.print_title_rows
            if source_sheet.print_title_cols:
                target.print_title_cols = source_sheet.print_title_cols
            target.oddHeader = copy(source_sheet.oddHeader)
            target.oddFooter = copy(source_sheet.oddFooter)
            target.row_breaks = copy(source_sheet.row_breaks)
            target.col_breaks = copy(source_sheet.col_breaks)
        row_offset += source_sheet.max_row


def consolidate_v3_workbooks(
    workbooks: Iterable[tuple[str, str, bytes]],
) -> bytes:
    """Combine rendered report workbooks into the fixed canonical sheet pack."""
    by_name = {name: (state, content) for name, state, content in workbooks}
    expected = [name for name, _report_type, _state in CANONICAL_PRODUCT_SHEETS]
    missing = [name for name in expected if name not in by_name]
    if missing:
        raise ValueError(f"Canonical workbook is missing sheets: {missing!r}")

    # Each rendered source already contains its canonical structure.  Start
    # from empty sheets so dynamic row insertion cannot collide with stale
    # template cells or merge ranges left behind in a second copy.
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in expected:
        workbook.create_sheet(name)
    for name, _report_type, expected_state in CANONICAL_PRODUCT_SHEETS:
        state, content = by_name[name]
        sheet = workbook[name]
        _copy_workbook_into_sheet(content, sheet)
        sheet.sheet_state = expected_state if expected_state != "visible" else state

    workbook.properties.title = "Accord Canonical Payroll Reports"
    workbook.properties.created = datetime.now(timezone.utc).replace(tzinfo=None)
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
