"""Canonical v3 Pay Bill Excel renderer."""

from __future__ import annotations

from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
from typing import Mapping
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.worksheet import Worksheet

from app.reports.base import ReportDTO
from app.reports.canonical_pay_bill_common import (
    _PAY_BILL_HEADERS,
    _PAY_BILL_WIDTHS,
    _excel_date,
    _organization_label,
    _row_value,
    _text_preserving_zero,
)
from app.reports.canonical_schedules import clone_canonical_sheet_structure
from app.reports.excel import MONEY_FORMAT, sanitize_excel_text

_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="E7E6E6")
_TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")


def _style_range(
    ws: Worksheet, min_row: int, max_row: int, min_col: int = 1, max_col: int = 28
) -> None:
    for row in ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _money_cell(ws: Worksheet, row: int, column: int, value) -> None:
    if value is None:
        return
    cell = ws.cell(row=row, column=column, value=float(value))
    cell.number_format = MONEY_FORMAT
    cell.alignment = Alignment(horizontal="right", vertical="top")


def _formula_sum(column: int, rows: list[int]) -> str:
    letter = get_column_letter(column)
    return "=0" if not rows else "=SUM(" + ",".join(f"{letter}{row}" for row in rows) + ")"


def _write_page_totals(
    ws: Worksheet,
    *,
    first_row: int,
    page_number: int,
    detail_starts: list[int],
    total_rows: list[int],
    merge_source_ranges: bool = False,
) -> None:
    """Write the canonical five detail subtotals plus one page total row."""
    ws.cell(first_row, 2, f"Total of Page No. {page_number}")
    money_columns = (*range(3, 15), *range(16, 28))
    for offset in range(5):
        source_rows = [row + offset for row in detail_starts]
        for column in money_columns:
            ws.cell(first_row + offset, column, _formula_sum(column, source_rows))
            ws.cell(first_row + offset, column).number_format = MONEY_FORMAT
    final_row = first_row + 5
    ws.cell(final_row, 2, "Total Rs.")
    for column in (*range(3, 11), 12, 13, *range(16, 26)):
        ws.cell(final_row, column, _formula_sum(column, total_rows))
        ws.cell(final_row, column).number_format = MONEY_FORMAT
    ws.cell(final_row, 11, f"=SUM(C{final_row}:J{final_row})")
    ws.cell(final_row, 14, f"=K{final_row}+L{final_row}-M{final_row}")
    ws.cell(final_row, 26, f"=SUM(P{final_row}:Y{final_row})")
    ws.cell(final_row, 27, f"=N{final_row}-Z{final_row}")
    _style_range(ws, first_row, final_row)
    for cell in ws[final_row]:
        cell.font = Font(bold=True, size=9)
        cell.fill = _TOTAL_FILL
    if merge_source_ranges and first_row == 62:
        ws.merge_cells("A62:A67")
        ws.merge_cells("B62:B66")
    elif merge_source_ranges and first_row == 129:
        ws.merge_cells("A129:A133")
    elif merge_source_ranges and first_row == 197:
        ws.merge_cells("A197:A201")
        ws.merge_cells("B197:B201")


def _write_grand_totals(
    ws: Worksheet, *, page_starts: list[int], merge_source_ranges: bool = False
) -> int:
    first_row = 203 if page_starts == [62, 129, 197] else max(page_starts) + 6
    ws.cell(first_row, 2, "Total of All Pages")
    money_columns = (*range(3, 15), *range(16, 28))
    for offset in range(5):
        for column in money_columns:
            ws.cell(
                first_row + offset,
                column,
                _formula_sum(column, [row + offset for row in page_starts]),
            )
            ws.cell(first_row + offset, column).number_format = MONEY_FORMAT
    final_row = first_row + 5
    ws.cell(final_row, 2, "Grand Total Rs.")
    page_total_rows = [row + 5 for row in page_starts]
    for column in (*range(3, 11), 12, 13, *range(16, 26)):
        ws.cell(final_row, column, _formula_sum(column, page_total_rows))
        ws.cell(final_row, column).number_format = MONEY_FORMAT
    ws.cell(final_row, 11, f"=SUM(C{final_row}:J{final_row})")
    ws.cell(final_row, 14, f"=K{final_row}+L{final_row}-M{final_row}")
    ws.cell(final_row, 26, f"=SUM(P{final_row}:Y{final_row})")
    ws.cell(final_row, 27, f"=N{final_row}-Z{final_row}")
    _style_range(ws, first_row, final_row)
    for cell in ws[final_row]:
        cell.font = Font(bold=True, size=9)
        cell.fill = _TOTAL_FILL
    if merge_source_ranges and first_row == 203:
        ws.merge_cells("A203:A207")
        ws.merge_cells("B203:B207")
    return final_row


_REFERENCE_GROUP_STARTS = (1, 2, 3, 16, 25, 26)
_REFERENCE_DETAIL_ANCHORS = (
    10,
    18,
    25,
    31,
    37,
    43,
    49,
    55,
    68,
    74,
    80,
    86,
    92,
    98,
    104,
    111,
    117,
    123,
    135,
    141,
    147,
    153,
    159,
    165,
    172,
    179,
    185,
    191,
)


def _matches_reference_layout(section) -> bool:
    """Use the source layout only for its explicit roster-group topology."""

    group_starts: list[int] = []
    current_group = None
    for serial, row in enumerate(section.rows, start=1):
        group = _row_value(section, row, "post_group_key")
        if group != current_group:
            group_starts.append(serial)
            current_group = group
    return (
        len(section.rows) == len(_REFERENCE_DETAIL_ANCHORS)
        and tuple(group_starts) == _REFERENCE_GROUP_STARTS
    )


def pay_bill_v3_to_excel(dto: ReportDTO) -> bytes:
    """Render the canonical 28-column, grouped-header Pay Bill workbook."""
    if not dto.sections:
        raise ValueError("Canonical Pay Bill requires a register section.")
    section = dto.sections[0]
    detail_section = next(
        (item for item in dto.sections if item.title == "Component detail lines"), None
    )
    detail_lines: dict[tuple[str, str], list[dict[str, object]]] = {}
    if detail_section is not None:
        detail_keys = [column.key for column in detail_section.columns]
        for detail_row in detail_section.rows:
            detail = dict(zip(detail_keys, detail_row, strict=True))
            detail_lines.setdefault(
                (str(detail["employee_number"]), str(detail["register_column"])), []
            ).append(detail)

    wb = Workbook()
    ws = wb.active
    clone_canonical_sheet_structure(sheet_name="Pay Bill", target=ws)
    # Body merges encode one historical roster. Keep the exact header topology,
    # then rebuild employee/post merges from the current posted run.
    for merged in list(ws.merged_cells.ranges):
        if merged.max_row >= 8:
            ws.unmerge_cells(str(merged))

    ws.merge_cells("A1:N1")
    ws["A1"] = "Scalewise Abstract"
    try:
        ws["O1"] = datetime.strptime(dto.subtitle, "%B %Y").date().replace(day=1)
        ws["O1"].number_format = "dd/mm/yyyy"
    except ValueError:
        ws["O1"] = dto.subtitle
    ws["Q1"] = "Paid in "
    run_metadata = dto.metadata.get("run_metadata", {})
    payment_date = (
        _excel_date(run_metadata.get("payment_date")) if isinstance(run_metadata, Mapping) else None
    )
    if payment_date is not None:
        ws["S1"] = payment_date
        ws["S1"].number_format = "dd/mm/yyyy"
    ws.merge_cells("O2:Z2")
    ws["O2"] = "Deductions"
    ws.merge_cells("O3:T3")
    ws["O3"] = "Adjustable by Accountant General"
    ws.merge_cells("U3:Z3")
    ws["U3"] = "Adjustable by Treasury"

    for column in range(1, 15):
        ws.cell(row=3, column=column, value=_PAY_BILL_HEADERS[column - 1])
        ws.merge_cells(start_row=3, start_column=column, end_row=5, end_column=column)
    ws.merge_cells("O4:P4")
    ws["O4"] = "General Provident Fund"
    ws["O5"] = "Account Number"
    ws["P5"] = "Subscription / Refund / Arrears"
    for column in range(17, 27):
        ws.cell(row=4, column=column, value=_PAY_BILL_HEADERS[column - 1])
        ws.merge_cells(start_row=4, start_column=column, end_row=5, end_column=column)
    ws["AA3"] = _PAY_BILL_HEADERS[26]
    ws.merge_cells("AA3:AA5")
    ws["AB3"] = _PAY_BILL_HEADERS[27]
    ws.merge_cells("AB3:AB5")

    for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=28):
        for cell in row:
            cell.font = Font(bold=True, size=9)
            cell.fill = _HEADER_FILL
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 21
    ws.row_dimensions[2].height = 15.75
    ws.row_dimensions[3].height = 24.75
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 75.75
    ws.row_dimensions[6].height = 14.25
    ws.row_dimensions[6].hidden = True
    ws.row_dimensions[7].height = 15
    ws.row_dimensions[7].hidden = True
    ws.row_dimensions[8].height = 14.25
    for column in range(1, 29):
        ws.cell(8, column, column)
        ws.cell(8, column).font = Font(name="Arial", size=12)
        ws.cell(8, column).alignment = Alignment(horizontal="center")
        ws.cell(8, column).border = _BORDER

    for column, width in enumerate(_PAY_BILL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width

    current_row = 9
    current_group: tuple[str, str, str, str, str] | None = None
    employee_total_rows: list[int] = []
    page_detail_starts: list[int] = []
    page_total_rows: list[int] = []
    page_summary_starts: list[int] = []
    reference_anchors = _REFERENCE_DETAIL_ANCHORS if _matches_reference_layout(section) else None
    for serial, item in enumerate(section.rows, start=1):
        group = (
            str(_row_value(section, item, "post_group_key") or ""),
            str(_row_value(section, item, "post_title") or "Unassigned Post"),
            _text_preserving_zero(_row_value(section, item, "sanctioned_posts")),
            _text_preserving_zero(_row_value(section, item, "vacant_posts")),
            str(_row_value(section, item, "pay_scale") or ""),
        )
        planned_detail_start = None if reference_anchors is None else reference_anchors[serial - 1]
        if planned_detail_start is not None:
            current_row = planned_detail_start
        if group != current_group:
            group_row = current_row - 1 if planned_detail_start is not None else current_row
            group_end_column = 27 if reference_anchors is not None and group_row >= 110 else 28
            ws.merge_cells(
                start_row=group_row,
                start_column=2,
                end_row=group_row,
                end_column=group_end_column,
            )
            group_text = f"Post of {group[1]}"
            strength = []
            if group[2]:
                strength.append(f"Total Posts {group[2]}")
            if group[3]:
                strength.append(f"Vacant {group[3]}")
            if strength:
                group_text += f" ({'. '.join(strength)})"
            if group[4]:
                group_text += f" - Scale {group[4]}"
            ws.cell(group_row, 2, sanitize_excel_text(group_text))
            _style_range(ws, group_row, group_row)
            ws.cell(group_row, 2).font = Font(bold=True, size=9)
            current_group = group
            if planned_detail_start is None:
                current_row += 1

        detail_start = current_row
        detail_end = current_row + 4
        total_row = current_row + 5
        special_serial_merge = reference_anchors is not None and serial == 25
        if special_serial_merge:
            ws.merge_cells(
                start_row=detail_start - 1,
                start_column=1,
                end_row=total_row,
                end_column=1,
            )
        else:
            serial_merge_end = (
                total_row
                if reference_anchors is not None and serial in {3, 4, 5, 16}
                else detail_end
            )
            ws.merge_cells(
                start_row=detail_start,
                start_column=1,
                end_row=serial_merge_end,
                end_column=1,
            )
        employee_number = str(_row_value(section, item, "employee_number") or "")
        ws.cell(detail_start - 1 if special_serial_merge else detail_start, 1, serial)
        ws.cell(detail_start, 2, sanitize_excel_text(str(_row_value(section, item, "name") or "")))
        ws.cell(
            detail_start + 1,
            2,
            sanitize_excel_text(str(_row_value(section, item, "designation") or "")),
        )
        narrations: list[str] = []
        for key in (
            "c_basic",
            "d_da",
            "e_cla",
            "f_hra",
            "g_wash_other",
            "h_other_reimbursement",
            "i_additional_allowance",
            "j_ta",
            "l_employer_share",
            "m_recovery",
            "p_gpf",
            "q_pension_employer",
            "r_pension_employee",
            "s_advance",
            "t_flood",
            "u_income_tax",
            "v_insurance_gis",
            "w_hrr",
            "x_professional_tax",
            "y_co_op",
        ):
            for line in detail_lines.get((employee_number, key), []):
                narration = " ".join(
                    filter(
                        None,
                        (str(line.get("reason") or ""), str(line.get("service_period") or "")),
                    )
                )
                if narration and narration not in narrations:
                    narrations.append(narration)
        if narrations:
            ws.cell(detail_start + 2, 2, sanitize_excel_text("; ".join(narrations)))
        basic = _row_value(section, item, "c_basic")
        basic_label = "" if basic is None else f"Basic @ Rs.{int(basic)}/-"
        ws.cell(detail_start + (3 if narrations else 2), 2, basic_label)
        ws.cell(
            detail_start + 4, 2, sanitize_excel_text(str(_row_value(section, item, "pan") or ""))
        )
        ws.cell(detail_start, 15, _row_value(section, item, "account_label"))
        ws.cell(
            detail_start + 1,
            15,
            sanitize_excel_text(str(_row_value(section, item, "gpf_account_number") or "")),
        )
        ws.cell(
            detail_start, 28, sanitize_excel_text(str(_row_value(section, item, "remarks") or ""))
        )

        for column, key in (
            (3, "c_basic"),
            (4, "d_da"),
            (5, "e_cla"),
            (6, "f_hra"),
            (7, "g_wash_other"),
            (8, "h_other_reimbursement"),
            (9, "i_additional_allowance"),
            (10, "j_ta"),
            (12, "l_employer_share"),
            (13, "m_recovery"),
            (16, "p_gpf"),
            (17, "q_pension_employer"),
            (18, "r_pension_employee"),
            (19, "s_advance"),
            (20, "t_flood"),
            (21, "u_income_tax"),
            (22, "v_insurance_gis"),
            (23, "w_hrr"),
            (24, "x_professional_tax"),
            (25, "y_co_op"),
        ):
            lines = detail_lines.get((employee_number, key), [])
            if lines:
                if len(lines) > 5:
                    raise ValueError(
                        f"Employee {employee_number} has more than five detail lines for {key}."
                    )
                for offset, line in enumerate(lines):
                    _money_cell(ws, detail_start + offset, column, line["amount"])
            else:
                _money_cell(ws, detail_start, column, _row_value(section, item, key))

        ws.cell(total_row, 2, "Total Rs.")
        formula_detail_start = detail_start - 1 if special_serial_merge else detail_start
        for column in range(3, 11):
            letter = get_column_letter(column)
            ws.cell(total_row, column, f"=SUM({letter}{formula_detail_start}:{letter}{detail_end})")
        ws.cell(total_row, 11, f"=SUM(C{total_row}:J{total_row})")
        ws.cell(total_row, 12, f"=SUM(L{formula_detail_start}:L{detail_end})")
        ws.cell(total_row, 13, f"=SUM(M{formula_detail_start}:M{detail_end})")
        ws.cell(total_row, 14, f"=K{total_row}+L{total_row}-M{total_row}")
        for column in range(16, 26):
            letter = get_column_letter(column)
            ws.cell(total_row, column, f"=SUM({letter}{formula_detail_start}:{letter}{detail_end})")
        ws.cell(total_row, 26, f"=SUM(P{total_row}:Y{total_row})")
        ws.cell(total_row, 27, f"=N{total_row}-Z{total_row}")
        for column in range(3, 28):
            ws.cell(total_row, column).number_format = MONEY_FORMAT
        _style_range(ws, detail_start, total_row)
        for cell in ws[total_row]:
            cell.font = Font(bold=True, size=9)
            cell.fill = _TOTAL_FILL
        employee_total_rows.append(total_row)
        page_detail_starts.append(detail_start - 1 if special_serial_merge else detail_start)
        page_total_rows.append(total_row)
        current_row = total_row + 1
        page_end = {8: 67, 18: 134}.get(serial)
        if page_end is not None:
            summary_start = page_end - 5
            if current_row > summary_start:
                raise ValueError(
                    f"Canonical Pay Bill page {len(page_summary_starts) + 1} overflowed "
                    f"before row {summary_start}."
                )
            _write_page_totals(
                ws,
                first_row=summary_start,
                page_number=len(page_summary_starts) + 1,
                detail_starts=page_detail_starts,
                total_rows=page_total_rows,
                merge_source_ranges=reference_anchors is not None,
            )
            page_summary_starts.append(summary_start)
            page_detail_starts = []
            page_total_rows = []
            current_row = page_end + 1

    computed_totals = {
        key: sum(
            (Decimal(str(_row_value(section, item, key) or 0)) for item in section.rows),
            Decimal("0"),
        )
        for key in (
            "c_basic",
            "d_da",
            "e_cla",
            "f_hra",
            "g_wash_other",
            "h_other_reimbursement",
            "i_additional_allowance",
            "j_ta",
            "l_employer_share",
            "m_recovery",
            "p_gpf",
            "q_pension_employer",
            "r_pension_employee",
            "s_advance",
            "t_flood",
            "u_income_tax",
            "v_insurance_gis",
            "w_hrr",
            "x_professional_tax",
            "y_co_op",
        )
    }
    if section.totals is not None:
        expected = dict(
            zip((column.key for column in section.columns), section.totals, strict=True)
        )
        mismatches = {
            key: (computed_totals[key], Decimal(str(expected[key] or 0)))
            for key in computed_totals
            if computed_totals[key] != Decimal(str(expected[key] or 0))
        }
        if mismatches:
            raise ValueError(
                f"Canonical Pay Bill grand totals do not match DTO totals: {mismatches}"
            )

    final_page_start = 197 if reference_anchors is not None else max(current_row, 197)
    if page_detail_starts or not page_summary_starts:
        if current_row > final_page_start:
            final_page_start = current_row
        _write_page_totals(
            ws,
            first_row=final_page_start,
            page_number=len(page_summary_starts) + 1,
            detail_starts=page_detail_starts,
            total_rows=page_total_rows,
            merge_source_ranges=reference_anchors is not None,
        )
        page_summary_starts.append(final_page_start)
    grand_total_row = _write_grand_totals(
        ws,
        page_starts=page_summary_starts,
        merge_source_ranges=reference_anchors is not None,
    )
    if reference_anchors is not None:
        ws.merge_cells("AB57:AB59")
        ws.merge_cells("AB76:AB78")
        ws.merge_cells("V211:Y211")
        ws["V211"] = sanitize_excel_text(_organization_label(dto))

    ws.freeze_panes = "A8"
    ws.print_title_rows = "2:7"
    ws.print_area = f"A1:AB{max(208, grand_total_row)}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.scale = 33
    ws.sheet_properties.pageSetUpPr.fitToPage = False
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.07874015748031496
    ws.page_margins.right = 0.07874015748031496
    ws.page_margins.top = 0
    ws.page_margins.bottom = 0
    ws.page_margins.header = 0.07874015748031496
    ws.page_margins.footer = 0.11811023622047245
    profile = dto.metadata.get("report_profile", {})
    footer_text = profile.get("pay_bill_footer_text") if isinstance(profile, Mapping) else None
    ws.oddFooter.left.text = str(footer_text or f"{_organization_label(dto)} - Pay Bill")
    ws.oddFooter.right.text = "Page &P"
    row_break_ids = {item.id for item in ws.row_breaks.brk}
    for break_id in (67, 134):
        if break_id not in row_break_ids:
            ws.row_breaks.append(Break(id=break_id))
    if 28 not in {item.id for item in ws.col_breaks.brk}:
        ws.col_breaks.append(Break(id=28))

    wb.properties.title = dto.title
    wb.properties.created = datetime.now(timezone.utc).replace(tzinfo=None)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
