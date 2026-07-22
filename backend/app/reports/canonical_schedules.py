"""Canonical v3 schedule-sheet renderers.

These formatters use a PII-free structural workbook whose cells contain styles
only.  Every visible value and every formula is rebuilt from the immutable
report DTO; no canonical-source value is carried into an Accord export.
"""

from __future__ import annotations

from copy import copy
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Mapping

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.worksheet import Worksheet

from app.reports.amount_in_words import amount_in_words
from app.reports.base import ReportDTO, TableSection
from app.reports.excel import MONEY_FORMAT, sanitize_excel_text


TEMPLATE_PATH = Path(__file__).with_name("templates") / "canonical_schedule_structure.xlsx"

REPORT_SHEET_NAMES: dict[str, str] = {
    "income_tax_schedule": "Income Tax",
    "gpf_nagpur_schedule": "GPF-Nagpur",
    "professional_tax_schedule": "P.T.",
    "gpf_mumbai_schedule": "GPF-Mumbai",
    "gpf_advance_schedule": "GPF-IV",
    "gis_schedule": "GIS",
    "hba_schedule": "HBA Ad",
    "motor_car_advance_schedule": "Motor car Ad",
    "motorcycle_advance_schedule": "Motor cycale Ad (2)",
    "nps_contribution_schedule": "Pension Sub (2)",
    "festival_advance_schedule": "Festival",
    "accommodation_worli_schedule": "WORLI",
    "accommodation_mumbai_schedule": "Mumbai",
}


def _row(section: TableSection, values: tuple[Any, ...]) -> dict[str, Any]:
    return {column.key: value for column, value in zip(section.columns, values, strict=True)}


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%d-%m-%Y")
    return sanitize_excel_text(str(value))


def _profile(dto: ReportDTO) -> Mapping[str, Any]:
    value = dto.metadata.get("report_profile", {})
    return value if isinstance(value, Mapping) else {}


def _organization_label(dto: ReportDTO) -> str:
    profile = _profile(dto)
    return str(profile.get("legal_name") or profile.get("office_name") or dto.organization_name)


def _signatory_text(dto: ReportDTO, role: str) -> str:
    for item in _profile(dto).get("signatories") or []:
        if isinstance(item, Mapping) and item.get("role") == role:
            return "\n".join(
                value
                for value in (
                    str(item.get("name") or "").strip(),
                    str(item.get("designation") or "").strip(),
                )
                if value
            )
    return ""


def _month_label(value: str) -> str:
    try:
        parsed = datetime.strptime(value, "%B %Y")
    except ValueError:
        return value
    return parsed.strftime("%B-%y")


def _set(ws: Worksheet, coordinate: str, value: Any) -> None:
    ws[coordinate] = _text(value) if isinstance(value, (str, date, datetime)) else value


def _formula(ws: Worksheet, coordinate: str, formula: str) -> None:
    if not formula.startswith("="):
        raise ValueError("Excel formulas must start with '='.")
    ws[coordinate] = formula


def _money(ws: Worksheet, coordinate: str, value: Any) -> None:
    ws[coordinate] = 0 if value is None else value
    ws[coordinate].number_format = MONEY_FORMAT


def _grow_table(
    ws: Worksheet,
    *,
    row_count: int,
    capacity: int,
    insert_at: int,
    template_start: int,
    block_height: int = 1,
    print_title_rows: str,
) -> int:
    """Insert styled continuation rows and return the inserted row count.

    Excel repeats ``print_title_rows`` on every continuation page. This keeps
    the accepted form/header block exact without copying any source values.
    """
    extra_blocks = max(row_count - capacity, 0)
    if not extra_blocks:
        return 0
    inserted_rows = extra_blocks * block_height
    max_column = ws.max_column
    style_rows = []
    for offset in range(block_height):
        source_row = template_start + offset
        style_rows.append(
            [
                (
                    copy(ws.cell(source_row, column).font),
                    copy(ws.cell(source_row, column).fill),
                    copy(ws.cell(source_row, column).border),
                    copy(ws.cell(source_row, column).alignment),
                    copy(ws.cell(source_row, column).protection),
                    ws.cell(source_row, column).number_format,
                )
                for column in range(1, max_column + 1)
            ]
        )
    template_merges = [
        (item.min_row, item.min_col, item.max_row, item.max_col)
        for item in ws.merged_cells.ranges
        if template_start <= item.min_row and item.max_row < template_start + block_height
    ]
    trailing_merges = [
        (item.min_row, item.min_col, item.max_row, item.max_col)
        for item in list(ws.merged_cells.ranges)
        if item.min_row >= insert_at
    ]
    for min_row, min_col, max_row, max_col in trailing_merges:
        ws.unmerge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col,
        )
    ws.insert_rows(insert_at, amount=inserted_rows)
    for min_row, min_col, max_row, max_col in trailing_merges:
        ws.merge_cells(
            start_row=min_row + inserted_rows,
            start_column=min_col,
            end_row=max_row + inserted_rows,
            end_column=max_col,
        )
    for block in range(extra_blocks):
        target_start = insert_at + block * block_height
        for offset, row_styles in enumerate(style_rows):
            target_row = target_start + offset
            ws.row_dimensions[target_row].height = ws.row_dimensions[template_start + offset].height
            for column, styles in enumerate(row_styles, start=1):
                cell = ws.cell(target_row, column)
                (
                    cell.font,
                    cell.fill,
                    cell.border,
                    cell.alignment,
                    cell.protection,
                    cell.number_format,
                ) = styles
        for min_row, min_col, max_row, max_col in template_merges:
            delta = target_start - template_start
            ws.merge_cells(
                start_row=min_row + delta,
                start_column=min_col,
                end_row=max_row + delta,
                end_column=max_col,
            )

    if ws.print_area:
        area = ws.print_area.rsplit("!", 1)[-1].replace("'", "")
        min_col, min_row, max_col, max_row = range_boundaries(area)
        ws.print_area = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row + inserted_rows}"
        )
    ws.print_title_rows = print_title_rows
    continuation_pages = (row_count - 1) // capacity
    for page_index in range(continuation_pages):
        ws.row_breaks.append(Break(id=insert_at - 1 + page_index * capacity * block_height))
    return inserted_rows


def _common_heading(ws: Worksheet, *, dto: ReportDTO, month_cell: str, org_cell: str) -> None:
    _set(ws, month_cell, _month_label(dto.subtitle))
    _set(ws, org_cell, dto.organization_name)


def _income_tax(ws: Worksheet, dto: ReportDTO) -> None:
    section = dto.sections[0]
    inserted = _grow_table(
        ws,
        row_count=len(section.rows),
        capacity=25,
        insert_at=30,
        template_start=29,
        print_title_rows="1:4",
    )
    total_row = 30 + inserted
    footer_row = 32 + inserted
    first = _row(section, section.rows[0]) if section.rows else {}
    financial_year = _text(first.get("financial_year"))
    _set(ws, "A1", f"Schedule showing the recovery of Income Tax for {financial_year}")
    _set(ws, "A2", "For the Month:-")
    _set(ws, "C2", _month_label(dto.subtitle))
    _set(ws, "A3", "Name of the Office :-")
    _set(ws, "C3", dto.organization_name)
    for coordinate, label in zip(
        ("A4", "B4", "C4", "D4", "E4", "F4"),
        ("Sr. No.", "Employee Name", "Designation", "PAN No.", "Financial Year", "Income Tax"),
        strict=True,
    ):
        _set(ws, coordinate, label)
    for serial, values in enumerate(section.rows, start=1):
        item = _row(section, values)
        output_row = serial + 4
        _set(ws, f"A{output_row}", serial)
        for column, key in zip(
            "BCDE", ("name", "designation", "pan", "financial_year"), strict=True
        ):
            _set(ws, f"{column}{output_row}", item.get(key))
        _money(ws, f"F{output_row}", item.get("income_tax"))
    _set(ws, f"E{total_row}", "Total Rs.")
    _formula(ws, f"F{total_row}", f"=SUM(F5:F{total_row - 1})")
    _set(ws, f"D{footer_row}", _signatory_text(dto, "approving_officer"))


def _simple_statutory(
    ws: Worksheet,
    dto: ReportDTO,
    *,
    title: str,
    header_row: int,
    first_row: int,
    last_row: int,
    serial_column: str,
    name_column: str,
    designation_column: str,
    amount_column: str,
    amount_key: str,
    total_row: int,
    footer_cell: str,
) -> None:
    section = dto.sections[0]
    capacity = last_row - first_row + 1
    inserted = _grow_table(
        ws,
        row_count=len(section.rows),
        capacity=capacity,
        insert_at=total_row,
        template_start=last_row,
        print_title_rows=f"1:{header_row}",
    )
    last_row += inserted
    total_row += inserted
    footer_row = int("".join(character for character in footer_cell if character.isdigit()))
    footer_column = "".join(character for character in footer_cell if character.isalpha())
    footer_cell = f"{footer_column}{footer_row + inserted}"
    _set(ws, f"{serial_column}1", title)
    _set(ws, f"{serial_column}2", "For the Month :-")
    _set(ws, f"{designation_column}2", _month_label(dto.subtitle))
    _set(ws, f"{serial_column}3", "Name of the office :-")
    _set(ws, f"{designation_column}3", dto.organization_name)
    for column, label in (
        (serial_column, "Sr.No."),
        (name_column, "Name"),
        (designation_column, "Designation"),
        (amount_column, "Amount Rs."),
    ):
        _set(ws, f"{column}{header_row}", label)
    for serial, values in enumerate(section.rows, start=1):
        item = _row(section, values)
        output_row = first_row + serial - 1
        _set(ws, f"{serial_column}{output_row}", serial)
        _set(ws, f"{name_column}{output_row}", item.get("name"))
        _set(ws, f"{designation_column}{output_row}", item.get("designation"))
        _money(ws, f"{amount_column}{output_row}", item.get(amount_key))
    _set(ws, f"{designation_column}{total_row}", "Total Rs.")
    _formula(
        ws,
        f"{amount_column}{total_row}",
        f"=SUM({amount_column}{first_row}:{amount_column}{last_row})",
    )
    _set(ws, footer_cell, _signatory_text(dto, "approving_officer"))


def _professional_tax(ws: Worksheet, dto: ReportDTO) -> None:
    _simple_statutory(
        ws,
        dto,
        title="Schedule Showing the Recovery of Profession Tax",
        header_row=4,
        first_row=5,
        last_row=32,
        serial_column="B",
        name_column="C",
        designation_column="D",
        amount_column="E",
        amount_key="professional_tax",
        total_row=33,
        footer_cell="D35",
    )


def _gis(ws: Worksheet, dto: ReportDTO) -> None:
    _simple_statutory(
        ws,
        dto,
        title="Schedule Showing the recovery of Group Insurance Scheme",
        header_row=6,
        first_row=7,
        last_row=32,
        serial_column="B",
        name_column="C",
        designation_column="D",
        amount_column="E",
        amount_key="gis",
        total_row=33,
        footer_cell="D35",
    )


def _gpf(ws: Worksheet, dto: ReportDTO, *, jurisdiction: str) -> None:
    section = dto.sections[0]
    profile = _profile(dto)
    remittance_profiles = profile.get("gpf_remittance_profiles") or {}
    candidate = (
        remittance_profiles.get(jurisdiction.casefold(), {})
        if isinstance(remittance_profiles, Mapping)
        else {}
    )
    remittance = candidate if isinstance(candidate, Mapping) else {}
    organization_label = _organization_label(dto)
    first_row = 36
    last_row = 40 if jurisdiction == "Nagpur" else 45
    total_row = 41 if jurisdiction == "Nagpur" else 46
    words_row = 43 if jurisdiction == "Nagpur" else 48
    footer_row = 46 if jurisdiction == "Nagpur" else 51
    inserted = _grow_table(
        ws,
        row_count=len(section.rows),
        capacity=last_row - first_row + 1,
        insert_at=total_row,
        template_start=last_row,
        print_title_rows="27:34",
    )
    last_row += inserted
    total_row += inserted
    words_row += inserted
    footer_row += inserted
    _set(ws, "A2", "General Provident Fund schedule")
    _set(ws, "I2", "Form 150")
    _set(ws, "C5", "General Provident Fund remittance form")
    _set(ws, "C6", "Schedule of General Provident Fund deductions")
    _set(ws, "A7", "Important instructions")
    _set(ws, "B8", "Use this schedule only for the General Provident Fund accounts shown below.")
    _set(ws, "B9", "Enter the complete account number and jurisdiction for every subscriber.")
    _set(ws, "B10", "State the reason in Remarks when a subscription stops or changes.")
    _set(ws, "B12", "Identify new subscribers and transfers in Remarks.")
    _set(ws, "B14", "Prepare a separate schedule for each Accountant General jurisdiction.")
    _set(ws, "B16", f"Drawing and disbursing office: {organization_label}")
    _set(ws, "B17", f"Deductions from salary for {_month_label(dto.subtitle)}")
    _set(ws, "B18", "Accounts maintained by the Accountant General")
    _set(
        ws,
        "B20",
        remittance.get("office_name")
        or f"ACCOUNTANT GENERAL, MAHARASHTRA ({'II' if jurisdiction == 'Nagpur' else 'I'}), {jurisdiction.upper()}",
    )
    _set(
        ws,
        "B21",
        " | ".join(
            filter(
                None,
                (
                    ", ".join(remittance.get("address_lines") or []),
                    str(remittance.get("account_code") or ""),
                    str(remittance.get("authority_text") or ""),
                ),
            )
        ),
    )
    _set(ws, "A22", "Office of the Drawing and Disbursing Officer")
    _set(ws, "A23", organization_label)
    _set(ws, "A27", f"Drawing and Disbursing Officer: {_signatory_text(dto, 'approving_officer')}")
    _set(ws, "A29", "Deduction month:-")
    _set(ws, "C29", _month_label(dto.subtitle))
    _set(ws, "H29", "Due month:-")
    _set(ws, "J29", _month_label(dto.subtitle))
    headers = {
        "A30": "Sr. No.",
        "B30": "Account Number",
        "C30": "Employee Name",
        "D30": "Designation",
        "E30": "Basic Pay",
        "F30": "Monthly Subscription",
        "G30": "Advance Recovery",
        "H32": "Other Recovery",
        "I32": "Total Recovery",
        "J30": "Remarks",
    }
    for coordinate, value in headers.items():
        _set(ws, coordinate, value)
    for serial, values in enumerate(section.rows, start=1):
        item = _row(section, values)
        output_row = first_row + serial - 1
        _set(ws, f"A{output_row}", serial)
        _set(ws, f"B{output_row}", item.get("gpf_account_number"))
        _set(ws, f"C{output_row}", item.get("name"))
        _set(ws, f"D{output_row}", item.get("designation"))
        _money(ws, f"E{output_row}", item.get("basic_pay"))
        _money(ws, f"F{output_row}", item.get("subscription"))
        _money(ws, f"G{output_row}", item.get("advance_recovery"))
        _money(ws, f"H{output_row}", 0)
        _formula(ws, f"I{output_row}", f"=SUM(F{output_row}:G{output_row})")
    _set(ws, f"A{total_row}", "Grand Total :-")
    _formula(ws, f"I{total_row}", f"=SUM(I{first_row}:I{last_row})")
    _set(ws, f"B{words_row}", "In Words:")
    total = sum(
        (
            (item.get("subscription") or 0) + (item.get("advance_recovery") or 0)
            for item in (_row(section, values) for values in section.rows)
        ),
        start=Decimal("0"),
    )
    _set(ws, f"C{words_row}", amount_in_words(total))
    _set(ws, f"E{footer_row}", _signatory_text(dto, "approving_officer"))
    _set(ws, f"E{footer_row + 1}", dto.organization_name)


def _gpf_advance(ws: Worksheet, dto: ReportDTO) -> None:
    section = dto.sections[0]
    inserted = _grow_table(
        ws,
        row_count=len(section.rows),
        capacity=12,
        insert_at=23,
        template_start=22,
        print_title_rows="1:10",
    )
    total_row = 23 + inserted
    footer_row = 24 + inserted
    _set(ws, "A1", "Schedule showing General Provident Fund advance recovery")
    _set(ws, "A4", "For Month :-")
    _set(ws, "C4", _month_label(dto.subtitle))
    _set(ws, "A6", "Name of the office :-")
    _set(ws, "C6", dto.organization_name)
    for coordinate, label in {
        "A8": "Sr. No.",
        "B8": "Sanction Reference",
        "C8": "Name",
        "D8": "Scheduled Installment",
        "E8": "Recovery This Month",
        "F9": "Installments Recovered/Total",
        "G8": "Total",
    }.items():
        _set(ws, coordinate, label)
    for serial, values in enumerate(section.rows, start=1):
        item = _row(section, values)
        output_row = serial + 10
        _set(ws, f"A{output_row}", serial)
        _set(ws, f"B{output_row}", item.get("advance_reference"))
        _set(ws, f"C{output_row}", item.get("name"))
        _money(ws, f"D{output_row}", item.get("scheduled_installment_amount"))
        _money(ws, f"E{output_row}", item.get("installment_amount"))
        _set(ws, f"F{output_row}", item.get("installments_progress"))
        _formula(ws, f"G{output_row}", f"=E{output_row}")
    _set(ws, f"C{total_row}", "Total Rs.")
    _formula(ws, f"D{total_row}", f"=SUM(D11:D{total_row - 1})")
    _formula(ws, f"E{total_row}", f"=SUM(E11:E{total_row - 1})")
    _formula(ws, f"G{total_row}", f"=SUM(G11:G{total_row - 1})")
    _set(ws, f"E{footer_row}", _signatory_text(dto, "approving_officer"))


def _advance(
    ws: Worksheet,
    dto: ReportDTO,
    *,
    title: str,
    first_row: int,
    last_row: int,
    total_row: int,
    footer_cell: str,
    amount_header: str,
) -> None:
    section = dto.sections[0]
    capacity = last_row - first_row + 1
    insert_at = total_row - 1 if ws.title == "HBA Ad" else total_row
    inserted = _grow_table(
        ws,
        row_count=len(section.rows),
        capacity=capacity,
        insert_at=insert_at,
        template_start=last_row,
        print_title_rows=f"1:{first_row - 1}",
    )
    last_row += inserted
    total_row += inserted
    footer_row = int("".join(character for character in footer_cell if character.isdigit()))
    footer_column = "".join(character for character in footer_cell if character.isalpha())
    footer_cell = f"{footer_column}{footer_row + inserted}"
    _set(ws, "A1", title)
    _set(ws, "A2", "For the Month of :-")
    _set(ws, "C2", _month_label(dto.subtitle))
    _set(
        ws,
        "A3" if ws.title != "Motor car Ad" and ws.title != "Motor cycale Ad (2)" else "A4",
        "Name of office :-",
    )
    org_cell = "C3" if ws.title not in {"Motor car Ad", "Motor cycale Ad (2)"} else "C4"
    _set(ws, org_cell, dto.organization_name)
    header_row = first_row - 1
    for coordinate, label in {
        f"A{header_row}": "Sr. No.",
        f"B{header_row}": "Name",
        f"C{header_row}": "Designation",
        f"D{header_row}": "No. of Installment",
        f"E{header_row}": amount_header,
    }.items():
        _set(ws, coordinate, label)
    for serial, values in enumerate(section.rows, start=1):
        item = _row(section, values)
        output_row = first_row + serial - 1
        _set(ws, f"A{output_row}", serial)
        _set(ws, f"B{output_row}", item.get("name"))
        _set(ws, f"C{output_row}", item.get("designation"))
        _set(ws, f"D{output_row}", item.get("installments_progress"))
        _money(ws, f"E{output_row}", item.get("installment_amount"))
    total_label_cell = f"A{total_row}" if ws.title == "HBA Ad" else f"D{total_row}"
    _set(ws, total_label_cell, "Total Rs.")
    formula_last_row = total_row - 1 if ws.title == "HBA Ad" else last_row
    _formula(ws, f"E{total_row}", f"=SUM(E{first_row}:E{formula_last_row})")
    _set(ws, footer_cell, _signatory_text(dto, "approving_officer"))


def _hba(ws: Worksheet, dto: ReportDTO) -> None:
    _advance(
        ws,
        dto,
        title="Recovery Statement of House Building Advance",
        first_row=5,
        last_row=10,
        total_row=12,
        footer_cell="D14",
        amount_header="H.B. Advance Recovery Rs.",
    )
    _set(ws, "E2", "Head - 7610")


def _motor(ws: Worksheet, dto: ReportDTO, *, motorcycle: bool) -> None:
    _advance(
        ws,
        dto,
        title=f"Schedule Showing the recovery of Motor {'Cycle' if motorcycle else 'Car'} Advance",
        first_row=6,
        last_row=6,
        total_row=7,
        footer_cell="D9",
        amount_header=f"Motor {'Cycle' if motorcycle else 'Car'} Advance Recovery Rs.",
    )
    _set(ws, "D2", "Major Head - 7610")


def _festival(ws: Worksheet, dto: ReportDTO) -> None:
    _advance(
        ws,
        dto,
        title="Schedule Showing the recovery of Festival Advance",
        first_row=5,
        last_row=29,
        total_row=30,
        footer_cell="D32",
        amount_header="Festival Advance Recovery Rs.",
    )


def _nps(ws: Worksheet, dto: ReportDTO) -> None:
    section = dto.sections[0]
    profile = _profile(dto)
    inserted = _grow_table(
        ws,
        row_count=len(section.rows),
        capacity=10,
        insert_at=41,
        template_start=38,
        block_height=3,
        print_title_rows="1:10",
    )
    total_row = 41 + inserted
    _set(ws, "B1", "FORM - 2")
    _set(ws, "B3", "NPS contribution schedule for State Government employees")
    _set(ws, "B4", "Name of office :-")
    _set(ws, "B5", profile.get("legal_name") or profile.get("office_name") or dto.organization_name)
    _set(ws, "B6", "Region/Treasury/Sub Treasury Code")
    _set(ws, "F4", f"DDO: {profile.get('ddo_name') or ''} ({profile.get('ddo_code') or ''})")
    _set(ws, "H4", f"Dept. code: {profile.get('department_code') or ''}")
    _set(ws, "I4", f"Treasury: {profile.get('treasury_code') or ''}")
    headers = {
        "B7": "Sr. No.",
        "C7": "Pension Account No.",
        "C8": "Sevarth I.D. No.",
        "C9": "PRAN No.",
        "D7": "Name of Employees",
        "E7": "Month",
        "F7": "Basic Pay Rs.",
        "G7": "D.A. Rs.",
        "H7": "NPS Employee Contribution Rs.",
        "I7": "NPS Employer Contribution Rs.",
        "J7": "Remarks",
    }
    for coordinate, value in headers.items():
        _set(ws, coordinate, value)
    for serial, values in enumerate(section.rows, start=1):
        item = _row(section, values)
        output_row = 11 + (serial - 1) * 3
        _set(ws, f"B{output_row}", serial)
        _set(ws, f"C{output_row}", item.get("pension_account"))
        _set(ws, f"C{output_row + 1}", item.get("sevarth_id"))
        _set(ws, f"C{output_row + 2}", item.get("pran"))
        _set(ws, f"D{output_row}", item.get("name"))
        _set(ws, f"E{output_row}", item.get("month"))
        _money(ws, f"F{output_row}", item.get("basic_pay"))
        _money(ws, f"G{output_row}", item.get("dearness_allowance"))
        _money(ws, f"H{output_row}", item.get("employee_contribution"))
        _money(ws, f"I{output_row}", item.get("employer_contribution"))
        _set(ws, f"J{output_row}", item.get("remarks"))
    _formula(ws, f"H{total_row}", f"=SUM(H11:H{total_row - 1})")
    _formula(ws, f"I{total_row}", f"=SUM(I11:I{total_row - 1})")
    allocation_row = total_row + 1
    employee_head = profile.get("nps_employee_account_head") or "Employee contribution"
    employer_head = profile.get("nps_employer_account_head") or "Employer contribution"
    _set(ws, f"B{allocation_row}", employee_head)
    _money(ws, f"I{allocation_row}", 0)
    _set(ws, f"B{allocation_row + 1}", employer_head)
    _money(ws, f"I{allocation_row + 1}", 0)
    _set(ws, f"B{allocation_row + 2}", employee_head)
    _formula(ws, f"I{allocation_row + 2}", f"=H{total_row}")
    _set(ws, f"B{allocation_row + 3}", employer_head)
    _formula(ws, f"I{allocation_row + 3}", f"=I{total_row}")
    heads = profile.get("head_of_account") or {}
    grand_total_row = 46 + inserted
    amount_words_row = 47 + inserted
    _set(ws, f"B{grand_total_row}", "Grand Total")
    _formula(ws, f"I{grand_total_row}", f"=SUM(I{allocation_row}:I{allocation_row + 3})")
    _set(ws, f"B{amount_words_row}", "Total Amount Rupees")
    _formula(ws, f"C{amount_words_row}", f"=I{grand_total_row}")
    total = sum(
        (
            (item.get("employee_contribution") or 0) + (item.get("employer_contribution") or 0)
            for item in (_row(section, values) for values in section.rows)
        ),
        start=Decimal("0"),
    )
    _set(ws, f"D{amount_words_row}", amount_in_words(total))
    _set(ws, f"G{amount_words_row}", f"Major Head: {heads.get('major_head') or ''}")
    _set(ws, f"B{50 + inserted}", "CERTIFICATE")
    _set(
        ws,
        f"B{51 + inserted}",
        "Certified that the details in this schedule have been verified and are correct.",
    )
    _set(ws, f"G{55 + inserted}", _signatory_text(dto, "approving_officer"))
    _set(ws, f"G{56 + inserted}", f"Fund source: {profile.get('fund_source') or ''}")
    _set(ws, f"G{57 + inserted}", f"Plan status: {profile.get('plan_status') or ''}")


def _accommodation(ws: Worksheet, dto: ReportDTO, *, location: str) -> None:
    section = dto.sections[0]
    maximum = 1 if location == "Worli" else 3
    insert_at = 9 if location == "Worli" else 11
    inserted = _grow_table(
        ws,
        row_count=len(section.rows),
        capacity=maximum,
        insert_at=insert_at,
        template_start=8 if location == "Worli" else 10,
        print_title_rows="1:6",
    )
    last_column = "H" if location == "Worli" else "J"
    _set(ws, "A1", "0216 LICENCE FEE")
    _set(ws, "A2", f"Schedule showing the recovery of House Rent ({location})")
    _set(ws, "A3", "For the Month :-")
    _set(ws, "C3", _month_label(dto.subtitle))
    _set(ws, "A4", "Name of the office :-")
    _set(ws, "C4", dto.organization_name)
    header_values = [
        "Sr. No.",
        "Name",
        "Designation",
        "Address",
        "HRA",
        "House Rent",
        "House Rent Service Charges",
    ]
    if location == "Mumbai":
        header_values += ["Parking Charges", "Additional Parking Charges"]
    header_values.append("Total Rs.")
    for column, value in enumerate(header_values, start=1):
        _set(ws, f"{chr(64 + column)}5", value)
        _set(ws, f"{chr(64 + column)}6", column)
    for serial, values in enumerate(section.rows, start=1):
        item = _row(section, values)
        output_row = 7 + serial
        _set(ws, f"A{output_row}", serial)
        _set(ws, f"B{output_row}", item.get("name"))
        _set(ws, f"C{output_row}", item.get("designation"))
        _set(ws, f"D{output_row}", item.get("quarters_address"))
        _money(ws, f"E{output_row}", item.get("informational_foregone_hra"))
        _money(ws, f"F{output_row}", item.get("house_rent"))
        _money(ws, f"G{output_row}", item.get("service_charge"))
        recovery_keys = ["house_rent", "service_charge"]
        if location == "Mumbai":
            _money(ws, f"H{output_row}", item.get("parking_charge"))
            _money(ws, f"I{output_row}", item.get("additional_parking_charge"))
            recovery_keys += ["parking_charge", "additional_parking_charge"]
        explicit_total = sum(
            (Decimal(str(item.get(key) or 0)) for key in recovery_keys),
            start=Decimal("0"),
        )
        actual_total = Decimal(str(item.get("license_fee_actual") or 0))
        if explicit_total != actual_total:
            raise ValueError(
                "Accommodation recovery buckets do not reconcile to license_fee_actual: "
                f"employee {item.get('employee_number') or serial}, "
                f"buckets={explicit_total}, actual={actual_total}."
            )
        if location == "Mumbai":
            # Informational foregone HRA (E) is displayed but never recovered.
            _formula(ws, f"J{output_row}", f"=SUM(F{output_row}:I{output_row})")
        else:
            _formula(ws, f"H{output_row}", f"=SUM(F{output_row}:G{output_row})")
    if location == "Mumbai":
        total_row = 11 + inserted
        _set(ws, f"G{total_row}", "Total Rs.")
        _formula(ws, f"J{total_row}", f"=SUM(J8:J{total_row - 1})")
        _set(ws, f"D{13 + inserted}", _signatory_text(dto, "approving_officer"))
    else:
        _set(ws, f"F{10 + inserted}", _signatory_text(dto, "approving_officer"))
    if last_column not in {"H", "J"}:  # pragma: no cover - defensive invariant
        raise AssertionError("invalid accommodation layout")


_RENDERERS: dict[str, Callable[[Worksheet, ReportDTO], None]] = {
    "Income Tax": _income_tax,
    "P.T.": _professional_tax,
    "GIS": _gis,
    "GPF-Nagpur": lambda ws, dto: _gpf(ws, dto, jurisdiction="Nagpur"),
    "GPF-Mumbai": lambda ws, dto: _gpf(ws, dto, jurisdiction="Mumbai"),
    "GPF-IV": _gpf_advance,
    "HBA Ad": _hba,
    "Motor car Ad": lambda ws, dto: _motor(ws, dto, motorcycle=False),
    "Motor cycale Ad (2)": lambda ws, dto: _motor(ws, dto, motorcycle=True),
    "Pension Sub (2)": _nps,
    "Festival": _festival,
    "WORLI": lambda ws, dto: _accommodation(ws, dto, location="Worli"),
    "Mumbai": lambda ws, dto: _accommodation(ws, dto, location="Mumbai"),
}


def clone_canonical_sheet_structure(*, sheet_name: str, target: Worksheet) -> None:
    """Copy one PII-free structural template sheet into an empty worksheet."""
    # Reading ``target["A1"]`` creates a real empty cell in openpyxl.  That
    # phantom cell expands B-starting canonical schedules to A1 on save.
    if target._cells:  # noqa: SLF001 - openpyxl exposes no non-mutating emptiness check
        raise ValueError("Canonical structural target worksheet must be empty.")

    template_workbook = load_workbook(TEMPLATE_PATH, data_only=False)
    if sheet_name not in template_workbook.sheetnames:
        raise ValueError(f"No canonical structural template for {sheet_name!r}.")
    source = template_workbook[sheet_name]
    target.title = sheet_name
    min_column, min_row, max_column, max_row = range_boundaries(source.calculate_dimension())
    for row in source.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_column,
        max_col=max_column,
    ):
        for source_cell in row:
            target_cell = target[source_cell.coordinate]
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.protection = copy(source_cell.protection)
            target_cell.number_format = source_cell.number_format
    for merged_range in source.merged_cells.ranges:
        target.merge_cells(str(merged_range))
    for index, dimension in source.row_dimensions.items():
        target_dimension = target.row_dimensions[index]
        target_dimension.height = dimension.height
        target_dimension.hidden = dimension.hidden
        target_dimension.outlineLevel = dimension.outlineLevel
        target_dimension.collapsed = dimension.collapsed
    for index, dimension in source.column_dimensions.items():
        target_dimension = target.column_dimensions[index]
        target_dimension.min = dimension.min
        target_dimension.max = dimension.max
        target_dimension.width = dimension.width
        target_dimension.hidden = dimension.hidden
        target_dimension.outlineLevel = dimension.outlineLevel
        target_dimension.collapsed = dimension.collapsed
        target_dimension.bestFit = dimension.bestFit
    target.sheet_state = source.sheet_state
    target.sheet_view.showGridLines = source.sheet_view.showGridLines
    target.sheet_view.zoomScale = source.sheet_view.zoomScale
    target.sheet_view.zoomScaleNormal = source.sheet_view.zoomScaleNormal
    target.freeze_panes = source.freeze_panes
    target.sheet_format.defaultColWidth = source.sheet_format.defaultColWidth
    target.sheet_format.defaultRowHeight = source.sheet_format.defaultRowHeight
    target.page_setup = copy(source.page_setup)
    target.page_margins = copy(source.page_margins)
    target.print_options = copy(source.print_options)
    target.sheet_properties.pageSetUpPr = copy(source.sheet_properties.pageSetUpPr)
    if source.print_area:
        target.print_area = source.print_area.rsplit("!", 1)[-1]
    if source.print_title_rows:
        target.print_title_rows = source.print_title_rows
    if source.print_title_cols:
        target.print_title_cols = source.print_title_cols
    target.row_breaks = copy(source.row_breaks)
    target.col_breaks = copy(source.col_breaks)


def populate_canonical_schedule_sheet(
    target: Worksheet,
    dto: ReportDTO,
    *,
    sheet_name: str | None = None,
) -> None:
    """Clone and populate one canonical schedule in a composite workbook."""
    resolved_name = sheet_name or REPORT_SHEET_NAMES.get(dto.report_type)
    if resolved_name is None or resolved_name not in _RENDERERS:
        raise ValueError(f"No canonical schedule layout for {dto.report_type!r}.")
    if dto.template_version != "v3":
        raise ValueError("Canonical schedule layouts require template_version='v3'.")
    if not dto.sections:
        raise ValueError(f"{resolved_name} requires a schedule section.")
    clone_canonical_sheet_structure(sheet_name=resolved_name, target=target)
    _RENDERERS[resolved_name](target, dto)


def canonical_schedule_to_excel(dto: ReportDTO, *, sheet_name: str | None = None) -> bytes:
    """Render one v3 schedule using the canonical structural template."""
    resolved_name = sheet_name or REPORT_SHEET_NAMES.get(dto.report_type)
    if resolved_name is None or resolved_name not in _RENDERERS:
        raise ValueError(f"No canonical schedule layout for {dto.report_type!r}.")
    if dto.template_version != "v3":
        raise ValueError("Canonical schedule layouts require template_version='v3'.")
    if not dto.sections:
        raise ValueError(f"{resolved_name} requires a schedule section.")

    workbook = Workbook()
    worksheet = workbook.active
    populate_canonical_schedule_sheet(worksheet, dto, sheet_name=resolved_name)
    # A one-sheet workbook cannot be saved with its sole sheet hidden.  The
    # composite exporter reapplies the contract state after copying the sheet.
    worksheet.sheet_state = "visible"
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
