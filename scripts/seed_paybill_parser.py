"""Parse the canonical source workbook into typed employee seed rows."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from decimal import ROUND_CEILING, ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import httpx
import openpyxl


class SeedError(RuntimeError):
    pass


def clean(value: Any) -> str | None:
    if value is None:
        return None
    if not isinstance(value, str):
        return str(value).strip() or None
    text = " ".join(value.split()).strip()
    return text or None


def money(value: Any, *, context: str = "money value") -> int:
    """Parse a source money cell without hiding malformed nonblank values."""
    if value is None or value == "":
        return 0
    if isinstance(value, bool):
        raise SeedError(f"{context}: boolean is not a money value")
    text = str(value).replace(",", "").strip()
    if not text:
        return 0
    try:
        parsed = Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise SeedError(f"{context}: invalid nonblank money value {value!r}") from exc
    if not parsed.is_finite():
        raise SeedError(f"{context}: non-finite money value {value!r}")
    return int(parsed.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def optional_money(value: Any, *, context: str = "money value") -> int | None:
    """Preserve a source blank while validating every entered amount."""

    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    return money(value, context=context)


def money_str(value: int | Decimal) -> str:
    return f"{Decimal(value).quantize(Decimal('0.01'))}"


def norm_name(value: str) -> str:
    text = value.lower()
    for prefix in ("shri.", "shri ", "smt.", "smt ", "mr.", "mrs.", "ms."):
        if text.startswith(prefix):
            text = text[len(prefix) :]
            break
    return re.sub(r"[^a-z0-9]", "", text)


@dataclass
class Accommodation:
    location: str
    license_fee: int
    foregone_hra: int
    address: str | None = None
    house_rent: int | None = None
    service_charge: int | None = None
    parking_charge: int | None = None
    additional_parking_charge: int | None = None


@dataclass
class PayBillGroup:
    """One source Pay Bill heading, distinct from an employee's designation."""

    order: int
    heading: str
    sanctioned_strength: int | None = None
    vacant_count: int | None = None
    pay_scale: str | None = None
    owner_designation: str | None = None


@dataclass
class EmployeeRow:
    sr: int
    name: str
    designation: str
    pan: str | None
    basic: int
    da: int
    cla: int
    hra: int
    wash: int
    other: int
    additional_allowance: int
    ta: int
    gross: int
    gpf: int
    nps_employer: int
    nps_employee: int
    hba: int
    income_tax: int
    professional_tax: int
    gpf_account: str | None = None
    pension_account: str | None = None
    regime: str | None = None  # gpf_mumbai | gpf_nagpur | gpf | nps | epf
    sevarth_id: str | None = None
    pran: str | None = None
    epf_number: str | None = None
    gis: int = 0
    hba_installments: str | None = None
    accommodation: Accommodation | None = None
    bank_account: str | None = None
    ifsc: str | None = None
    bank_name: str | None = None
    bank_branch: str | None = None
    pay_bill_group: PayBillGroup | None = None
    extra: dict[str, Any] = field(default_factory=dict)


def _require(resp: httpx.Response, *, context: str) -> Any:
    if resp.status_code >= 400:
        raise SeedError(f"{context}: {resp.status_code} {resp.text}")
    if resp.status_code == 204 or not resp.content:
        return None
    return resp.json()


def _source_formula_amount(
    values_ws: Any,
    formula_ws: Any,
    *,
    row: int,
    column: int,
    overrides: dict[tuple[int, int], int],
    context: str,
) -> int:
    """Read a cached value or evaluate the workbook's simple ROUND formula."""
    key = (row, column)
    if key in overrides:
        return overrides[key]
    value = values_ws.cell(row, column).value
    try:
        return money(value, context=context)
    except SeedError:
        formula = formula_ws.cell(row, column).value
        if not isinstance(formula, str) or not formula.startswith("=ROUND"):
            raise
        compact = re.sub(r"\s+", "", formula.upper())
        rate_match = re.search(r"\*(\d+(?:\.\d+)?)%", compact)
        refs = re.findall(r"\$?([A-Z]+)\$?(\d+)", compact.split("*", 1)[0])
        if rate_match is None or not refs or not compact.endswith(",0)"):
            raise SeedError(f"{context}: unsupported source formula {formula!r}")
        from openpyxl.utils import column_index_from_string

        basis = sum(
            _source_formula_amount(
                values_ws,
                formula_ws,
                row=int(ref_row),
                column=column_index_from_string(ref_col),
                overrides=overrides,
                context=f"{context} reference {ref_col}{ref_row}",
            )
            for ref_col, ref_row in refs
        )
        raw = Decimal(basis) * Decimal(rate_match.group(1)) / Decimal("100")
        rounding = ROUND_CEILING if compact.startswith("=ROUNDUP") else ROUND_HALF_UP
        return int(raw.quantize(Decimal("1"), rounding=rounding))


def _block_amount(
    values_ws: Any,
    formula_ws: Any,
    *,
    start_row: int,
    total_row: int,
    column: int,
    overrides: dict[tuple[int, int], int],
    context: str,
) -> int:
    try:
        return money(values_ws.cell(total_row, column).value, context=context)
    except SeedError:
        formula = formula_ws.cell(total_row, column).value
        if not isinstance(formula, str) or not formula.startswith("="):
            raise
        return sum(
            _source_formula_amount(
                values_ws,
                formula_ws,
                row=row,
                column=column,
                overrides=overrides,
                context=f"{context} source row {row}",
            )
            for row in range(start_row, total_row)
        )


def _recover_basic(values_ws: Any, *, start_row: int, total_row: int, name: str) -> int:
    try:
        return money(values_ws.cell(start_row, 3).value, context=f"{name} basic pay")
    except SeedError as cell_error:
        for row in range(start_row, total_row):
            label = clean(values_ws.cell(row, 2).value) or ""
            match = re.search(r"\bBasic\s*@\s*Rs\.?\s*([0-9,]+)(?:/-)?", label, re.IGNORECASE)
            if match:
                return money(match.group(1), context=f"{name} basic-pay narration")
        raise SeedError(
            f"{name}: invalid basic cell and no Basic @ Rs... narration"
        ) from cell_error


def _parse_pay_bill_group(text: str, *, order: int) -> PayBillGroup:
    """Parse non-employee heading facts without inventing omitted source values."""

    heading_text = re.split(r"\(\s*Total\s+Posts", text, maxsplit=1, flags=re.IGNORECASE)[0]
    heading = re.sub(r"^Post\s+of\s+", "", heading_text, flags=re.IGNORECASE).strip()
    heading = " ".join(heading.split())
    # Normalize the accepted source typo; group identity comes from ``order``.
    if heading.casefold() == "assissant account officer":
        heading = "Assistant Accounts Officer"

    strength_match = re.search(
        r"Total\s+Posts\s*(\d+)?\s*\.?\s*Vacant\s*(\d+)?",
        text,
        flags=re.IGNORECASE,
    )
    sanctioned = (
        int(strength_match.group(1)) if strength_match and strength_match.group(1) else None
    )
    vacant = int(strength_match.group(2)) if strength_match and strength_match.group(2) else None
    scale_match = re.search(r"\bScale\s*[-:]?\s*([^)]*)", text, flags=re.IGNORECASE)
    pay_scale = clean(scale_match.group(1)) if scale_match else None
    return PayBillGroup(
        order=order,
        heading=heading,
        sanctioned_strength=sanctioned,
        vacant_count=vacant,
        pay_scale=pay_scale,
    )


def parse_paybill(path: Path) -> list[EmployeeRow]:
    wb = openpyxl.load_workbook(path, data_only=True)
    formula_wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["Pay Bill"]
    formula_ws = formula_wb["Pay Bill"]
    employees: list[EmployeeRow] = []
    pending_sr: int | None = None
    current_group: PayBillGroup | None = None
    group_count = 0

    r = 1
    while r <= ws.max_row:
        b = clean(ws.cell(r, 2).value)
        a = ws.cell(r, 1).value
        if b and b.startswith("Post of"):
            group_count += 1
            current_group = _parse_pay_bill_group(b, order=group_count)
            if isinstance(a, int):
                pending_sr = a
        looks_like_name = isinstance(b, str) and b.lower().startswith(("shri", "smt"))
        if looks_like_name:
            rr = r + 1
            while rr <= ws.max_row:
                row_label = clean(ws.cell(rr, 2).value)
                if row_label == "Total Rs.":
                    break
                if isinstance(ws.cell(rr, 1).value, int) or (
                    row_label and row_label.startswith("Post of")
                ):
                    rr -= 1
                    break
                rr += 1
            if rr > ws.max_row:
                raise SeedError(f"{b}: employee block has no Total Rs. row")
            name = b
            sr = int(a) if isinstance(a, int) else pending_sr or len(employees) + 1
            try:
                money(ws.cell(r, 3).value, context=f"{name} basic pay")
                basic_was_recovered = False
            except SeedError:
                basic_was_recovered = True
            base_basic = _recover_basic(ws, start_row=r, total_row=rr, name=name)
            overrides = {(r, 3): base_basic}
            amount = lambda column, label: _block_amount(  # noqa: E731
                ws,
                formula_ws,
                start_row=r,
                total_row=rr,
                column=column,
                overrides=overrides,
                context=f"{name} {label}",
            )
            designation = clean(ws.cell(r + 1, 2).value) or "Staff"
            if current_group is None:
                raise SeedError(f"{name}: employee row has no preceding Pay Bill group heading")
            if current_group.owner_designation is None:
                current_group.owner_designation = designation
            pan = None
            gpf_account = None
            for block_row in range(r, rr):
                cell = clean(ws.cell(block_row, 2).value)
                if cell and re.fullmatch(r"[A-Z]{5}[0-9]{4}[A-Z]", cell):
                    pan = cell
                account = clean(ws.cell(block_row, 15).value)
                if (
                    account
                    and account
                    not in {
                        "G.P.F.No.",
                        "Account No.",
                        "Pension A/C",
                        "Adjuesable by AG",
                    }
                    and any(ch.isdigit() for ch in account)
                ):
                    gpf_account = account

            basic = amount(3, "basic pay")
            if basic_was_recovered:
                basic = sum(
                    _source_formula_amount(
                        ws,
                        formula_ws,
                        row=block_row,
                        column=3,
                        overrides=overrides,
                        context=f"{name} basic source row {block_row}",
                    )
                    for block_row in range(r, rr)
                )
            da = amount(4, "DA")
            cla = amount(5, "CLA")
            hra = amount(6, "HRA")
            wash = amount(7, "wash allowance")
            other = amount(8, "other allowance")
            additional = amount(9, "additional allowance")
            ta = amount(10, "transport allowance")
            employees.append(
                EmployeeRow(
                    sr=sr,
                    name=name,
                    designation=designation,
                    pan=pan,
                    basic=basic,
                    da=da,
                    cla=cla,
                    hra=hra,
                    wash=wash,
                    other=other,
                    additional_allowance=additional,
                    ta=ta,
                    gross=sum((basic, da, cla, hra, wash, other, additional, ta)),
                    gpf=amount(16, "GPF"),
                    nps_employer=amount(17, "pension employer share"),
                    nps_employee=amount(18, "pension employee share"),
                    hba=amount(19, "advance recovery"),
                    income_tax=amount(21, "income tax"),
                    professional_tax=amount(24, "professional tax"),
                    gpf_account=gpf_account,
                    pay_bill_group=current_group,
                )
            )
            pending_sr = None
            r = rr + 1
            continue
        r += 1

    # Enrich from schedules and the employee-level bank advice.
    gpf_mumbai: dict[str, str] = {}
    gpf_nagpur: dict[str, str] = {}
    for sheet, bucket in (("GPF-Mumbai", gpf_mumbai), ("GPF-Nagpur", gpf_nagpur)):
        schedule = wb[sheet]
        for row_number, row in enumerate(schedule.iter_rows(min_row=1, values_only=True), start=1):
            if isinstance(row[0], int) and isinstance(row[2], str):
                money(row[4], context=f"{sheet}!E{row_number}")
                acct = clean(row[1])
                nm = clean(row[2])
                if nm and acct and not acct.startswith("ज्या"):
                    bucket[norm_name(nm)] = acct

    gis_map: dict[str, int] = {}
    schedule = wb["GIS"]
    for row_number, row in enumerate(schedule.iter_rows(min_row=1, values_only=True), start=1):
        if isinstance(row[1], int) and isinstance(row[2], str):
            gis_map[norm_name(clean(row[2]) or "")] = money(row[4], context=f"GIS!E{row_number}")

    it_map: dict[str, dict[str, Any]] = {}
    schedule = wb["Income Tax"]
    for row_number, row in enumerate(schedule.iter_rows(min_row=5, values_only=True), start=5):
        if isinstance(row[0], int) and isinstance(row[1], str):
            it_map[norm_name(clean(row[1]) or "")] = {
                "pan": clean(row[3]),
                "income_tax": money(row[5], context=f"Income Tax!F{row_number}"),
            }

    hba_map: dict[str, dict[str, Any]] = {}
    schedule = wb["HBA Ad"]
    for row_number, row in enumerate(schedule.iter_rows(min_row=1, values_only=True), start=1):
        if isinstance(row[0], int) and isinstance(row[1], str):
            hba_map[norm_name(clean(row[1]) or "")] = {
                "installments": clean(row[3]),
                "amount": money(row[4], context=f"HBA Ad!E{row_number}"),
            }

    nps_map: dict[str, dict[str, Any]] = {}
    schedule = wb["Pension Sub (2)"]
    paybill_by_name = {norm_name(employee.name): employee for employee in employees}
    r = 1
    while r <= schedule.max_row:
        if isinstance(schedule.cell(r, 2).value, int) and isinstance(
            schedule.cell(r, 4).value, str
        ):
            nm = clean(schedule.cell(r, 4).value) or ""
            key = norm_name(nm)
            paybill_employee = paybill_by_name.get(key)
            try:
                nps_ee = money(schedule.cell(r, 8).value, context=f"Pension Sub (2)!H{r}")
                nps_er = money(schedule.cell(r, 9).value, context=f"Pension Sub (2)!I{r}")
            except SeedError:
                if paybill_employee is None:
                    raise
                nps_ee = paybill_employee.nps_employee
                nps_er = paybill_employee.nps_employer
            nps_map[key] = {
                "pension_acct": clean(schedule.cell(r, 3).value),
                "nps_ee": nps_ee,
                "nps_er": nps_er,
                "sevarth": clean(schedule.cell(r + 1, 3).value)
                if r + 1 <= schedule.max_row
                else None,
                "pran": clean(schedule.cell(r + 2, 3).value) if r + 2 <= schedule.max_row else None,
            }
            r += 3
            continue
        r += 1

    bank_map: dict[str, dict[str, str | None]] = {}
    schedule = wb["Bank Tip"]
    for row_number, row in enumerate(schedule.iter_rows(min_row=14, values_only=True), start=14):
        if not isinstance(row[1], int) or not isinstance(row[2], str):
            continue
        name = clean(row[2])
        account = clean(row[4])
        ifsc = clean(row[5])
        bank_and_branch = clean(row[3])
        if not name or not account or not ifsc or not bank_and_branch:
            raise SeedError(f"Bank Tip row {row_number}: incomplete employee bank details")
        bank_map[norm_name(name)] = {
            "account": account,
            "ifsc": ifsc,
            # The source has one combined "Bank Name and Branch" column. Keep
            # it intact instead of guessing where the institution name ends.
            "bank_name": " ".join(bank_and_branch.split()),
            "branch": None,
        }

    acc_map: dict[str, Accommodation] = {}
    for sheet, location in (("WORLI", "worli"), ("Mumbai", "mumbai")):
        schedule = wb[sheet]
        for row_number, row in enumerate(schedule.iter_rows(min_row=1, values_only=True), start=1):
            if isinstance(row[0], int) and isinstance(row[1], str):
                house_rent = optional_money(row[5], context=f"{sheet}!F{row_number} house rent")
                service_charge = optional_money(
                    row[6], context=f"{sheet}!G{row_number} service charge"
                )
                parking_charge = (
                    optional_money(row[7], context=f"{sheet}!H{row_number} parking charge")
                    if location == "mumbai"
                    else None
                )
                additional_parking_charge = (
                    optional_money(
                        row[8],
                        context=f"{sheet}!I{row_number} additional parking charge",
                    )
                    if location == "mumbai"
                    else None
                )
                fee = sum(
                    value or 0
                    for value in (
                        house_rent,
                        service_charge,
                        parking_charge,
                        additional_parking_charge,
                    )
                )
                acc_map[norm_name(clean(row[1]) or "")] = Accommodation(
                    location=location,
                    license_fee=fee,
                    foregone_hra=money(row[4], context=f"{sheet}!E{row_number}"),
                    address=clean(row[3]),
                    house_rent=house_rent,
                    service_charge=service_charge,
                    parking_charge=parking_charge,
                    additional_parking_charge=additional_parking_charge,
                )

    wb.close()
    formula_wb.close()

    for emp in employees:
        key = norm_name(emp.name)
        if key in gpf_nagpur:
            emp.regime = "gpf_nagpur"
            emp.gpf_account = gpf_nagpur[key]
        elif key in gpf_mumbai:
            emp.regime = "gpf_mumbai"
            emp.gpf_account = gpf_mumbai[key]
        elif key in nps_map:
            emp.regime = "nps"
            info = nps_map[key]
            emp.sevarth_id = info.get("sevarth")
            emp.pran = info.get("pran")
            emp.pension_account = info.get("pension_acct")
            emp.gpf_account = None
            emp.nps_employee = int(info["nps_ee"])
            emp.nps_employer = int(info["nps_er"])
        elif emp.nps_employee > 0 and emp.nps_employee == emp.nps_employer:
            emp.regime = "epf"
        elif emp.nps_employee > 0 or emp.nps_employer > 0:
            emp.regime = "nps"
        elif emp.gpf_account:
            emp.regime = "gpf"
        elif emp.gpf > 0:
            raise SeedError(f"{emp.name}: GPF deduction has no source jurisdiction/account")
        else:
            raise SeedError(f"{emp.name}: retirement regime cannot be determined from the workbook")

        if key in gis_map:
            emp.gis = gis_map[key]
        if key in it_map:
            emp.income_tax = int(it_map[key]["income_tax"])
            if it_map[key]["pan"]:
                emp.pan = it_map[key]["pan"]
        if key in hba_map:
            emp.hba = int(hba_map[key]["amount"])
            emp.hba_installments = hba_map[key]["installments"]
        if key in acc_map:
            emp.accommodation = acc_map[key]
            emp.hra = 0
        if key in bank_map:
            bank = bank_map[key]
            emp.bank_account = bank["account"]
            emp.ifsc = bank["ifsc"]
            emp.bank_name = bank["bank_name"]
            emp.bank_branch = bank["branch"]

    employees.sort(key=lambda e: e.sr)
    return employees
