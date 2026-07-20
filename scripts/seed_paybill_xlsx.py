#!/usr/bin/env python3
"""Seed a running Accord org from the June 2026 MSIDC pay-bill spreadsheet.

Parses real employee / pay / deduction rows from the xlsx (not the synthetic
golden fixture), wipes existing master data for the active org, then posts via
the public HTTP APIs.

Usage:
  backend/.venv/bin/python scripts/seed_paybill_xlsx.py \\
    --xlsx "/Users/darshan/Downloads/Pay bill - June 2026 Regular Staff.xlsx"
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import Any
from uuid import UUID

import httpx
import openpyxl

EFFECTIVE_FROM = "2026-01-01"
PROFESSIONAL_TAX_DEFAULT = 200

COMPONENT_SPECS: list[tuple[str, str, str]] = [
	("BASIC", "Basic Pay", "earning"),
	("DA", "Dearness Allowance", "earning"),
	("CLA", "City Compensatory Allowance", "earning"),
	("HRA", "House Rent Allowance", "earning"),
	("WASH_ALLOWANCE", "Wash Allowance", "earning"),
	("TRANSPORT", "Transport / PTA", "earning"),
	("OTHER_ALLOWANCE", "Other Allowances", "earning"),
	("GPF_SUBSCRIPTION", "GPF Subscription", "ag_deduction"),
	("NPS_EMPLOYEE", "NPS Employee Contribution", "ag_deduction"),
	("NPS_EMPLOYER_TRANSFER", "NPS Employer Transfer", "ag_deduction"),
	("EPF_EMPLOYEE", "EPF Employee Contribution", "ag_deduction"),
	("EPF_EMPLOYER", "EPF Employer Contribution", "employer_contribution"),
	("EPF_EMPLOYER_TRANSFER", "EPF Employer Transfer", "ag_deduction"),
	("INCOME_TAX", "Income Tax", "treasury_deduction"),
	("PROFESSIONAL_TAX", "Professional Tax", "treasury_deduction"),
	("GIS", "Group Insurance Scheme", "treasury_deduction"),
	("HBA_INSTALLMENT", "House Building Advance", "external_recovery"),
	("ACCOMMODATION_LICENSE_FEE", "Accommodation License Fee", "external_recovery"),
]


class SeedError(RuntimeError):
	pass


def clean(value: Any) -> str | None:
	if value is None:
		return None
	if not isinstance(value, str):
		return str(value).strip() or None
	text = " ".join(value.split()).strip()
	return text or None


def money(value: Any) -> int:
	if value is None or value == "":
		return 0
	if isinstance(value, bool):
		return 0
	if isinstance(value, (int, float)):
		return int(round(value))
	if isinstance(value, str):
		text = value.replace(",", "").strip()
		if not text:
			return 0
		try:
			return int(round(float(text)))
		except ValueError:
			return 0
	return 0


def money_str(value: int | Decimal) -> str:
	return f"{Decimal(value).quantize(Decimal('0.01'))}"


def norm_name(value: str) -> str:
	text = value.lower()
	for prefix in ("shri.", "shri ", "smt.", "smt ", "mr.", "mrs.", "ms."):
		if text.startswith(prefix):
			text = text[len(prefix) :]
			break
	return re.sub(r"[^a-z0-9]", "", text)


@dataclass
class Accommodation:
	location: str
	license_fee: int
	foregone_hra: int
	address: str | None = None


@dataclass
class EmployeeRow:
	sr: int
	name: str
	designation: str
	pan: str
	basic: int
	da: int
	cla: int
	hra: int
	wash: int
	other: int
	ta: int
	gross: int
	gpf: int
	nps_employer: int
	nps_employee: int
	hba: int
	income_tax: int
	gpf_account: str | None = None
	regime: str = "gpf"  # gpf_mumbai | gpf_nagpur | nps | epf
	sevarth_id: str | None = None
	pran: str | None = None
	epf_number: str | None = None
	gis: int = 0
	professional_tax: int = PROFESSIONAL_TAX_DEFAULT
	hba_installments: str | None = None
	accommodation: Accommodation | None = None
	extra: dict[str, Any] = field(default_factory=dict)


def _require(resp: httpx.Response, *, context: str) -> Any:
	if resp.status_code >= 400:
		raise SeedError(f"{context}: {resp.status_code} {resp.text}")
	if resp.status_code == 204 or not resp.content:
		return None
	return resp.json()


def parse_paybill(path: Path) -> list[EmployeeRow]:
	wb = openpyxl.load_workbook(path, data_only=True)
	ws = wb["Pay Bill"]
	employees: list[EmployeeRow] = []

	def is_name_row(row_idx: int) -> bool:
		b = clean(ws.cell(row_idx, 2).value)
		if not b or b.startswith("Post of") or b == "Total Rs.":
			return False
		if b.startswith("Basic @") or re.fullmatch(r"[A-Z]{5}[0-9]{4}[A-Z]", b or ""):
			return False
		# Name rows have a numeric basic in col 3
		return money(ws.cell(row_idx, 3).value) > 0 and not b.endswith("Engineer") or (
			money(ws.cell(row_idx, 3).value) > 0
			and (
				b.lower().startswith("shri")
				or b.lower().startswith("smt")
				or " " in b
			)
		)

	r = 1
	while r <= ws.max_row:
		b = clean(ws.cell(r, 2).value)
		a = ws.cell(r, 1).value
		looks_like_name = isinstance(b, str) and (
			b.lower().startswith("shri") or b.lower().startswith("smt")
		)
		has_basic = money(ws.cell(r, 3).value) > 0
		if looks_like_name and has_basic:
			name = b
			sr = int(a) if isinstance(a, int) else len(employees) + 1
			# gather block until Total Rs.
			block_rows: list[list[Any]] = []
			rr = r
			while rr <= ws.max_row:
				block_rows.append([ws.cell(rr, c).value for c in range(1, 35)])
				bb = clean(ws.cell(rr, 2).value)
				if bb == "Total Rs.":
					break
				if rr > r and isinstance(ws.cell(rr, 1).value, int):
					block_rows.pop()
					rr -= 1
					break
				# next employee without sr (Amol case): name-like + basic after a Total
				if (
					rr > r
					and bb
					and (bb.lower().startswith("shri") or bb.lower().startswith("smt"))
					and money(ws.cell(rr, 3).value) > 0
				):
					block_rows.pop()
					rr -= 1
					break
				rr += 1

			first = block_rows[0]
			total = block_rows[-1] if clean(block_rows[-1][1]) == "Total Rs." else first
			designation = clean(block_rows[1][1]) if len(block_rows) > 1 else "Staff"
			pan = None
			gpf_account = None
			for rowvals in block_rows:
				cell = clean(rowvals[1])
				if cell and re.fullmatch(r"[A-Z]{5}[0-9]{4}[A-Z]", cell):
					pan = cell
				c15 = clean(rowvals[14]) if len(rowvals) > 14 else None
				if c15 and c15 not in {"G.P.F.No.", "Account No.", "Pension A/C", "Adjuesable by AG"}:
					if any(ch.isdigit() for ch in c15):
						gpf_account = c15

			if not pan:
				raise SeedError(f"Missing PAN for {name}")

			employees.append(
				EmployeeRow(
					sr=sr,
					name=name,
					designation=designation or "Staff",
					pan=pan,
					basic=money(total[2]),
					da=money(total[3]),
					cla=money(total[4]),
					hra=money(total[5]),
					wash=money(total[6]),
					other=money(total[7]),
					ta=money(total[9]),
					gross=money(total[10]),
					gpf=money(total[15]),
					nps_employer=money(total[16]),
					nps_employee=money(total[17]),
					hba=money(total[18]),
					income_tax=money(total[20]),
					gpf_account=gpf_account,
				)
			)
			r = rr + 1
			continue
		r += 1

	# Enrich from schedules
	gpf_mumbai: dict[str, str] = {}
	gpf_nagpur: dict[str, str] = {}
	for sheet, bucket in (("GPF-Mumbai", gpf_mumbai), ("GPF-Nagpur", gpf_nagpur)):
		ws = wb[sheet]
		for row in ws.iter_rows(min_row=1, values_only=True):
			if isinstance(row[0], int) and isinstance(row[2], str) and money(row[4]) >= 0:
				acct = clean(row[1])
				nm = clean(row[2])
				if nm and acct and not acct.startswith("ज्या"):
					bucket[norm_name(nm)] = acct

	gis_map: dict[str, int] = {}
	ws = wb["GIS"]
	for row in ws.iter_rows(min_row=1, values_only=True):
		if isinstance(row[1], int) and isinstance(row[2], str):
			gis_map[norm_name(clean(row[2]) or "")] = money(row[4])

	it_map: dict[str, dict[str, Any]] = {}
	ws = wb["Income Tax"]
	for row in ws.iter_rows(min_row=5, values_only=True):
		if isinstance(row[0], int) and isinstance(row[1], str):
			it_map[norm_name(clean(row[1]) or "")] = {
				"pan": clean(row[3]),
				"designation": clean(row[2]),
				"income_tax": money(row[5]),
			}

	hba_map: dict[str, dict[str, Any]] = {}
	ws = wb["HBA Ad"]
	for row in ws.iter_rows(min_row=1, values_only=True):
		if isinstance(row[0], int) and isinstance(row[1], str):
			hba_map[norm_name(clean(row[1]) or "")] = {
				"installments": clean(row[3]),
				"amount": money(row[4]),
			}

	nps_map: dict[str, dict[str, Any]] = {}
	ws = wb["Pension Sub (2)"]
	r = 1
	while r <= ws.max_row:
		if isinstance(ws.cell(r, 2).value, int) and isinstance(ws.cell(r, 4).value, str):
			nm = clean(ws.cell(r, 4).value) or ""
			nps_map[norm_name(nm)] = {
				"pension_acct": clean(ws.cell(r, 3).value),
				"nps_ee": money(ws.cell(r, 8).value),
				"nps_er": money(ws.cell(r, 9).value),
				"sevarth": clean(ws.cell(r + 1, 3).value) if r + 1 <= ws.max_row else None,
				"pran": ws.cell(r + 2, 3).value if r + 2 <= ws.max_row else None,
			}
			r += 3
			continue
		r += 1

	acc_map: dict[str, Accommodation] = {}
	ws = wb["WORLI"]
	for row in ws.iter_rows(min_row=1, values_only=True):
		if isinstance(row[0], int) and isinstance(row[1], str):
			fee = money(row[5]) + money(row[6])
			acc_map[norm_name(clean(row[1]) or "")] = Accommodation(
				location="worli",
				license_fee=fee,
				foregone_hra=money(row[4]),
				address=clean(row[3]),
			)
	ws = wb["Mumbai"]
	for row in ws.iter_rows(min_row=1, values_only=True):
		if isinstance(row[0], int) and isinstance(row[1], str):
			fee = money(row[5]) + money(row[6]) + money(row[7]) + money(row[8])
			acc_map[norm_name(clean(row[1]) or "")] = Accommodation(
				location="mumbai",
				license_fee=fee,
				foregone_hra=money(row[4]),
				address=clean(row[3]),
			)

	wb.close()

	for emp in employees:
		key = norm_name(emp.name)
		if key in gpf_nagpur:
			emp.regime = "gpf_nagpur"
			emp.gpf_account = gpf_nagpur[key]
		elif key in gpf_mumbai:
			emp.regime = "gpf_mumbai"
			emp.gpf_account = gpf_mumbai[key]
		elif key in nps_map:
			emp.regime = "nps"
			info = nps_map[key]
			emp.sevarth_id = info.get("sevarth")
			pran = info.get("pran")
			emp.pran = str(pran) if pran is not None else None
			if info.get("nps_ee"):
				emp.nps_employee = int(info["nps_ee"])
			if info.get("nps_er"):
				emp.nps_employer = int(info["nps_er"])
		elif emp.nps_employee > 0 and emp.nps_employee == emp.nps_employer:
			emp.regime = "epf"
			emp.epf_number = f"EPF/{emp.pan}"
		elif emp.nps_employee > 0 or emp.nps_employer > 0:
			emp.regime = "nps"
		elif emp.gpf > 0:
			# default GPF mumbai when schedule membership unknown
			emp.regime = "gpf_mumbai"
		else:
			emp.regime = "epf"
			emp.epf_number = f"EPF/{emp.pan}"

		if key in gis_map:
			emp.gis = gis_map[key]
		if key in it_map and it_map[key]["income_tax"]:
			emp.income_tax = int(it_map[key]["income_tax"])
			if it_map[key]["pan"]:
				emp.pan = it_map[key]["pan"]
		if key in hba_map:
			emp.hba = int(hba_map[key]["amount"])
			emp.hba_installments = hba_map[key]["installments"]
		if key in acc_map:
			emp.accommodation = acc_map[key]
			# Quarters holders: HRA is foregone, not paid
			emp.hra = 0

		if not emp.sevarth_id:
			emp.sevarth_id = f"MSIDC{emp.sr:04d}"

	employees.sort(key=lambda e: e.sr)
	return employees


def wipe_org_master_data(org_id: str, *, dsn_env: dict[str, str]) -> None:
	import subprocess

	# Validate before any SQL construction; psql :'org_id' quotes the bound value.
	org_uuid = str(UUID(org_id))
	tables = [
		"payroll_report_snapshots",
		"payroll_result_lines",
		"payroll_employee_results",
		"payroll_approvals",
		"payroll_run_inputs",
		"payroll_run_employees",
		"payroll_run_versions",
		"payroll_runs",
		"payroll_periods",
		"export_artifacts",
		"report_configurations",
		"accommodation_charge_versions",
		"accommodation_assignments",
		"advance_installment_versions",
		"advance_accounts",
		"recurring_instruction_versions",
		"recurring_instructions",
		"employee_bank_account_versions",
		"employee_pay_versions",
		"employee_posting_versions",
		"employee_profile_versions",
		"employees",
		"component_rate_versions",
		"pay_components",
		"posts",
		"offices",
		"jobs",
		"outbox_events",
		"idempotency_keys",
	]
	# Escape immutability triggers and break run↔version FK before deleting versions.
	sql = (
		"BEGIN;\n"
		"SET LOCAL accord.allow_immutable_ddl = 'on';\n"
		"UPDATE payroll_runs\n"
		"   SET current_version_id = NULL\n"
		" WHERE organization_id = :'org_id'::uuid;\n"
	)
	for table in tables:
		sql += f"DELETE FROM {table} WHERE organization_id = :'org_id'::uuid;\n"
	sql += "COMMIT;\n"
	env = {**os.environ, **dsn_env}
	proc = subprocess.run(
		[
			"psql",
			"-v",
			"ON_ERROR_STOP=1",
			"-v",
			f"org_id={org_uuid}",
			"-d",
			dsn_env.get("PGDATABASE", "accord"),
			"-c",
			sql,
		],
		env=env,
		capture_output=True,
		text=True,
	)
	if proc.returncode != 0:
		raise SeedError(f"wipe failed: {proc.stderr or proc.stdout}")
	print(f"Wiped master data for org {org_uuid}")


def seed(base_url: str, xlsx: Path, *, pg: dict[str, str]) -> None:
	employees = parse_paybill(xlsx)
	print(f"Parsed {len(employees)} employees from {xlsx.name}")
	for emp in employees:
		print(
			f"  {emp.sr:2} {emp.name[:32]:32} {emp.regime:12} "
			f"basic={emp.basic} da={emp.da} hra={emp.hra} gpf={emp.gpf} "
			f"nps={emp.nps_employee}/{emp.nps_employer} it={emp.income_tax} gis={emp.gis}"
		)

	with httpx.Client(base_url=base_url.rstrip("/"), timeout=60.0) as client:
		login = client.get("/api/auth/login", follow_redirects=False)
		if login.status_code not in {200, 302}:
			raise SeedError(f"login failed: {login.status_code}")
		me = _require(client.get("/api/auth/me"), context="me")
		if me.get("access_state") != "active" or not me.get("organization"):
			raise SeedError(
				"No active organization membership. Bootstrap with "
				"scripts/provision_organization.py and ensure this user is a member, then re-run."
			)
		org = me["organization"]
		try:
			org_id = str(UUID(org["id"]))
		except (KeyError, TypeError, ValueError) as exc:
			raise SeedError(f"organization id is not a UUID: {org.get('id')!r}") from exc
		print(f"Seeding into {org['name']} ({org['slug']})")

		wipe_org_master_data(org_id, dsn_env=pg)

		# Offices
		offices: dict[str, UUID] = {}
		for name, jurisdiction in (
			("MSIDC Mumbai HQ", "mumbai"),
			("MSIDC Nagpur GPF Circle", "nagpur"),
			("MSIDC Worli Quarters", "worli"),
		):
			body = _require(
				client.post(
					"/api/offices",
					json={"name": name, "jurisdiction": jurisdiction},
				),
				context=f"office {name}",
			)
			offices[jurisdiction] = UUID(body["id"])
			print(f"  office {name}")

		# Posts from designations
		post_ids: dict[str, UUID] = {}
		for designation in sorted({e.designation for e in employees}):
			class_name = "I"
			dl = designation.lower()
			if "clerk" in dl:
				class_name = "III"
			elif "junior" in dl or "deputy" in dl or "assistant" in dl:
				class_name = "II"
			body = _require(
				client.post(
					"/api/posts",
					json={"designation": designation, "class_name": class_name},
				),
				context=f"post {designation}",
			)
			post_ids[designation] = UUID(body["id"])
			print(f"  post {designation}")

		# Components
		component_ids: dict[str, UUID] = {}
		for order, (code, name, classification) in enumerate(COMPONENT_SPECS, start=1):
			body = _require(
				client.post(
					"/api/pay-components",
					json={
						"code": code,
						"name": name,
						"classification": classification,
						"display_order": order,
					},
				),
				context=f"component {code}",
			)
			cid = UUID(body["id"])
			component_ids[code] = cid
			if code not in {"HBA_INSTALLMENT", "ACCOMMODATION_LICENSE_FEE"}:
				_require(
					client.post(
						f"/api/pay-components/{cid}/rate-versions",
						json={
							"effective_from": EFFECTIVE_FROM,
							"calc_kind": "fixed_recurring_amount",
							"amount": "0.00",
							"rounding_rule": "ROUND_HALF_UP_RUPEE",
						},
					),
					context=f"rate {code}",
				)
			print(f"  component {code}")

		for emp in employees:
			if emp.regime == "gpf_nagpur":
				office_id = offices["nagpur"]
				regime, gpf_jurisdiction = "gpf", "nagpur"
			elif emp.regime == "gpf_mumbai":
				office_id = offices["mumbai"]
				regime, gpf_jurisdiction = "gpf", "mumbai"
			elif emp.regime == "nps":
				office_id = (
					offices["worli"]
					if emp.accommodation and emp.accommodation.location == "worli"
					else offices["mumbai"]
				)
				regime, gpf_jurisdiction = "nps", None
			elif emp.regime == "epf":
				office_id = offices["mumbai"]
				regime, gpf_jurisdiction = "epf", None
			else:
				office_id = offices["mumbai"]
				regime, gpf_jurisdiction = "gpf", "mumbai"

			profile: dict[str, Any] = {
				"name": emp.name,
				"sevarth_id": emp.sevarth_id,
				"pan": emp.pan,
				"date_of_birth": "1975-01-15",
				"date_of_joining": "2005-06-01",
				"retirement_regime": regime,
			}
			if regime == "gpf":
				profile["gpf_jurisdiction"] = gpf_jurisdiction
				profile["gpf_account_number"] = emp.gpf_account or f"GPF/{emp.pan}"
			elif regime == "nps":
				profile["pran"] = emp.pran or f"1100{emp.sr:08d}"
			else:
				profile["epf_number"] = emp.epf_number or f"EPF/{emp.pan}"

			# Fake-but-unique bank details (sheet bank tip is aggregate, not per-employee)
			body = _require(
				client.post(
					"/api/employees",
					json={
						"employee_number": f"MSIDC{emp.sr:03d}",
						"effective_from": EFFECTIVE_FROM,
						"profile": profile,
						"posting": {
							"office_id": str(office_id),
							"post_id": str(post_ids[emp.designation]),
						},
						"pay": {
							"pay_matrix_level": "L10",
							"basic_pay": money_str(emp.basic),
						},
						"bank": {
							"account_number": f"{10000000000000 + emp.sr}",
							"ifsc": "ICIC0001234",
							"bank_name": "ICICI Bank",
							"branch": "Vashi",
							"is_primary_salary": True,
						},
					},
				),
				context=f"employee {emp.name}",
			)
			employee_id = UUID(body["id"])

			recurring: list[tuple[str, int]] = [
				("DA", emp.da),
				("CLA", emp.cla),
				("HRA", emp.hra),
				("WASH_ALLOWANCE", emp.wash),
				("TRANSPORT", emp.ta),
				("OTHER_ALLOWANCE", emp.other),
				("INCOME_TAX", emp.income_tax),
				("PROFESSIONAL_TAX", emp.professional_tax),
				("GIS", emp.gis),
			]
			if regime == "gpf" and emp.gpf:
				recurring.append(("GPF_SUBSCRIPTION", emp.gpf))
			elif regime == "nps":
				if emp.nps_employee:
					recurring.append(("NPS_EMPLOYEE", emp.nps_employee))
				if emp.nps_employer:
					recurring.append(("NPS_EMPLOYER_TRANSFER", emp.nps_employer))
			elif regime == "epf":
				if emp.nps_employee:
					recurring.append(("EPF_EMPLOYEE", emp.nps_employee))
					recurring.append(("EPF_EMPLOYER", emp.nps_employee))
					recurring.append(("EPF_EMPLOYER_TRANSFER", emp.nps_employee))

			for code, amount in recurring:
				if amount <= 0:
					continue
				_require(
					client.post(
						f"/api/employees/{employee_id}/recurring-instructions",
						json={
							"component_id": str(component_ids[code]),
							"effective_from": EFFECTIVE_FROM,
							"amount": money_str(amount),
							"reason": f"Pay bill June 2026 {code}",
						},
					),
					context=f"{emp.name} {code}",
				)

			if emp.hba > 0:
				total_inst = 24
				recovered = 0
				if emp.hba_installments and "/" in emp.hba_installments:
					parts = emp.hba_installments.split("/")
					try:
						recovered = max(int(parts[0]) - 1, 0)
						total_inst = int(parts[1])
					except ValueError:
						pass
				principal = emp.hba * max(total_inst - recovered, 1)
				_require(
					client.post(
						f"/api/employees/{employee_id}/advances",
						json={
							"advance_type": "hba",
							"principal": money_str(principal),
							"sanctioned_on": EFFECTIVE_FROM,
							"reference": f"HBA-{emp.pan}",
							"installment": {
								"installment_amount": money_str(emp.hba),
								"installments_total": total_inst,
								"installments_recovered_opening": recovered,
								"effective_from": EFFECTIVE_FROM,
							},
						},
					),
					context=f"{emp.name} HBA",
				)

			if emp.accommodation and emp.accommodation.license_fee > 0:
				charge: dict[str, Any] = {
					"license_fee": money_str(emp.accommodation.license_fee),
					"effective_from": EFFECTIVE_FROM,
				}
				if emp.accommodation.foregone_hra > 0:
					charge["informational_hra_foregone"] = money_str(
						emp.accommodation.foregone_hra
					)
				_require(
					client.post(
						f"/api/employees/{employee_id}/accommodation",
						json={
							"quarters_location": emp.accommodation.location,
							"quarters_identifier": emp.accommodation.address
							or f"Q-{emp.sr}",
							"charge": charge,
						},
					),
					context=f"{emp.name} accommodation",
				)

			print(f"  employee {emp.sr:02d} {emp.name}")

		listed = _require(
			client.get("/api/employees", params={"page_size": 100}),
			context="verify",
		)
		print(f"Done. {listed.get('total')} real employees loaded into {org['name']}.")


def main() -> int:
	parser = argparse.ArgumentParser(description=__doc__)
	parser.add_argument(
		"--xlsx",
		type=Path,
		default=Path("/Users/darshan/Downloads/Pay bill - June 2026 Regular Staff.xlsx"),
	)
	parser.add_argument("--base-url", default="http://127.0.0.1:8000")
	parser.add_argument("--pghost", default=os.environ.get("PGHOST", "127.0.0.1"))
	parser.add_argument("--pgport", default=os.environ.get("PGPORT", "5433"))
	parser.add_argument("--pguser", default=os.environ.get("PGUSER", "accord"))
	parser.add_argument("--pgpassword", default=os.environ.get("PGPASSWORD", "accord"))
	parser.add_argument("--pgdatabase", default=os.environ.get("PGDATABASE", "accord"))
	args = parser.parse_args()
	if not args.xlsx.is_file():
		print(f"error: spreadsheet not found: {args.xlsx}", file=sys.stderr)
		return 1
	pg = {
		"PGHOST": args.pghost,
		"PGPORT": str(args.pgport),
		"PGUSER": args.pguser,
		"PGPASSWORD": args.pgpassword,
		"PGDATABASE": args.pgdatabase,
	}
	try:
		seed(args.base_url, args.xlsx, pg=pg)
	except SeedError as exc:
		print(f"error: {exc}", file=sys.stderr)
		return 1
	return 0


if __name__ == "__main__":
	raise SystemExit(main())
