#!/usr/bin/env python3
"""Build the PII-free structural template for canonical payroll schedules.

The accepted June workbook is an input to this developer tool only.  The
output is a new workbook containing no source cell values or relationships.
Only layout metadata and cell style objects are copied.
"""

from __future__ import annotations

import argparse
from copy import copy
from datetime import datetime
from io import BytesIO
from pathlib import Path
import re
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import range_boundaries


CANONICAL_SHEETS = (
    "office tip",
    "Bank Tip",
    "PaySlip",
    " Face ",
    "Pay Bill",
    "Income Tax",
    "GPF-Nagpur",
    "P.T.",
    "GPF-Mumbai",
    "GPF-IV",
    "GIS",
    "HBA Ad",
    "Motor car Ad",
    "Motor cycale Ad (2)",
    "Pension Sub (2)",
    "Festival",
    "WORLI",
    "Mumbai",
)

_FIXED_TIME = datetime(2000, 1, 1)
_ZIP_TIME = (2000, 1, 1, 0, 0, 0)
_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _copy_cell_style(source, target) -> None:
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)
    target.number_format = source.number_format


def _copy_layout(source, target) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(source.calculate_dimension())
    for row in source.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for source_cell in row:
            _copy_cell_style(source_cell, target[source_cell.coordinate])

    for merged_range in source.merged_cells.ranges:
        target.merge_cells(str(merged_range))

    for index, source_dimension in source.row_dimensions.items():
        target_dimension = target.row_dimensions[index]
        target_dimension.height = source_dimension.height
        target_dimension.hidden = source_dimension.hidden
        target_dimension.outlineLevel = source_dimension.outlineLevel
        target_dimension.collapsed = source_dimension.collapsed

    for index, source_dimension in source.column_dimensions.items():
        target_dimension = target.column_dimensions[index]
        target_dimension.min = source_dimension.min
        target_dimension.max = source_dimension.max
        target_dimension.width = source_dimension.width
        target_dimension.hidden = source_dimension.hidden
        target_dimension.outlineLevel = source_dimension.outlineLevel
        target_dimension.collapsed = source_dimension.collapsed
        target_dimension.bestFit = source_dimension.bestFit

    target.sheet_state = source.sheet_state
    target.sheet_view.showGridLines = source.sheet_view.showGridLines
    target.sheet_view.zoomScale = source.sheet_view.zoomScale
    target.sheet_view.zoomScaleNormal = source.sheet_view.zoomScaleNormal
    target.freeze_panes = source.freeze_panes
    target.sheet_format.defaultColWidth = source.sheet_format.defaultColWidth
    target.sheet_format.defaultRowHeight = source.sheet_format.defaultRowHeight
    target.page_setup = copy(source.page_setup)
    target.page_margins = copy(source.page_margins)
    target.print_options = copy(source.print_options)
    target.sheet_properties.pageSetUpPr = copy(source.sheet_properties.pageSetUpPr)
    if source.print_area:
        target.print_area = source.print_area.rsplit("!", 1)[-1]
    if source.print_title_rows:
        target.print_title_rows = source.print_title_rows
    if source.print_title_cols:
        target.print_title_cols = source.print_title_cols
    target.row_breaks = copy(source.row_breaks)
    target.col_breaks = copy(source.col_breaks)


def _deterministic_xlsx(content: bytes) -> bytes:
    source = ZipFile(BytesIO(content))
    output = BytesIO()
    with source, ZipFile(output, "w", compression=ZIP_DEFLATED, compresslevel=9) as target:
        for name in sorted(source.namelist()):
            payload = source.read(name)
            if name == "docProps/core.xml":
                payload = re.sub(
                    rb"<dcterms:modified[^>]*>.*?</dcterms:modified>",
                    (
                        b'<dcterms:modified xsi:type="dcterms:W3CDTF">'
                        b"2000-01-01T00:00:00Z</dcterms:modified>"
                    ),
                    payload,
                )
            info = ZipInfo(name, _ZIP_TIME)
            info.compress_type = ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            target.writestr(info, payload)
    return output.getvalue()


def validate_template(content: bytes) -> None:
    """Fail closed unless ``content`` is a value-free structural workbook."""
    with ZipFile(BytesIO(content)) as archive:
        names = set(archive.namelist())
        forbidden_fragments = (
            "sharedstrings",
            "externallink",
            "connection",
            "comment",
            "person",
            "threadedcomment",
            "customxml",
            "docprops/custom",
            "/media/",
            "/drawings/",
            "vml",
        )
        unsafe_names = sorted(
            name
            for name in names
            if any(fragment in name.casefold() for fragment in forbidden_fragments)
        )
        if unsafe_names:
            raise ValueError(f"Unsafe OOXML parts in structural template: {unsafe_names!r}")

        workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        defined_names = workbook_root.find(f"{{{_MAIN_NS}}}definedNames")
        if defined_names is not None:
            allowed = {"_xlnm.Print_Area", "_xlnm.Print_Titles"}
            unexpected = [
                item.attrib.get("name", "")
                for item in defined_names
                if item.attrib.get("name") not in allowed
            ]
            if unexpected:
                raise ValueError(f"Unexpected defined names: {unexpected!r}")

        for name in sorted(names):
            if not name.startswith("xl/worksheets/") or not name.endswith(".xml"):
                continue
            root = ElementTree.fromstring(archive.read(name))
            for cell in root.iter(f"{{{_MAIN_NS}}}c"):
                if len(cell):
                    raise ValueError(f"Worksheet payload found in {name}:{cell.attrib.get('r')}")
            for forbidden_tag in (
                "f",
                "v",
                "is",
                "hyperlinks",
                "drawing",
                "legacyDrawing",
                "oleObjects",
                "controls",
            ):
                if root.find(f".//{{{_MAIN_NS}}}{forbidden_tag}") is not None:
                    raise ValueError(f"Unsafe {forbidden_tag} element found in {name}")
            header_footer = root.find(f"{{{_MAIN_NS}}}headerFooter")
            if header_footer is not None and any(
                (item.text or "").strip() for item in header_footer
            ):
                raise ValueError(f"Header/footer payload found in {name}")


def build_template(source_path: Path) -> bytes:
    source = load_workbook(source_path, data_only=False, read_only=False)
    missing = [name for name in CANONICAL_SHEETS if name not in source.sheetnames]
    if missing:
        raise ValueError(f"Source workbook is missing canonical sheets: {missing!r}")

    target = Workbook()
    target.remove(target.active)
    target.properties.creator = "Accord"
    target.properties.lastModifiedBy = "Accord"
    target.properties.title = "Accord canonical schedule structure"
    target.properties.subject = "PII-free layout template"
    target.properties.description = "Styles and worksheet layout only; no source values."
    target.properties.created = _FIXED_TIME
    target.properties.modified = _FIXED_TIME

    for name in CANONICAL_SHEETS:
        target_sheet = target.create_sheet(name)
        _copy_layout(source[name], target_sheet)

    buffer = BytesIO()
    target.save(buffer)
    content = _deterministic_xlsx(buffer.getvalue())
    validate_template(content)
    return content


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_bytes(build_template(args.source))


if __name__ == "__main__":
    main()
