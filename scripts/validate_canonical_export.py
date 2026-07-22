#!/usr/bin/env python3
"""Validate an Accord v3 export against the normalized canonical contract.

The validator is intentionally separate from the report formatters.  Structural
checks run against the original OOXML package.  Formula results are checked on
a temporary LibreOffice-recalculated copy so the supplied workbook is never
modified.
"""

from __future__ import annotations

import argparse
from collections.abc import Iterable, Mapping, Sequence
from decimal import Decimal, InvalidOperation
import importlib.util
import json
from pathlib import Path
import re
import shutil
import subprocess
import tempfile
from typing import Any, NamedTuple
import zipfile
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


_EXTRACTOR_PATH = Path(__file__).with_name("extract_canonical_export_contract.py")
_EXTRACTOR_SPEC = importlib.util.spec_from_file_location(
    "accord_extract_canonical_export_contract", _EXTRACTOR_PATH
)
if _EXTRACTOR_SPEC is None or _EXTRACTOR_SPEC.loader is None:  # pragma: no cover
    raise RuntimeError(f"Could not load canonical contract extractor: {_EXTRACTOR_PATH}")
_EXTRACTOR = importlib.util.module_from_spec(_EXTRACTOR_SPEC)
_EXTRACTOR_SPEC.loader.exec_module(_EXTRACTOR)
extract_contract = _EXTRACTOR.extract_contract


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CONTRACT_PATH = (
    ROOT / "fixtures" / "sanitized" / "june-2026" / "canonical_export_contract.json"
)
_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NUMERIC_TOLERANCE = Decimal("0.000001")
_ERROR_VALUES = frozenset({"#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!"})


class ValidationIssue(NamedTuple):
    """One PII-safe acceptance failure."""

    code: str
    location: str
    message: str
    expected: object | None = None
    actual: object | None = None

    def render(self) -> str:
        detail = f"[{self.code}] {self.location}: {self.message}"
        if self.expected is not None or self.actual is not None:
            detail += f" (expected={self.expected!r}, actual={self.actual!r})"
        return detail


def _issue(
    issues: list[ValidationIssue],
    code: str,
    location: str,
    message: str,
    *,
    expected: object | None = None,
    actual: object | None = None,
) -> None:
    issues.append(ValidationIssue(code, location, message, expected, actual))


def load_contract(path: Path = DEFAULT_CONTRACT_PATH) -> dict[str, Any]:
    contract = json.loads(path.read_text(encoding="utf-8"))
    if contract.get("contract_version") != 1:
        raise ValueError(
            f"Unsupported canonical contract version: {contract.get('contract_version')}"
        )
    return contract


def _decimal(value: object) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _equal_scalar(expected: object, actual: object) -> bool:
    expected_decimal = _decimal(expected)
    actual_decimal = _decimal(actual)
    if expected_decimal is not None and actual_decimal is not None:
        return abs(expected_decimal - actual_decimal) <= _NUMERIC_TOLERANCE
    return expected == actual


def _normalize_range(value: object) -> str | None:
    """Normalize harmless Excel quoting/absolute-reference differences."""
    if value is None:
        return None
    normalized = str(value).strip().replace("$", "")
    if "!" in normalized:
        sheet, normalized = normalized.rsplit("!", 1)
        sheet = sheet.strip("'").replace("''", "'")
        # Only the local range is structural; sheet identity is checked separately.
        del sheet
    return normalized.upper()


def _expanded_columns(items: Sequence[Mapping[str, object]]) -> dict[int, dict[str, object]]:
    expanded: dict[int, dict[str, object]] = {}
    for item in items:
        start = int(item["min"])
        end = int(item["max"])
        attributes = {key: value for key, value in item.items() if key not in {"min", "max"}}
        for column in range(start, end + 1):
            expanded[column] = attributes
    return expanded


def _expanded_rows(items: Sequence[Mapping[str, object]]) -> dict[int, dict[str, object]]:
    return {
        int(item["row"]): {key: value for key, value in item.items() if key != "row"}
        for item in items
    }


def _compare_attribute_maps(
    issues: list[ValidationIssue],
    *,
    code: str,
    location: str,
    expected: Mapping[object, Mapping[str, object]],
    actual: Mapping[object, Mapping[str, object]],
) -> None:
    for key in sorted(set(expected) | set(actual), key=str):
        expected_attributes = expected.get(key)
        actual_attributes = actual.get(key)
        if expected_attributes is None or actual_attributes is None:
            _issue(
                issues,
                code,
                f"{location}[{key}]",
                "dimension entry differs",
                expected=expected_attributes,
                actual=actual_attributes,
            )
            continue
        for attribute in sorted(set(expected_attributes) | set(actual_attributes)):
            expected_value = expected_attributes.get(attribute)
            actual_value = actual_attributes.get(attribute)
            if not _equal_scalar(expected_value, actual_value):
                _issue(
                    issues,
                    code,
                    f"{location}[{key}].{attribute}",
                    "dimension attribute differs",
                    expected=expected_value,
                    actual=actual_value,
                )


def _compare_flat_map(
    issues: list[ValidationIssue],
    *,
    code: str,
    location: str,
    expected: Mapping[str, object],
    actual: Mapping[str, object],
) -> None:
    for key in sorted(set(expected) | set(actual)):
        expected_value = expected.get(key)
        actual_value = actual.get(key)
        if not _equal_scalar(expected_value, actual_value):
            _issue(
                issues,
                code,
                f"{location}.{key}",
                "setting differs",
                expected=expected_value,
                actual=actual_value,
            )


def _normalized_payslip_merges(starts: Sequence[int]) -> set[str]:
    merges: set[str] = set()
    for start in starts:
        merges.add(f"A{start}:A{start + 16}")
        for min_col, max_col in (
            (2, 4),
            (5, 7),
            (8, 10),
            (11, 12),
            (14, 15),
            (16, 19),
            (20, 24),
        ):
            merges.add(f"{get_column_letter(min_col)}{start}:{get_column_letter(max_col)}{start}")
        for min_col, max_col in ((2, 12), (13, 14), (15, 18), (19, 24)):
            merges.add(
                f"{get_column_letter(min_col)}{start + 1}:{get_column_letter(max_col)}{start + 1}"
            )
        for min_col, max_col in ((2, 7), (8, 16), (17, 24)):
            merges.add(
                f"{get_column_letter(min_col)}{start + 2}:{get_column_letter(max_col)}{start + 2}"
            )
        for row in range(start + 3, start + 14):
            for min_col, max_col in (
                (2, 5),
                (6, 7),
                (8, 11),
                (12, 13),
                (14, 16),
                (17, 21),
                (22, 24),
            ):
                merges.add(f"{get_column_letter(min_col)}{row}:{get_column_letter(max_col)}{row}")
        for min_col, max_col in ((2, 5), (6, 7), (8, 15), (16, 23)):
            merges.add(
                f"{get_column_letter(min_col)}{start + 14}:{get_column_letter(max_col)}{start + 15}"
            )
    return merges


def _validate_normalized_payslip_structure(
    issues: list[ValidationIssue],
    actual: Mapping[str, Any],
    semantics: Mapping[str, Any],
) -> None:
    structure = semantics["normalized_structure"]
    starts = [int(value) for value in semantics["block_starts"]]
    location = "sheet['PaySlip']"
    for field in ("used_range", "print_area"):
        if _normalize_range(actual.get(field)) != _normalize_range(structure[field]):
            _issue(
                issues,
                f"sheet.{field}",
                location,
                f"normalized PaySlip {field.replace('_', ' ')} differs",
                expected=structure[field],
                actual=actual.get(field),
            )

    expected_columns = {
        index: {"width": width} for index, width in enumerate(structure["column_widths"], start=1)
    }
    _compare_attribute_maps(
        issues,
        code="sheet.column_dimensions",
        location=f"{location}.columns",
        expected=expected_columns,
        actual=_expanded_columns(actual["column_dimensions"]),
    )

    expected_rows: dict[int, dict[str, object]] = {}
    for start in starts:
        expected_rows[start] = {"height": structure["start_row_height"]}
        for row in range(start + 1, start + 14):
            expected_rows[row] = {"height": structure["default_row_height"]}
        expected_rows[start + 13] = {"height": structure["total_row_height"]}
        expected_rows[start + 14] = {"height": structure["net_row_height"]}
    _compare_attribute_maps(
        issues,
        code="sheet.row_dimensions",
        location=f"{location}.rows",
        expected=expected_rows,
        actual=_expanded_rows(actual["row_dimensions"]),
    )

    expected_merges = _normalized_payslip_merges(starts)
    actual_merges = set(actual["merged_cells"])
    if expected_merges != actual_merges:
        _issue(
            issues,
            "sheet.merged_cells",
            location,
            "normalized PaySlip repeated-block merges differ",
            expected={"missing": sorted(expected_merges - actual_merges)},
            actual={"unexpected": sorted(actual_merges - expected_merges)},
        )

    expected_breaks = [starts[index - 1] + 17 for index in range(2, len(starts), 2)]
    if actual["manual_row_breaks"] != expected_breaks:
        _issue(
            issues,
            "sheet.manual_row_breaks",
            location,
            "normalized PaySlip page breaks differ",
            expected=expected_breaks,
            actual=actual["manual_row_breaks"],
        )
    for key, expected in structure["page_setup"].items():
        actual_value = actual["page_setup"].get(key)
        if not _equal_scalar(expected, actual_value):
            _issue(
                issues,
                "sheet.page_setup",
                f"{location}.page_setup.{key}",
                "normalized PaySlip page setting differs",
                expected=expected,
                actual=actual_value,
            )


def validate_structure(workbook_path: Path, contract: Mapping[str, Any]) -> list[ValidationIssue]:
    """Compare topology and print layout without reading employee cell values."""
    issues: list[ValidationIssue] = []
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    expected_sheets = list(contract["sheets"])
    expected_names = [sheet["name"] for sheet in expected_sheets]
    if workbook.sheetnames != expected_names:
        _issue(
            issues,
            "workbook.sheet_order",
            "workbook",
            "sheet names or order differ",
            expected=expected_names,
            actual=workbook.sheetnames,
        )
        return issues

    try:
        actual_contract = extract_contract(workbook_path)
    except (KeyError, OSError, ValueError, zipfile.BadZipFile) as error:
        _issue(issues, "workbook.ooxml", "workbook", f"could not extract structure: {error}")
        return issues

    for expected, actual in zip(expected_sheets, actual_contract["sheets"], strict=True):
        name = str(expected["name"])
        location = f"sheet[{name!r}]"
        if expected["state"] != actual["state"]:
            _issue(
                issues,
                "sheet.state",
                location,
                "visibility differs",
                expected=expected["state"],
                actual=actual["state"],
            )
        if name == "PaySlip":
            payslip_semantics = contract["canonical_semantics"]["front_sheets"]["PaySlip"]
            if "normalized_structure" in payslip_semantics:
                _validate_normalized_payslip_structure(issues, actual, payslip_semantics)
                continue
        for field in ("used_range", "print_area", "print_titles"):
            if _normalize_range(expected.get(field)) != _normalize_range(actual.get(field)):
                _issue(
                    issues,
                    f"sheet.{field}",
                    location,
                    f"{field.replace('_', ' ')} differs",
                    expected=expected.get(field),
                    actual=actual.get(field),
                )

        expected_merges = set(expected["merged_cells"])
        actual_merges = set(actual["merged_cells"])
        if expected_merges != actual_merges:
            _issue(
                issues,
                "sheet.merged_cells",
                location,
                "merged ranges differ",
                expected={
                    "missing": sorted(expected_merges - actual_merges),
                    "unexpected": [],
                },
                actual={
                    "missing": [],
                    "unexpected": sorted(actual_merges - expected_merges),
                },
            )

        _compare_attribute_maps(
            issues,
            code="sheet.column_dimensions",
            location=f"{location}.columns",
            expected=_expanded_columns(expected["column_dimensions"]),
            actual=_expanded_columns(actual["column_dimensions"]),
        )
        _compare_attribute_maps(
            issues,
            code="sheet.row_dimensions",
            location=f"{location}.rows",
            expected=_expanded_rows(expected["row_dimensions"]),
            actual=_expanded_rows(actual["row_dimensions"]),
        )

        for field in ("hidden_rows", "hidden_columns"):
            expected_value = sorted(
                tuple(item) if isinstance(item, list) else item for item in expected[field]
            )
            actual_value = sorted(
                tuple(item) if isinstance(item, list) else item for item in actual[field]
            )
            if expected_value != actual_value:
                _issue(
                    issues,
                    f"sheet.{field}",
                    location,
                    f"{field.replace('_', ' ')} differ",
                    expected=expected_value,
                    actual=actual_value,
                )

        for field in ("page_setup", "page_margins", "print_options"):
            _compare_flat_map(
                issues,
                code=f"sheet.{field}",
                location=f"{location}.{field}",
                expected=expected[field],
                actual=actual[field],
            )
        for field in ("manual_row_breaks", "manual_column_breaks"):
            if sorted(expected[field]) != sorted(actual[field]):
                _issue(
                    issues,
                    f"sheet.{field}",
                    location,
                    f"{field.replace('_', ' ')} differ",
                    expected=sorted(expected[field]),
                    actual=sorted(actual[field]),
                )
    return issues


def _merged_anchor_value(sheet: Worksheet, row: int, column: int) -> object:
    for merged in sheet.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= column <= merged.max_col:
            return sheet.cell(merged.min_row, merged.min_col).value
    return sheet.cell(row, column).value


def _normalized_text(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").casefold()).strip()


_HEADER_TERMS: tuple[tuple[str, ...], ...] = (
    ("sr", "no"),
    ("employee", "name"),
    ("basic", "pay"),
    ("dearness", "allowance"),
    ("city", "compensatory"),
    ("house", "rent", "allowance"),
    ("wash", "child", "other"),
    ("reimbursement", "salary", "difference"),
    ("additional", "allowance"),
    ("ta", "pta", "honorarium"),
    ("gross", "salary"),
    ("employer", "share"),
    ("festival", "recovery"),
    ("gross", "salary", "recovery"),
    ("account", "number"),
    ("subscription", "refund", "arrears"),
    ("pension", "employer", "share"),
    ("pension", "employee", "share"),
    ("advance",),
    ("flood", "advance"),
    ("income", "tax"),
    ("pli", "gis"),
    ("house", "rent", "service", "charges"),
    ("professional", "tax"),
    ("co", "operative", "recovery"),
    ("total", "deductions"),
    ("net", "amount", "payable"),
    ("remarks",),
)


def _formula(cell: Cell) -> str:
    value = cell.value
    return "" if not isinstance(value, str) else re.sub(r"\s+|\$", "", value.upper())


def _validate_total_formula_roles(
    issues: list[ValidationIssue], sheet: Worksheet, row: int, *, role: str
) -> None:
    expected = {
        "K": f"=SUM(C{row}:J{row})",
        "N": f"=K{row}+L{row}-M{row}",
        "Z": f"=SUM(P{row}:Y{row})",
        "AA": f"=N{row}-Z{row}",
    }
    for column, formula in expected.items():
        actual = _formula(sheet[f"{column}{row}"])
        if actual != formula:
            _issue(
                issues,
                "pay_bill.formula_role",
                f"Pay Bill!{column}{row}",
                f"{role} formula role differs",
                expected=formula,
                actual=actual,
            )


def _sum_references(formula: str) -> set[str] | None:
    """Return the cells in a simple SUM formula, expanding same-column ranges."""
    match = re.fullmatch(r"=SUM\((.*)\)", formula)
    if match is None:
        return None
    references: set[str] = set()
    for part in match.group(1).split(","):
        range_match = re.fullmatch(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", part)
        if range_match is not None:
            start_column, start_row, end_column, end_row = range_match.groups()
            if start_column != end_column:
                return None
            references.update(
                f"{start_column}{row}" for row in range(int(start_row), int(end_row) + 1)
            )
            continue
        if re.fullmatch(r"[A-Z]+\d+", part) is None:
            return None
        references.add(part)
    return references


def _validate_sum_references(
    issues: list[ValidationIssue],
    sheet: Worksheet,
    *,
    row: int,
    column: int,
    source_rows: Sequence[int],
    role: str,
) -> None:
    letter = get_column_letter(column)
    expected = {f"{letter}{source_row}" for source_row in source_rows}
    actual_formula = _formula(sheet.cell(row, column))
    actual = _sum_references(actual_formula)
    if actual != expected:
        _issue(
            issues,
            "pay_bill.total_references",
            f"Pay Bill!{letter}{row}",
            f"{role} does not sum the canonical source cells",
            expected=sorted(expected),
            actual=actual_formula,
        )


def _serial_cells(sheet: Worksheet) -> list[tuple[int, int]]:
    serials: list[tuple[int, int]] = []
    for row in range(9, sheet.max_row + 1):
        value = sheet.cell(row, 1).value
        if isinstance(value, bool):
            continue
        number = _decimal(value)
        if number is not None and number == number.to_integral_value() and number > 0:
            serials.append((row, int(number)))
    return serials


def validate_pay_bill_semantics(
    formula_sheet: Worksheet,
    contract: Mapping[str, Any],
    *,
    value_sheet: Worksheet | None = None,
    require_june_totals: bool = True,
) -> list[ValidationIssue]:
    """Validate grouped headers, employee blocks, formula roles, and June totals."""
    issues: list[ValidationIssue] = []
    pay_bill_contract = contract["pay_bill"]
    merged = {str(item) for item in formula_sheet.merged_cells.ranges}
    for role, cell_range in pay_bill_contract["header_groups"].items():
        if cell_range not in merged:
            _issue(
                issues,
                "pay_bill.header_group",
                "Pay Bill",
                f"missing {role.replace('_', ' ')} merge",
                expected=cell_range,
                actual=None,
            )
    group_text = {
        "O2": ("deductions",),
        "O3": ("adjustable", "accountant", "general"),
        "U3": ("adjustable", "treasury"),
    }
    for coordinate, terms in group_text.items():
        text = _normalized_text(formula_sheet[coordinate].value)
        if any(term not in text for term in terms):
            _issue(
                issues,
                "pay_bill.header_group_text",
                f"Pay Bill!{coordinate}",
                "group header is not canonical",
                expected=" ".join(terms),
                actual=text,
            )

    ordinals = [formula_sheet.cell(8, column).value for column in range(1, 29)]
    if ordinals != list(range(1, 29)):
        _issue(
            issues,
            "pay_bill.ordinal_row",
            "Pay Bill!A8:AB8",
            "canonical column ordinals differ",
            expected=list(range(1, 29)),
            actual=ordinals,
        )

    for column, required_terms in enumerate(_HEADER_TERMS, start=1):
        combined = " ".join(
            dict.fromkeys(
                _normalized_text(_merged_anchor_value(formula_sheet, row, column))
                for row in range(2, 6)
                if _merged_anchor_value(formula_sheet, row, column) is not None
            )
        )
        if any(term not in combined for term in required_terms):
            _issue(
                issues,
                "pay_bill.column_header",
                f"Pay Bill!{get_column_letter(column)}2:{get_column_letter(column)}5",
                "column header is missing canonical terms",
                expected=" ".join(required_terms),
                actual=combined,
            )

    serial_cells = _serial_cells(formula_sheet)
    serial_numbers = [serial for _row, serial in serial_cells]
    if serial_numbers != list(range(1, len(serial_numbers) + 1)):
        _issue(
            issues,
            "pay_bill.employee_serials",
            "Pay Bill!A9:A208",
            "employee serial numbers are not contiguous",
            expected=f"1..{len(serial_numbers)}",
            actual=serial_numbers,
        )

    employee_total_rows: list[int] = []
    for serial_row, serial in serial_cells:
        merged_range = next(
            (
                item
                for item in formula_sheet.merged_cells.ranges
                if item.min_col == item.max_col == 1 and item.min_row == serial_row
            ),
            None,
        )
        if merged_range is None:
            _issue(
                issues,
                "pay_bill.employee_block",
                f"Pay Bill!A{serial_row}",
                f"employee {serial} does not have a merged detail block",
            )
            continue
        search_end = min(merged_range.max_row + 2, formula_sheet.max_row)
        total_row = next(
            (
                row
                for row in range(merged_range.min_row, search_end + 1)
                if _normalized_text(formula_sheet.cell(row, 2).value) == "total rs"
            ),
            None,
        )
        if total_row is None:
            _issue(
                issues,
                "pay_bill.employee_total",
                f"Pay Bill!A{serial_row}:AB{search_end}",
                f"employee {serial} has no total row",
            )
            continue
        employee_total_rows.append(total_row)
        _validate_total_formula_roles(issues, formula_sheet, total_row, role="employee total")
        detail_end = total_row - 1
        for column in (*range(3, 11), 12, 13, *range(16, 26)):
            letter = get_column_letter(column)
            expected_formula = f"=SUM({letter}{serial_row}:{letter}{detail_end})"
            actual_formula = _formula(formula_sheet.cell(total_row, column))
            if actual_formula != expected_formula:
                _issue(
                    issues,
                    "pay_bill.detail_formula",
                    f"Pay Bill!{letter}{total_row}",
                    "employee detail sum differs",
                    expected=expected_formula,
                    actual=actual_formula,
                )

    page_starts = [
        row
        for row in range(9, formula_sheet.max_row + 1)
        if _normalized_text(formula_sheet.cell(row, 2).value).startswith("total of page no")
    ]
    if not page_starts:
        _issue(issues, "pay_bill.page_totals", "Pay Bill", "no page total blocks found")
    for start in page_starts:
        _validate_total_formula_roles(issues, formula_sheet, start + 5, role="page total")
        for row in range(start, start + 5):
            for column in (*range(3, 15), *range(16, 28)):
                if not _formula(formula_sheet.cell(row, column)).startswith("=SUM("):
                    _issue(
                        issues,
                        "pay_bill.page_detail_total",
                        f"Pay Bill!{get_column_letter(column)}{row}",
                        "page detail total is not a SUM formula",
                    )

    grand_starts = [
        row
        for row in range(9, formula_sheet.max_row + 1)
        if _normalized_text(formula_sheet.cell(row, 2).value) == "total of all pages"
    ]
    if len(grand_starts) != 1:
        _issue(
            issues,
            "pay_bill.grand_total",
            "Pay Bill",
            "expected one grand-total block",
            expected=1,
            actual=len(grand_starts),
        )
    else:
        grand_total_row = grand_starts[0] + 5
        _validate_total_formula_roles(issues, formula_sheet, grand_total_row, role="grand total")
        for row in range(grand_starts[0], grand_total_row):
            for column in (*range(3, 15), *range(16, 28)):
                if not _formula(formula_sheet.cell(row, column)).startswith("=SUM("):
                    _issue(
                        issues,
                        "pay_bill.grand_detail_total",
                        f"Pay Bill!{get_column_letter(column)}{row}",
                        "grand detail total is not a SUM formula",
                    )

        if len(serial_cells) == 28:
            pay_bill_sheet_contract = next(
                item for item in contract["sheets"] if item["name"] == "Pay Bill"
            )
            expected_grand_start = 203
            print_range = _normalize_range(pay_bill_sheet_contract.get("print_area"))
            print_match = re.fullmatch(r"[A-Z]+\d+:[A-Z]+(\d+)", print_range or "")
            if print_match is not None:
                expected_grand_start = int(print_match.group(1)) - 5
            expected_page_starts = [
                *(int(row) - 5 for row in pay_bill_sheet_contract["manual_row_breaks"]),
                expected_grand_start - 6,
            ]
            if page_starts != expected_page_starts:
                _issue(
                    issues,
                    "pay_bill.page_layout",
                    "Pay Bill",
                    "28-person page total blocks differ",
                    expected=expected_page_starts,
                    actual=page_starts,
                )
            if grand_starts[0] != expected_grand_start:
                _issue(
                    issues,
                    "pay_bill.page_layout",
                    "Pay Bill",
                    "28-person grand-total block differs",
                    expected=expected_grand_start,
                    actual=grand_starts[0],
                )

            lower_bound = 9
            for page_start in page_starts:
                page_serial_rows = [
                    row for row, _serial in serial_cells if lower_bound <= row < page_start
                ]
                page_employee_total_rows = [
                    row for row in employee_total_rows if lower_bound <= row < page_start
                ]
                for offset in range(5):
                    for column in (*range(3, 15), *range(16, 28)):
                        _validate_sum_references(
                            issues,
                            formula_sheet,
                            row=page_start + offset,
                            column=column,
                            source_rows=[row + offset for row in page_serial_rows],
                            role="page detail total",
                        )
                for column in (*range(3, 11), 12, 13, *range(16, 26)):
                    _validate_sum_references(
                        issues,
                        formula_sheet,
                        row=page_start + 5,
                        column=column,
                        source_rows=page_employee_total_rows,
                        role="page total",
                    )
                lower_bound = page_start + 6

            for offset in range(5):
                for column in (*range(3, 15), *range(16, 28)):
                    _validate_sum_references(
                        issues,
                        formula_sheet,
                        row=grand_starts[0] + offset,
                        column=column,
                        source_rows=[row + offset for row in page_starts],
                        role="grand detail total",
                    )
            for column in (*range(3, 11), 12, 13, *range(16, 26)):
                _validate_sum_references(
                    issues,
                    formula_sheet,
                    row=grand_total_row,
                    column=column,
                    source_rows=[row + 5 for row in page_starts],
                    role="grand total",
                )

            if value_sheet is None and require_june_totals:
                _issue(
                    issues,
                    "pay_bill.june_totals",
                    "Pay Bill",
                    "28-person export requires recalculated values for June reconciliation",
                )
            elif value_sheet is not None:
                totals = pay_bill_contract["normalized_june_2026_totals"]
                actual_totals = {
                    "employee_count": Decimal(len(serial_cells)),
                    "salary_earnings": sum(
                        (_decimal(value_sheet.cell(grand_total_row, column).value) or Decimal(0))
                        for column in range(3, 11)
                    ),
                    "employer_share": _decimal(value_sheet[f"L{grand_total_row}"].value),
                    "gross_bill": _decimal(value_sheet[f"N{grand_total_row}"].value),
                    "total_deductions": _decimal(value_sheet[f"Z{grand_total_row}"].value),
                    "net_payable": _decimal(value_sheet[f"AA{grand_total_row}"].value),
                }
                for key, expected_value in totals.items():
                    actual_value = actual_totals[key]
                    if actual_value is None or not _equal_scalar(expected_value, actual_value):
                        _issue(
                            issues,
                            "pay_bill.june_totals",
                            f"Pay Bill!{key}",
                            "normalized June total differs",
                            expected=expected_value,
                            actual=actual_value,
                        )
    return issues


def _is_present(value: object) -> bool:
    return value is not None and (not isinstance(value, str) or bool(value.strip()))


def _validate_body(
    issues: list[ValidationIssue], sheet: Worksheet, spec: Mapping[str, Any]
) -> list[int]:
    start, end = (int(value) for value in spec["rows"])
    serial_column = str(spec["serial_column"])
    active: list[tuple[int, int]] = []
    expected_count = int(spec["expected_count"])
    for row in range(start, end + 1):
        serial = _decimal(sheet[f"{serial_column}{row}"].value)
        has_payload = any(
            _is_present(sheet[f"{column}{row}"].value)
            for column in spec.get("required_columns", [])
        ) or any(
            (_decimal(sheet[f"{column}{row}"].value) or Decimal(0)) != 0
            for column in spec.get("amount_columns", [])
        )
        if (
            serial is not None
            and serial > 0
            and serial == serial.to_integral_value()
            and (expected_count > 0 or has_payload)
        ):
            active.append((row, int(serial)))

    serials = [serial for _row, serial in active]
    if len(active) != expected_count:
        _issue(
            issues,
            "sheet.body_empty" if not active and expected_count else "sheet.body_count",
            sheet.title,
            "canonical body row count differs",
            expected=expected_count,
            actual=len(active),
        )
    if serials != list(range(1, len(serials) + 1)):
        _issue(
            issues,
            "sheet.body_serials",
            sheet.title,
            "body serials are not contiguous",
            expected=f"1..{len(serials)}",
            actual=serials,
        )
    for row, serial in active:
        for column in spec.get("required_columns", []):
            coordinate = f"{column}{row}"
            if not _is_present(sheet[coordinate].value):
                _issue(
                    issues,
                    "sheet.required_value",
                    f"{sheet.title}!{coordinate}",
                    f"active row {serial} is missing a required field",
                )
    if expected_count == 0:
        for row in range(start, end + 1):
            for column in spec.get("required_columns", []):
                coordinate = f"{column}{row}"
                if _is_present(sheet[coordinate].value):
                    _issue(
                        issues,
                        "sheet.placeholder_body",
                        f"{sheet.title}!{coordinate}",
                        "zero-total canonical schedule contains a populated placeholder row",
                    )
    return [row for row, _serial in active]


def _validate_formula_roles(
    issues: list[ValidationIssue], sheet: Worksheet, roles: Mapping[str, str], *, code: str
) -> None:
    for coordinate, expected in roles.items():
        actual = _formula(sheet[coordinate])
        normalized_expected = re.sub(r"\s+|\$", "", str(expected).upper())
        if actual != normalized_expected:
            _issue(
                issues,
                code,
                f"{sheet.title}!{coordinate}",
                "canonical formula role differs",
                expected=normalized_expected,
                actual=actual,
            )


def _validate_semantic_totals(
    issues: list[ValidationIssue],
    formula_sheet: Worksheet,
    value_sheet: Worksheet | None,
    totals: Sequence[Mapping[str, Any]],
) -> None:
    for total in totals:
        coordinate = str(total["cell"])
        sum_range = total.get("sum_range")
        if sum_range is not None:
            expected_formula = f"=SUM({sum_range})"
            actual_formula = _formula(formula_sheet[coordinate])
            if actual_formula != expected_formula:
                _issue(
                    issues,
                    "sheet.total_formula",
                    f"{formula_sheet.title}!{coordinate}",
                    "canonical total formula differs",
                    expected=expected_formula,
                    actual=actual_formula,
                )
        if value_sheet is not None:
            actual_value = _decimal(value_sheet[coordinate].value)
            if actual_value is None or not _equal_scalar(total["expected"], actual_value):
                _issue(
                    issues,
                    "sheet.total_value",
                    f"{formula_sheet.title}!{coordinate}",
                    "canonical aggregate differs",
                    expected=total["expected"],
                    actual=actual_value,
                )


def _validate_payslips(
    issues: list[ValidationIssue],
    formula_sheet: Worksheet,
    value_sheet: Worksheet | None,
    spec: Mapping[str, Any],
) -> None:
    credited = Decimal(0)
    populated = 0
    for ordinal, start_value in enumerate(spec["block_starts"], start=1):
        start = int(start_value)
        required = (f"E{start}", f"K{start}")
        if all(_is_present(formula_sheet[cell].value) for cell in required):
            populated += 1
        else:
            _issue(
                issues,
                "payslip.required_value",
                f"PaySlip!A{start}:X{start + 15}",
                f"payslip block {ordinal} is empty or missing identity fields",
            )
        total_row = start + 13
        detail_start = start + 4
        roles = {
            f"F{total_row}": f"=SUM(F{detail_start}:F{total_row - 1})",
            f"L{total_row}": f"=SUM(L{detail_start}:L{total_row - 1})",
            f"V{total_row}": f"=SUM(V{detail_start}:V{total_row - 1})",
        }
        _validate_formula_roles(issues, formula_sheet, roles, code="payslip.formula_role")
        credit_coordinate = f"F{start + 14}"
        if not _is_present(formula_sheet[credit_coordinate].value):
            _issue(
                issues,
                "payslip.amount_credited",
                f"PaySlip!{credit_coordinate}",
                "amount credited is missing",
            )
        if value_sheet is not None:
            value = _decimal(value_sheet[credit_coordinate].value)
            if value is None:
                _issue(
                    issues,
                    "payslip.amount_credited",
                    f"PaySlip!{credit_coordinate}",
                    "amount credited is not numeric",
                )
            else:
                credited += value
    if populated != len(spec["block_starts"]):
        _issue(
            issues,
            "payslip.body_count",
            "PaySlip",
            "canonical payslip block count differs",
            expected=len(spec["block_starts"]),
            actual=populated,
        )
    if value_sheet is not None and not _equal_scalar(spec["expected_total"], credited):
        _issue(
            issues,
            "payslip.total_value",
            "PaySlip",
            "amounts credited do not reconcile to canonical net payable",
            expected=spec["expected_total"],
            actual=credited,
        )


def _value_at(workbook: Any, reference: str) -> Decimal | None:
    sheet_name, coordinate = reference.rsplit("!", 1)
    if sheet_name not in workbook.sheetnames:
        return None
    return _decimal(workbook[sheet_name][coordinate].value)


def _validate_reconciliations(
    issues: list[ValidationIssue], value_workbook: Any, specs: Sequence[Mapping[str, Any]]
) -> None:
    for spec in specs:
        members = [_value_at(value_workbook, item) for item in spec["members"]]
        target_members = [
            _value_at(value_workbook, item) for item in spec.get("target_members", [])
        ]
        target = _value_at(value_workbook, spec["target"]) if spec.get("target") else None
        expected = _decimal(spec["expected"])
        actual: object
        valid = all(value is not None for value in members + target_members)
        if spec["operation"] == "equal":
            valid = (
                valid and bool(members) and all(_equal_scalar(expected, value) for value in members)
            )
            actual = members
        else:
            member_total = sum((value or Decimal(0) for value in members), Decimal(0))
            target_total = (
                sum((value or Decimal(0) for value in target_members), Decimal(0))
                if target_members
                else target
            )
            valid = (
                valid
                and target_total is not None
                and _equal_scalar(expected, member_total)
                and _equal_scalar(expected, target_total)
            )
            actual = {"schedule_total": member_total, "pay_bill_total": target_total}
        if not valid:
            _issue(
                issues,
                "workbook.reconciliation",
                str(spec["name"]),
                "canonical schedules do not reconcile",
                expected=spec["expected"],
                actual=actual,
            )


def validate_non_pay_bill_semantics(
    formula_workbook: Any,
    contract: Mapping[str, Any],
    *,
    value_workbook: Any | None = None,
    require_june_totals: bool = True,
) -> list[ValidationIssue]:
    """Validate all canonical front sheets and schedules without exposing PII."""
    issues: list[ValidationIssue] = []
    semantics = contract.get("canonical_semantics", {})
    sections = (semantics.get("front_sheets", {}), semantics.get("schedules", {}))
    for specs in sections:
        for sheet_name, spec in specs.items():
            if sheet_name not in formula_workbook.sheetnames:
                _issue(issues, "sheet.missing", sheet_name, "canonical semantic sheet is missing")
                continue
            formula_sheet = formula_workbook[sheet_name]
            value_sheet = (
                value_workbook[sheet_name]
                if value_workbook is not None and sheet_name in value_workbook.sheetnames
                else None
            )
            active_rows: list[int] = []
            if "body" in spec:
                active_rows = _validate_body(issues, formula_sheet, spec["body"])
            if "formula_roles" in spec:
                _validate_formula_roles(
                    issues, formula_sheet, spec["formula_roles"], code="sheet.formula_role"
                )
            if "row_formula" in spec:
                target = str(spec["row_formula"]["target_column"])
                sources = spec["row_formula"]["source_columns"]
                for row in active_rows:
                    expected = f"=SUM({sources[0]}{row}:{sources[-1]}{row})"
                    equivalent = f"=SUM({'+'.join(f'{column}{row}' for column in sources)})"
                    actual = _formula(formula_sheet[f"{target}{row}"])
                    if actual not in {expected, equivalent}:
                        _issue(
                            issues,
                            "sheet.row_formula",
                            f"{sheet_name}!{target}{row}",
                            "row total includes the wrong recovery buckets",
                            expected=expected,
                            actual=actual,
                        )
            for coordinate in spec.get("required_cells", []):
                if not _is_present(formula_sheet[coordinate].value):
                    _issue(
                        issues,
                        "sheet.required_value",
                        f"{sheet_name}!{coordinate}",
                        "canonical form field is empty",
                    )
            destination = spec.get("destination")
            if destination:
                for coordinate in destination["required_cells"]:
                    if not _is_present(formula_sheet[coordinate].value):
                        _issue(
                            issues,
                            "gpf.destination",
                            f"{sheet_name}!{coordinate}",
                            "GPF remittance destination field is empty",
                        )
                actual = _normalized_text(formula_sheet[destination["cell"]].value)
                if destination["jurisdiction"] not in actual:
                    _issue(
                        issues,
                        "gpf.jurisdiction",
                        f"{sheet_name}!{destination['cell']}",
                        "GPF destination does not identify the schedule jurisdiction",
                        expected=destination["jurisdiction"],
                        actual=actual,
                    )
            if sheet_name == "PaySlip":
                _validate_payslips(issues, formula_sheet, value_sheet, spec)
            elif sheet_name == "Pension Sub (2)":
                starts = [int(row) for row in spec["block_starts"]]
                for ordinal, row in enumerate(starts, start=1):
                    for coordinate in (
                        f"C{row}",
                        f"C{row + 1}",
                        f"C{row + 2}",
                        f"D{row}",
                        f"F{row}",
                        f"G{row}",
                        f"H{row}",
                        f"I{row}",
                    ):
                        if not _is_present(formula_sheet[coordinate].value):
                            _issue(
                                issues,
                                "nps.required_value",
                                f"{sheet_name}!{coordinate}",
                                f"NPS block {ordinal} is incomplete",
                            )
            _validate_semantic_totals(issues, formula_sheet, value_sheet, spec.get("totals", []))

    if require_june_totals and value_workbook is None:
        _issue(
            issues,
            "semantics.june_totals",
            "workbook",
            "canonical schedules require recalculated values for June reconciliation",
        )
    elif value_workbook is not None:
        _validate_reconciliations(issues, value_workbook, semantics.get("reconciliations", []))
    return issues


def recalculate_with_libreoffice(workbook_path: Path, soffice: str | None = None) -> Path:
    """Return a recalculated temporary workbook path owned by a persistent temp dir."""
    executable = soffice or shutil.which("soffice") or shutil.which("libreoffice")
    if executable is None:
        raise RuntimeError("LibreOffice (soffice) is required for formula-result validation")
    temp_dir = Path(tempfile.mkdtemp(prefix="accord-canonical-recalc-"))
    input_dir = temp_dir / "input"
    output_dir = temp_dir / "output"
    profile_dir = temp_dir / "profile"
    input_dir.mkdir()
    output_dir.mkdir()
    staged = input_dir / "workbook.xlsx"
    shutil.copyfile(workbook_path, staged)
    try:
        result = subprocess.run(
            [
                executable,
                f"-env:UserInstallation={profile_dir.as_uri()}",
                "--headless",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(output_dir),
                str(staged),
            ],
            check=False,
            capture_output=True,
            text=True,
            timeout=120,
        )
    except (OSError, subprocess.SubprocessError) as error:
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise RuntimeError(f"LibreOffice recalculation failed: {error}") from error
    recalculated = output_dir / staged.name
    if result.returncode != 0 or not recalculated.is_file():
        shutil.rmtree(temp_dir, ignore_errors=True)
        diagnostic = (result.stderr or result.stdout).strip().splitlines()
        summary = diagnostic[-1] if diagnostic else f"exit code {result.returncode}"
        raise RuntimeError(f"LibreOffice recalculation failed: {summary}")
    return recalculated


def _formula_error_cells(workbook_path: Path) -> list[str]:
    """Return formula-cell coordinates with cached OOXML error results."""
    errors: list[str] = []
    with zipfile.ZipFile(workbook_path) as archive:
        _workbook, sheets, _defined_names = _EXTRACTOR._workbook_parts(archive)
        for sheet in sheets:
            sheet_name = sheet["name"]
            xml_name = sheet["path"]
            if xml_name not in archive.namelist():
                continue
            root = ET.fromstring(archive.read(xml_name))
            for cell in root.iter(f"{{{_MAIN_NS}}}c"):
                formula = cell.find(f"{{{_MAIN_NS}}}f")
                value = cell.find(f"{{{_MAIN_NS}}}v")
                cached = "" if value is None else str(value.text or "").upper()
                if formula is not None and (cell.attrib.get("t") == "e" or cached in _ERROR_VALUES):
                    errors.append(f"{sheet_name}!{cell.attrib.get('r', '?')}")
    return errors


def validate_workbook(
    workbook_path: Path,
    *,
    contract_path: Path = DEFAULT_CONTRACT_PATH,
    recalculate: bool = True,
    soffice: str | None = None,
) -> list[ValidationIssue]:
    """Run the complete canonical export acceptance suite."""
    contract = load_contract(contract_path)
    issues = validate_structure(workbook_path, contract)
    formula_workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    if "Pay Bill" not in formula_workbook.sheetnames:
        return issues

    recalculated_path: Path | None = None
    value_sheet: Worksheet | None = None
    value_workbook: Any | None = None
    if recalculate:
        try:
            recalculated_path = recalculate_with_libreoffice(workbook_path, soffice)
            formula_errors = _formula_error_cells(recalculated_path)
            for coordinate in formula_errors:
                _issue(
                    issues,
                    "formula.cached_error",
                    coordinate,
                    "formula still returns an error after LibreOffice recalculation",
                )
            value_workbook = load_workbook(recalculated_path, data_only=True, read_only=False)
            value_sheet = value_workbook["Pay Bill"]
        except (OSError, RuntimeError, subprocess.SubprocessError, zipfile.BadZipFile) as error:
            _issue(issues, "recalculation.failed", "workbook", str(error))
    issues.extend(
        validate_pay_bill_semantics(
            formula_workbook["Pay Bill"],
            contract,
            value_sheet=value_sheet,
            require_june_totals=recalculate,
        )
    )
    issues.extend(
        validate_non_pay_bill_semantics(
            formula_workbook,
            contract,
            value_workbook=value_workbook,
            require_june_totals=recalculate,
        )
    )
    if recalculated_path is not None:
        shutil.rmtree(recalculated_path.parents[1], ignore_errors=True)
    return issues


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Validate an Accord v3 XLSX against the canonical June export contract."
    )
    parser.add_argument("xlsx", type=Path, help="generated v3 .xlsx workbook")
    parser.add_argument("--contract", type=Path, default=DEFAULT_CONTRACT_PATH)
    parser.add_argument("--soffice", help="path to the LibreOffice soffice executable")
    parser.add_argument(
        "--max-issues",
        type=int,
        default=200,
        help="maximum detailed differences to print (default: 200)",
    )
    parser.add_argument(
        "--no-recalculate",
        action="store_true",
        help="skip cached-result and June-total checks (never use for release acceptance)",
    )
    return parser


def main(argv: Iterable[str] | None = None) -> int:
    args = _build_parser().parse_args(argv)
    if args.max_issues < 1:
        print("FAIL: --max-issues must be at least 1")
        return 2
    if args.xlsx.suffix.casefold() != ".xlsx" or not args.xlsx.is_file():
        print(f"FAIL: workbook does not exist or is not .xlsx: {args.xlsx}")
        return 2
    try:
        issues = validate_workbook(
            args.xlsx,
            contract_path=args.contract,
            recalculate=not args.no_recalculate,
            soffice=args.soffice,
        )
    except (json.JSONDecodeError, OSError, ValueError, zipfile.BadZipFile) as error:
        print(f"FAIL: could not validate workbook: {error}")
        return 2
    if issues:
        print(f"FAIL: {len(issues)} canonical export difference(s) found")
        for issue in issues[: args.max_issues]:
            print(f"- {issue.render()}")
        if len(issues) > args.max_issues:
            print(f"- ... {len(issues) - args.max_issues} more difference(s) not shown")
        return 1
    if args.no_recalculate:
        print("PASS (STRUCTURE ONLY): formula results and June totals were not accepted")
    else:
        print("PASS: workbook matches the canonical export contract")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
